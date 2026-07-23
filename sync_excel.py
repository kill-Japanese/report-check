# -*- coding: utf-8 -*-
"""
Excel 同步模块 - 以 GitHub Excel 为唯一数据源
功能：
  1. 将 Web 操作（新增/删除/归档/编辑）写回原始 Excel
  2. 提交并推送到 GitHub
  3. 重新生成 HTML 报表
  4. 启动时从 GitHub 拉取最新数据
"""

import os
import sys
import json
import subprocess
import threading
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import auth

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
COLLAB_FILE = os.path.join(DATA_DIR, '协作数据.json')
EXCEL_FILE = os.path.join(BASE_DIR, '超声波户表脚本.xlsx')
HTML_FILE = os.path.join(BASE_DIR, '项目延期点检表.html')

# Excel 列配置（0-based 索引）
# 【严重修复】原使用U列(20)/V列(21)会覆盖延期计算公式，
# 导致Excel延期工时计算错误。改为使用完全空闲的A列(0)和B列(1)。
# 注意：openpyxl写入时使用 1-based 列号，所以 COL_ARCHIVED+1 = 第2列(B列)
COL_ARCHIVED = 0   # 第1列(A列)用于存放归档标志（原U列/20，改为A列/0，避免覆盖公式）
COL_DELETED = 1    # 第2列(B列)用于存放删除标志（原V列/21，改为B列/1，软删除避免合并单元格破坏）
COL_APPROVAL_STATUS = 23    # 第24列(X列) 审批状态: PENDING_ARCHIVE / PENDING_UNARCHIVE / PENDING_EDIT / 空
COL_APPROVAL_SUBMITTER = 24 # 第25列(Y列) 申请人|申请时间: 格式 "username|2026-07-23T10:00:00"
COL_APPROVAL_DETAIL = 25    # 第26列(Z列) 变更内容(JSON): 编辑时使用 {"负责人":"张三→李四","工时":"5→8"}
COL_APPROVAL_TYPE = 26      # 第27列(AA列) 操作类型: archive / unarchive / edit

# 操作记录Sheet配置
OPERATIONS_SHEET = '操作记录'
# 操作记录表头（A-L列，0-based索引对应列号-1）
OPERATION_HEADERS = [
    '操作ID',      # A列
    '操作时间',     # B列
    '操作人',       # C列
    '操作类型',     # D列: archive/unarchive/edit/delete/batch_archive/batch_unarchive/approval_submit/approval_approve/approval_reject
    '项目ID列表',   # E列: JSON数组
    '项目名列表',   # F列: JSON数组
    '变更前内容',   # G列: JSON对象（edit操作）
    '变更后内容',   # H列: JSON对象（edit操作/审批意见）
    '状态',         # I列: direct/pending/approved/rejected
    '关联审批ID',   # J列
    '审批人',       # K列
    '审批时间',     # L列
]

# ==================== 项目数据缓存（解决502超时关键）====================
# 【关键修复】每次请求都用pandas读Excel很慢（3-10秒），加缓存避免重复读取
# 缓存失效条件：Excel文件修改时间变化 或 显式调用 invalidate_projects_cache()
_projects_cache = None
_projects_cache_mtime = None
_cache_lock = threading.Lock()


def invalidate_projects_cache():
    """清除项目数据缓存（写入Excel后必须调用）"""
    global _projects_cache, _projects_cache_mtime
    with _cache_lock:
        _projects_cache = None
        _projects_cache_mtime = None
        print('[缓存] 已清除项目数据缓存')

# ==================== Git 仓库保障 ====================

def ensure_git_repo() -> tuple[bool, str]:
    """确保 Git 仓库存在并且配置正确（Render 部署环境保障）
    
    【关键修复】Render 部署时可能不会保留 .git 目录，
    导致所有 Git 操作失败，数据无法持久化。
    
    修复策略：
    1. 检查 .git 目录是否存在
    2. 如果不存在，尝试从环境变量或已知配置初始化 Git 仓库
    3. 检查 remote origin 是否配置正确
    4. 确保用户信息配置正确（用于 commit）
    """
    git_dir = os.path.join(BASE_DIR, '.git')
    
    try:
        # 1. 检查 .git 目录是否存在
        if not os.path.exists(git_dir):
            print(f"[Git] .git 目录不存在，正在初始化...")
            
            # 初始化 Git 仓库
            init = subprocess.run(
                ['git', 'init'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=10
            )
            if init.returncode != 0:
                return False, f'Git 初始化失败: {init.stderr[:200]}'
            
            # 尝试从环境变量获取 remote URL
            remote_url = os.environ.get('GIT_REMOTE_URL', '')
            if not remote_url:
                # 尝试从常见的 Render 环境变量中推断
                # Render 会设置一些环境变量，但不包含完整的 repo URL
                # 这里使用一个默认的占位符，用户需要在环境变量中配置
                print(f"[Git] 警告: 未设置 GIT_REMOTE_URL 环境变量")
            
            if remote_url:
                subprocess.run(
                    ['git', 'remote', 'add', 'origin', remote_url],
                    capture_output=True, cwd=BASE_DIR, timeout=10
                )
                print(f"[Git] 已设置 remote origin: {remote_url[:50]}...")
        else:
            # .git 存在，检查 remote 配置
            remote_check = subprocess.run(
                ['git', 'remote', '-v'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=5
            )
            if 'origin' not in remote_check.stdout:
                print(f"[Git] 警告: 未配置 remote origin")
        
        # 2. 确保 Git 用户信息配置正确（用于 commit）
        subprocess.run(
            ['git', 'config', 'user.email', 'server@report-check.local'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        subprocess.run(
            ['git', 'config', 'user.name', 'Report Check Server'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        
        # 3. 确保分支名为 main
        branch = subprocess.run(
            ['git', 'branch', '--show-current'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=5
        )
        current_branch = branch.stdout.strip()
        if current_branch and current_branch != 'main':
            subprocess.run(
                ['git', 'branch', '-M', 'main'],
                capture_output=True, cwd=BASE_DIR, timeout=5
            )
            print(f"[Git] 已将分支 {current_branch} 重命名为 main")
        
        return True, 'Git 仓库就绪'
        
    except Exception as e:
        return False, f'Git 仓库初始化失败: {str(e)}'


# ==================== Git 操作 ====================

def git_pull() -> tuple[bool, str]:
    """从 GitHub 拉取最新数据（双路径：git命令优先，GitHub API兜底）
    
    【关键修复】Render 环境中 git 命令可能不存在，
    自动检测并回退到 GitHub REST API 模式。
    """
    # 检查 git 命令是否可用
    git_available = False
    try:
        result = subprocess.run(['git', '--version'], capture_output=True, timeout=5)
        git_available = (result.returncode == 0) and os.path.exists(os.path.join(BASE_DIR, '.git'))
    except (FileNotFoundError, OSError):
        git_available = False
    
    if not git_available:
        # 回退到 GitHub API 模式
        try:
            from github_sync import github_api_pull
            print('[sync] git 不可用，使用 GitHub API 模式拉取')
            return github_api_pull()
        except Exception as api_e:
            return False, f'git不可用且API拉取失败: {api_e}'
    
    # git 可用，使用原有逻辑
    try:
        ensure_ok, ensure_msg = ensure_git_repo()
        if not ensure_ok:
            return False, f'Git仓库不可用: {ensure_msg}'
        if not os.path.exists(os.path.join(BASE_DIR, '.git')):
            return False, '未检测到 Git 仓库'

        fetch = subprocess.run(
            ['git', 'fetch', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        if fetch.returncode != 0:
            return False, f'fetch失败: {fetch.stderr[:200]}'

        # 【关键修复】先检查本地是否有未提交的变更
        # 如果有，说明本地数据比远程新，不能被远程覆盖！
        # 这是用户/项目消失的核心原因：本地未推送的变更被远程旧文件覆盖
        status_check = subprocess.run(
            ['git', 'status', '--porcelain', '用户管理.xlsx', '超声波户表脚本.xlsx', 'data/'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        has_local_changes = bool(status_check.stdout.strip())
        
        # 检查本地是否比远程新（有未推送的commit）
        ahead_check = subprocess.run(
            ['git', 'rev-list', '--count', 'origin/main..HEAD'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        try:
            ahead_count = int(ahead_check.stdout.strip())
        except ValueError:
            ahead_count = 0
        
        if has_local_changes or ahead_count > 0:
            # 本地有未推送的变更，先尝试推送再拉取
            print(f'[sync] 检测到本地有未推送变更（工作区:{has_local_changes}, 提交:{ahead_count}），先尝试推送...')
            push_ok, push_msg = git_push('启动时同步本地未推送数据')
            if not push_ok:
                print(f'[sync] 警告: 本地变更推送失败（{push_msg[:60]}），将保留本地文件不被覆盖')
                # 【关键修复】git pull 前完整备份Excel，防止本地新增行被覆盖
                excel_backup = _backup_excel_file()
                archive_flags = _backup_archive_deleted_flags()
                # 推送失败，只拉取不覆盖关键文件
                pull = subprocess.run(
                    ['git', 'pull', 'origin', 'main', '--no-edit', '--no-commit'],
                    capture_output=True, text=True, cwd=BASE_DIR, timeout=30
                )
                # 如果有冲突，取消合并（保留本地版本）
                if pull.returncode != 0:
                    subprocess.run(['git', 'merge', '--abort'], capture_output=True, cwd=BASE_DIR, timeout=10)
                # 【关键修复】git pull 后恢复Excel和归档标志
                _restore_excel_file(excel_backup)
                _restore_archive_deleted_flags(archive_flags)
                return True, '（拉取成功；本地有未推送变更已保留）'
            print(f'[sync] 本地变更已推送: {push_msg[:60]}')

        critical_files = ['用户管理.xlsx', '超声波户表脚本.xlsx']
        restored = []
        for f in critical_files:
            fpath = os.path.join(BASE_DIR, f)
            existed_before = os.path.exists(fpath)
            # 【关键修复】只在本地文件不存在时才从远程checkout
            # 本地文件已存在的情况下绝不覆盖！
            if not existed_before:
                checkout = subprocess.run(
                    ['git', 'checkout', 'origin/main', '--', f],
                    capture_output=True, text=True, cwd=BASE_DIR, timeout=10
                )
                if checkout.returncode == 0 and os.path.exists(fpath):
                    restored.append(f'{f}(新建)')

        # 【关键修复1】git pull 前完整备份Excel，防止本地新增行被远程覆盖
        excel_backup = _backup_excel_file()
        
        # 【关键修复2】git pull 前备份归档/删除标志，防止被远程旧版本覆盖
        archive_flags = _backup_archive_deleted_flags()
        
        pull = subprocess.run(
            ['git', 'pull', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        
        # 【关键修复3】git pull 后恢复Excel（防止本地新增行被覆盖）
        _restore_excel_file(excel_backup)
        
        # 【关键修复4】git pull 后恢复归档/删除标志
        _restore_archive_deleted_flags(archive_flags)

        msg_parts = ['拉取成功']
        if restored:
            msg_parts.append(f'已同步 {len(restored)} 个文件: {", ".join(restored)}')
        return True, '（' + '；'.join(msg_parts) + '）'
    except subprocess.TimeoutExpired:
        return False, '拉取超时'
    except Exception as e:
        return False, f'拉取失败: {str(e)}'

def _backup_archive_deleted_flags() -> dict:
    """备份 Excel 中的归档/删除标志（防止 git pull 覆盖）
    
    返回: {row_num: {'archived': bool, 'deleted': bool}}
    """
    from openpyxl import load_workbook
    flags = {}
    try:
        if not os.path.exists(EXCEL_FILE):
            return flags
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['任务计划表']
        for row_num in range(4, ws.max_row + 1):
            archived_val = ws.cell(row=row_num, column=COL_ARCHIVED + 1).value
            deleted_val = ws.cell(row=row_num, column=COL_DELETED + 1).value
            archived = str(archived_val).strip() in ('已归档', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y') if archived_val else False
            deleted = str(deleted_val).strip() in ('已删除', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y') if deleted_val else False
            if archived or deleted:
                flags[row_num] = {'archived': archived, 'deleted': deleted}
        wb.close()
        print(f'[sync] 已备份 {len(flags)} 个行的归档/删除标志')
    except Exception as e:
        print(f'[sync] 备份归档标志失败: {e}')
    return flags


def _restore_archive_deleted_flags(flags: dict):
    """恢复归档/删除标志到 Excel（git pull 后调用）"""
    if not flags:
        return
    from openpyxl import load_workbook
    try:
        if not os.path.exists(EXCEL_FILE):
            return
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        restored = 0
        for row_num, state in flags.items():
            if row_num > ws.max_row:
                continue
            if state['archived']:
                current = ws.cell(row=row_num, column=COL_ARCHIVED + 1).value
                if not current or str(current).strip() not in ('已归档', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y'):
                    ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='已归档')
                    restored += 1
            if state['deleted']:
                current = ws.cell(row=row_num, column=COL_DELETED + 1).value
                if not current or str(current).strip() not in ('已删除', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y'):
                    ws.cell(row=row_num, column=COL_DELETED + 1, value='已删除')
                    restored += 1
        if restored > 0:
            wb.save(EXCEL_FILE)
            invalidate_projects_cache()
            print(f'[sync] 已恢复 {restored} 个行的归档/删除标志')
        wb.close()
    except Exception as e:
        print(f'[sync] 恢复归档标志失败: {e}')


def _backup_excel_file() -> str:
    """完整备份 Excel 文件（git pull 前调用，防止新增行被远程覆盖）
    
    Returns:
        str: 备份文件路径，失败返回空字符串
    """
    import shutil
    try:
        if not os.path.exists(EXCEL_FILE):
            return ''
        backup_path = EXCEL_FILE + '.bak_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        shutil.copy2(EXCEL_FILE, backup_path)
        # 获取原始文件行数用于对比
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb['任务计划表']
        row_count = ws.max_row
        wb.close()
        print(f'[sync] 已完整备份Excel: {os.path.basename(backup_path)} ({row_count}行)')
        return backup_path
    except Exception as e:
        print(f'[sync] Excel备份失败: {e}')
        return ''


def _restore_excel_file(backup_path: str) -> bool:
    """从备份恢复 Excel 文件（git pull 后调用，确保本地变更不丢失）
    
    恢复策略：
    - 备份文件行数 > 当前文件行数 → 恢复（说明本地有新增行被远程覆盖了）
    - 备份文件行数 <= 当前文件行数 → 不恢复（远程可能有编辑或新增）
    
    Args:
        backup_path: 备份文件路径
    
    Returns:
        bool: 是否执行了恢复
    """
    import shutil
    if not backup_path or not os.path.exists(backup_path):
        return False
    try:
        from openpyxl import load_workbook
        
        # 获取备份文件行数
        wb_bak = load_workbook(backup_path, read_only=True)
        ws_bak = wb_bak['任务计划表']
        bak_rows = ws_bak.max_row
        wb_bak.close()
        
        # 获取当前文件行数
        if not os.path.exists(EXCEL_FILE):
            shutil.copy2(backup_path, EXCEL_FILE)
            print(f'[sync] Excel不存在，已从备份恢复 ({bak_rows}行)')
            return True
            
        wb_cur = load_workbook(EXCEL_FILE, read_only=True)
        ws_cur = wb_cur['任务计划表']
        cur_rows = ws_cur.max_row
        wb_cur.close()
        
        # 关键：备份行数更多 → 说明本地有新增行被远程覆盖了，必须恢复
        if bak_rows > cur_rows:
            shutil.copy2(backup_path, EXCEL_FILE)
            invalidate_projects_cache()
            print(f'[sync] 检测到远程覆盖: 备份{bak_rows}行 > 当前{cur_rows}行，已从备份恢复Excel')
            # 清理备份文件
            try:
                os.remove(backup_path)
            except:
                pass
            return True
        else:
            print(f'[sync] 无需恢复: 备份{bak_rows}行 <= 当前{cur_rows}行，保留远程版本')
            # 清理备份文件
            try:
                os.remove(backup_path)
            except:
                pass
            return False
    except Exception as e:
        print(f'[sync] Excel恢复失败: {e}')
        # 出错时保守处理：尝试恢复（以防数据丢失）
        try:
            if os.path.exists(backup_path):
                shutil.copy2(backup_path, EXCEL_FILE)
                invalidate_projects_cache()
                print(f'[sync] 异常恢复：已从备份还原Excel')
                return True
        except:
            pass
        return False


# Git 可用性缓存（避免每次都检查 git --version）
_git_available_cache = None


def git_push(message: str = '同步数据') -> tuple[bool, str]:
    """将变更提交并推送到 GitHub（极速优化版）
    
    【极速优化】99% 的场景下跳过不必要的网络请求：
    1. 不做 git fetch（省 3-5 秒网络请求）
    2. 不检查 ahead/behind（省 2 次 subprocess）
    3. 直接 git add + commit + push
    4. 只有 push 失败时才回退到完整流程（fetch+pull+重试）
    
    预期：从 10-15 秒降到 2-4 秒
    """
    global _git_available_cache
    
    # 检查 git 命令是否可用（带缓存）
    if _git_available_cache is None:
        try:
            result = subprocess.run(['git', '--version'], capture_output=True, timeout=5)
            _git_available_cache = (result.returncode == 0) and os.path.exists(os.path.join(BASE_DIR, '.git'))
        except (FileNotFoundError, OSError):
            _git_available_cache = False
    
    if not _git_available_cache:
        # 回退到 GitHub API 模式
        try:
            from github_sync import github_api_push
            print('[sync] git 不可用，使用 GitHub API 模式推送')
            return github_api_push(message)
        except Exception as api_e:
            return False, f'git不可用且API推送失败: {api_e}'
    
    # ============== 极速路径：直接 add + commit + push ==============
    try:
        # 1. 检查是否有变更（只检查关键文件）
        result = subprocess.run(
            ['git', 'status', '--porcelain',
             '超声波户表脚本.xlsx', '用户管理.xlsx',
             '项目延期点检表.html', 'data/'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=5
        )
        if not result.stdout.strip():
            return True, '无变更，无需推送'
        
        # 2. 只 add 关键文件（不用 -A，更快）
        subprocess.run(
            ['git', 'add', '超声波户表脚本.xlsx', '用户管理.xlsx',
             '项目延期点检表.html', 'data/'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        
        # 3. commit
        commit_msg = f'[数据同步] {message} - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        commit_result = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if commit_result.returncode != 0:
            if 'nothing to commit' in (commit_result.stdout + commit_result.stderr):
                return True, '无变更'
            return False, f'提交失败: {commit_result.stderr[:200]}'
        
        # 4. push（极速路径，不做 fetch）
        push_result = subprocess.run(
            ['git', 'push', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=15
        )
        if push_result.returncode == 0:
            return True, '已同步到 GitHub（极速路径）'
        
        # ============== push 失败：回退到完整流程（fetch+pull+重试）==============
        print(f'[sync] 极速push失败（{push_result.stderr[:80]}），回退到完整流程...')
        
        # 撤销刚才的 commit（避免重复提交）
        subprocess.run(['git', 'reset', '--soft', 'HEAD~1'], capture_output=True, cwd=BASE_DIR, timeout=5)
        
        # 【关键修复1】git pull 前完整备份Excel，防止新增行被远程覆盖
        excel_backup = _backup_excel_file()
        
        # 【关键修复2】git pull 前备份归档/删除标志，防止被远程旧版本覆盖
        archive_flags = _backup_archive_deleted_flags()
        
        # fetch
        fetch = subprocess.run(
            ['git', 'fetch', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        
        # pull
        pull = subprocess.run(
            ['git', 'pull', 'origin', 'main', '--no-edit'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        
        # 【关键修复3】git pull 后恢复Excel（防止新增行被覆盖）
        _restore_excel_file(excel_backup)
        
        # 【关键修复4】git pull 后恢复归档/删除标志
        _restore_archive_deleted_flags(archive_flags)
        if pull.returncode != 0:
            # 有冲突，回退到 GitHub API
            print(f'[sync] pull失败，回退到GitHub API: {pull.stderr[:80]}')
            try:
                from github_sync import github_api_push
                return github_api_push(message)
            except Exception as api_e:
                return False, f'同步失败: {pull.stderr[:200]}'
        
        # 重新 commit + push
        subprocess.run(
            ['git', 'add', '超声波户表脚本.xlsx', '用户管理.xlsx',
             '项目延期点检表.html', 'data/'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        commit_result2 = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if commit_result2.returncode != 0:
            if 'nothing to commit' in (commit_result2.stdout + commit_result2.stderr):
                return True, '无变更（同步完成）'
            return False, f'提交失败: {commit_result2.stderr[:200]}'
        
        push_result2 = subprocess.run(
            ['git', 'push', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        if push_result2.returncode == 0:
            return True, '已同步到 GitHub'
        
        # git push 还是失败，回退到 GitHub API
        subprocess.run(['git', 'reset', '--soft', 'HEAD~1'], capture_output=True, cwd=BASE_DIR, timeout=5)
        try:
            from github_sync import github_api_push
            print('[sync] git push 失败，回退到 GitHub API 模式')
            return github_api_push(message)
        except Exception as api_e:
            return False, f'推送失败（git+API均失败）: {push_result2.stderr[:200]}'
        
    except subprocess.TimeoutExpired:
        return False, '同步超时'
    except Exception as e:
        return False, f'同步失败: {str(e)}'

# ==================== 协作数据读写 ====================

def load_collab_data() -> dict:
    """加载协作数据（新增项目、归档、编辑等）"""
    default = {
        'localEdits': {}, 'notes': {}, 'checked': {},
        'archived': {}, 'customEmails': {}, 'newProjects': [],
        'deletedIds': [],
        'lastUpdate': datetime.now().isoformat()
    }
    if os.path.exists(COLLAB_FILE):
        try:
            with open(COLLAB_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # 确保所有默认键都存在
            for k, v in default.items():
                if k not in data:
                    data[k] = v
            return data
        except:
            pass
    return default

def save_collab_data(data: dict):
    """保存协作数据"""
    data['lastUpdate'] = datetime.now().isoformat()
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(COLLAB_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ==================== Excel 读写（核心） ====================

def _fmt_date(d, is_start=True):
    """日期格式化：将 datetime/Timestamp 转为字符串（与更新点检表.py一致）"""
    if d is None or (isinstance(d, float) and pd.isna(d)):
        return ''
    if d == '/' or str(d).strip() == '/':
        return '1900-01-01' if is_start else '2100-01-01'
    if isinstance(d, pd.Timestamp):
        return d.strftime('%Y-%m-%d')
    if isinstance(d, datetime):
        return d.strftime('%Y-%m-%d')
    return str(d)

def _cell_val(v):
    """安全获取单元格值，处理None/NaN"""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except:
        pass
    return v


def _not_empty(v) -> bool:
    """判断单元格值是否非空"""
    if v is None:
        return False
    try:
        if pd.isna(v):
            return False
    except:
        pass
    s = str(v).strip()
    return s != '' and s != 'nan'


def read_excel_projects() -> list:
    """从原始 Excel 读取所有项目资源（与更新点检表.py 逻辑一致）
    
    【502关键修复1】用 openpyxl 替代 pandas 读取：
      - pandas 导入需 2-5 秒，读 Excel 需 3-10 秒
      - openpyxl 导入几乎瞬间，读 Excel 只需 0.5-2 秒
    
    【502关键修复2】添加文件级缓存：Excel文件修改时间不变则直接返回缓存，
      避免每次请求都重读 Excel（0.5-2秒 → 几毫秒）
    """
    global _projects_cache, _projects_cache_mtime
    
    if not os.path.exists(EXCEL_FILE):
        return []
    
    # 检查缓存是否有效
    try:
        current_mtime = os.path.getmtime(EXCEL_FILE)
        with _cache_lock:
            if (_projects_cache is not None and 
                _projects_cache_mtime is not None and 
                _projects_cache_mtime == current_mtime):
                return _projects_cache
    except:
        pass
    
    # ============== 缓存失效，用 openpyxl 重新读取 ==============
    from openpyxl import load_workbook as _load_wb
    wb = _load_wb(EXCEL_FILE, data_only=True, read_only=True)
    ws = wb['任务计划表']
    
    projects = []
    current_dept = None
    current_project = None
    current_start = None
    current_end = None
    current_desc = None
    
    # 从第4行开始（Excel 1-based，前3行是表头）
    # idx 保持与 pandas 版本一致：df从0开始，第4行对应idx=3
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=3):
        # row 是 tuple，索引 0=A列, 1=B列, ... 4=E列, 5=F列, ...
        # 注意：openpyxl 的 row 可能比列数短，需要用安全访问
        
        def _get(col):
            """安全获取列值，避免索引越界"""
            if col < len(row):
                return row[col]
            return None
        
        # E列(4)：部门
        if _not_empty(_get(4)):
            val = _get(4)
            current_dept = str(val) if not isinstance(val, float) else val
        
        # F列(5)：项目名
        if _not_empty(_get(5)):
            current_project = str(_get(5))
            current_start = _cell_val(_get(6))   # G列：项目开始
            current_end = _cell_val(_get(7))     # H列：项目结束
            current_desc = str(_get(8)) if _not_empty(_get(8)) else ''  # I列：项目描述
        else:
            # 合并单元格的后续行，检查是否有独立值
            if _not_empty(_get(8)):
                current_desc = str(_get(8))
            if _not_empty(_get(6)):
                current_start = _cell_val(_get(6))
            if _not_empty(_get(7)):
                current_end = _cell_val(_get(7))
        
        # J列(9)：资源类型，K列(10)：资源名称
        resource_type = str(_get(9)) if _not_empty(_get(9)) else ''
        resource_name = str(_get(10)) if _not_empty(_get(10)) else ''
        
        # 清理资源名称
        if resource_name:
            resource_name = resource_name.strip()
            if resource_name.startswith('@'):
                resource_name = resource_name[1:]
            elif resource_name.startswith(' @'):
                resource_name = resource_name[2:]
            resource_name = resource_name.strip()
            if '(' in resource_name:
                resource_name = resource_name[:resource_name.index('(')].strip()
        
        has_resource = resource_type.strip() or resource_name.strip()
        
        # A列(0)：归档标志
        archived_flag = ''
        if _not_empty(_get(COL_ARCHIVED)):
            archived_flag = str(_get(COL_ARCHIVED)).strip()
        is_archived = archived_flag in ('已归档', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y')
        
        # B列(1)：删除标志（软删除）
        deleted_flag = ''
        if _not_empty(_get(COL_DELETED)):
            deleted_flag = str(_get(COL_DELETED)).strip()
        is_deleted = deleted_flag in ('已删除', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y')
        
        if is_deleted:
            continue
        
        if current_project and has_resource:
            projects.append({
                'id': row_idx,
                '部门': current_dept if current_dept and not (isinstance(current_dept, float) and (pd.isna(current_dept) if hasattr(pd, 'isna') else False)) else '',
                '项目': current_project,
                '项目开始时间': _fmt_date(current_start, True),
                '项目结束时间': _fmt_date(current_end, False),
                '项目描述': current_desc,
                '资源类型': resource_type,
                '资源名称': resource_name,
                '资源开始时间': _fmt_date(_cell_val(_get(11)), True),
                '资源结束时间': _fmt_date(_cell_val(_get(12)), False),
                '日平均工时': _get(13) if _not_empty(_get(13)) else 0,
                '已归档': is_archived,
            })
    
    wb.close()
    
    # 写入缓存
    try:
        with _cache_lock:
            _projects_cache = projects
            _projects_cache_mtime = os.path.getmtime(EXCEL_FILE)
            print(f'[缓存] 项目数据已缓存（{len(projects)}条，openpyxl）')
    except:
        pass
    
    return projects

def _find_merged_range(ws, row: int, col: int):
    """查找单元格所属的合并区域，返回 (min_row, max_row, min_col, max_col) 或 None"""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return None
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None


def _safe_write_cell(ws, row: int, col: int, value) -> bool:
    """
    安全地写入Excel单元格，自动处理合并单元格：
    - 如果单元格不是合并单元格，直接写入
    - 如果单元格是合并区域的左上角，直接写入
    - 如果单元格在合并区域内但不是左上角：
      1. 先取消合并
      2. 将原值填充到原合并区域的所有单元格
      3. 修改目标单元格的值
    返回: 是否成功写入
    """
    from openpyxl.cell.cell import MergedCell
    
    merged_range = _find_merged_range(ws, row, col)
    
    if merged_range is None:
        # 不是合并单元格，直接写入
        ws.cell(row=row, column=col, value=value)
        return True
    
    # 是合并单元格
    min_r, max_r = merged_range.min_row, merged_range.max_row
    min_c, max_c = merged_range.min_col, merged_range.max_col
    
    if row == min_r and col == min_c:
        # 恰好是合并区域的左上角，直接写入
        ws.cell(row=row, column=col, value=value)
        return True
    
    # 在合并区域内但不是左上角：需要先取消合并
    # 1. 先获取合并区域的原值
    original_value = ws.cell(row=min_r, column=min_c).value
    
    # 2. 取消合并
    ws.unmerge_cells(str(merged_range))
    
    # 3. 将原值填充到原合并区域的所有单元格
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(row=r, column=c, value=original_value)
    
    # 4. 写入目标单元格的新值
    ws.cell(row=row, column=col, value=value)
    return True


def apply_collab_to_excel() -> tuple[bool, str, int]:
    """
    将协作数据（新增/删除/归档/编辑）应用到原始 Excel
    返回: (成功, 消息, 变更数量)
    
    执行顺序（关键！避免行号漂移）：
    1. 先处理归档（不改变行号，使用初始映射）
    2. 再处理编辑（不改变行号，使用初始映射）
    3. 再处理删除（从大到小删除，使用初始映射）
    4. 最后处理新增（追加到末尾，不影响已有行号）
    """
    collab = load_collab_data()
    changes = 0
    errors = []
    
    if not os.path.exists(EXCEL_FILE):
        return False, '原始 Excel 文件不存在', 0
    
    # 读取所有现有项目（仅一次，作为行号映射的基准）
    existing_projects = read_excel_projects()
    # id是df的idx，Excel行号=idx+1（因为df从0开始，Excel从1开始，且前3行是表头）
    # 实际上：df.idx=3 对应 Excel第4行，所以 Excel行号 = idx + 1
    id_to_excel_row = {p['id']: p['id'] + 1 for p in existing_projects}
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        
        # ============== 1. 先处理归档（不改变行号） ==============
        archived = collab.get('archived', {})
        
        # 构建需要归档的行号集合
        archived_rows = set()
        if archived:
            for pid, arch_info in archived.items():
                try:
                    pid_int = int(pid)
                except:
                    pid_int = pid
                row_num = id_to_excel_row.get(pid_int) or id_to_excel_row.get(str(pid_int))
                if row_num and arch_info:
                    archived_rows.add(row_num)
        
        # 【修复】只处理需要变更的行，不盲目清空所有行的U列（避免覆盖原有公式）
        # - 需要归档的行：写入"已归档"
        # - 不需要归档但当前U列是"已归档"的行：清空（恢复）
        # - 其他行（原有公式）：不做任何修改
        for p in existing_projects:
            row_num = id_to_excel_row.get(p['id'])
            if not row_num:
                continue
            
            current_val = ws.cell(row=row_num, column=COL_ARCHIVED + 1).value
            is_current_archived = (str(current_val).strip() == '已归档') if current_val else False
            
            if row_num in archived_rows:
                # 需要归档
                if not is_current_archived:
                    ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='已归档')
                    changes += 1
                    print(f"   📦 归档: Excel第{row_num}行")
            else:
                # 不需要归档：只有之前被标记为"已归档"的才清空，保留原有公式
                if is_current_archived:
                    ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='')
                    changes += 1
                    print(f"   📤 取消归档: Excel第{row_num}行")
        
        # ============== 1.5. 处理编辑（localEdits）- 不改变行号，在删除前应用 ==============
        # 字段名到Excel列号的映射（1-based）
        FIELD_TO_COLUMN = {
            '部门': 5,           # E列
            '项目': 6,           # F列
            '项目开始时间': 7,   # G列
            '项目结束时间': 8,   # H列
            '项目描述': 9,       # I列
            '资源类型': 10,      # J列
            '资源名称': 11,      # K列
            '资源开始时间': 12,  # L列
            '资源结束时间': 13,  # M列
            '日平均工时': 14,    # N列
        }
        
        local_edits = collab.get('localEdits', {})
        if local_edits:
            # 先构建要删除的ID集合，跳过已删除项目的编辑
            deleted_ids_for_edit = set()
            for did in collab.get('deletedIds', []):
                try:
                    deleted_ids_for_edit.add(int(did))
                except:
                    deleted_ids_for_edit.add(did)
                    deleted_ids_for_edit.add(str(did))
            
            for pid, edits in local_edits.items():
                # pid 可能是整数或字符串
                try:
                    pid_int = int(pid)
                except:
                    pid_int = pid
                
                # 跳过已删除项目的编辑
                if pid_int in deleted_ids_for_edit or str(pid_int) in deleted_ids_for_edit:
                    continue
                
                row_num = id_to_excel_row.get(pid_int) or id_to_excel_row.get(str(pid_int))
                if not row_num:
                    continue
                
                if not isinstance(edits, dict):
                    continue
                
                for field, value in edits.items():
                    col = FIELD_TO_COLUMN.get(field)
                    if col:
                        # 跳过空日期值
                        if field in ('项目开始时间', '项目结束时间', '资源开始时间', '资源结束时间'):
                            if not value or value in ['', '1900-01-01', '2100-01-01']:
                                continue
                        try:
                            # 使用安全写入，自动处理合并单元格
                            write_ok = _safe_write_cell(ws, row_num, col, value)
                            if write_ok:
                                changes += 1
                                print(f"   ✏️  编辑: Excel第{row_num}行, {field}={value}")
                            else:
                                errors.append(f"编辑失败: 第{row_num}行 {field}")
                        except Exception as e:
                            # 单个字段编辑失败不影响其他操作
                            err_msg = f"编辑第{row_num}行{field}失败: {str(e)}"
                            errors.append(err_msg)
                            print(f"   ⚠️  {err_msg}")
        
        # ============== 2. 再处理删除（软删除：标记V列为"已删除"，不物理删除行） ==============
        # 注意：不能用 ws.delete_rows() 物理删除，因为Excel有大量合并单元格
        # 物理删除会导致合并区域错乱，丢失多行数据
        deleted_ids = set()
        for did in collab.get('deletedIds', []):
            try:
                deleted_ids.add(int(did))
            except:
                deleted_ids.add(did)
        
        if deleted_ids:
            for p in existing_projects:
                pid = p['id']
                if pid in deleted_ids or str(pid) in deleted_ids:
                    row_num = id_to_excel_row.get(pid) or id_to_excel_row.get(str(pid))
                    if row_num:
                        # 软删除：在V列(COL_DELETED+1)标记"已删除"
                        try:
                            _safe_write_cell(ws, row_num, COL_DELETED + 1, '已删除')
                            changes += 1
                            print(f"   🗑️  软删除: Excel第{row_num}行 ({p.get('项目', '')})")
                        except Exception as e:
                            err_msg = f"软删除第{row_num}行失败: {str(e)}"
                            errors.append(err_msg)
                            print(f"   ⚠️  {err_msg}")
        
        # ============== 3. 最后处理新增项目（追加到末尾） ==============
        new_projects = collab.get('newProjects', [])
        
        if new_projects:
            # 【修复】构建 Excel 中已有的项目集合，防止重复添加
            # （当 git_push 失败后重试时，协作数据未被清空，需要去重）
            existing_set = set()
            for p in existing_projects:
                key = (str(p.get('项目', '')).strip(), str(p.get('资源名称', '')).strip())
                existing_set.add(key)
            
            last_row = ws.max_row
            
            for np in new_projects:
                # 去重检查：项目名+资源名称完全一致则跳过
                np_key = (str(np.get('项目', '')).strip(), str(np.get('资源名称', '')).strip())
                if np_key in existing_set:
                    print(f"   ⏭️  跳过重复新增: {np_key[0]} / {np_key[1]}")
                    continue
                last_row += 1
                
                # 列5: 部门（列E=5）
                ws.cell(row=last_row, column=5, value=np.get('部门', ''))
                # 列6: 项目名（列F=6）
                ws.cell(row=last_row, column=6, value=np.get('项目', ''))
                # 列7: 项目开始（列G=7）
                start_val = np.get('项目开始时间', '')
                if start_val and start_val not in ['', '1900-01-01']:
                    ws.cell(row=last_row, column=7, value=start_val)
                # 列8: 项目结束（列H=8）
                end_val = np.get('项目结束时间', '')
                if end_val and end_val not in ['', '2100-01-01']:
                    ws.cell(row=last_row, column=8, value=end_val)
                # 列9: 项目描述（列I=9）
                ws.cell(row=last_row, column=9, value=np.get('项目描述', ''))
                # 列10: 资源类型（列J=10）
                ws.cell(row=last_row, column=10, value=np.get('资源类型', ''))
                # 列11: 资源名称（列K=11）
                ws.cell(row=last_row, column=11, value=np.get('资源名称', ''))
                # 列12: 资源开始（列L=12）
                res_start = np.get('资源开始时间', '')
                if res_start and res_start not in ['', '1900-01-01']:
                    ws.cell(row=last_row, column=12, value=res_start)
                # 列13: 资源结束（列M=13）
                res_end = np.get('资源结束时间', '')
                if res_end and res_end not in ['', '2100-01-01']:
                    ws.cell(row=last_row, column=13, value=res_end)
                # 列14: 日平均工时（列N=14）
                ws.cell(row=last_row, column=14, value=np.get('日平均工时', 0) or 0)
                # 列21: 归档标志（列U=21）
                if np.get('已归档'):
                    ws.cell(row=last_row, column=COL_ARCHIVED + 1, value='已归档')
                
                changes += 1
                print(f"   ➕ 新增: {np.get('项目', '')} / {np.get('资源名称', '')}")
        
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        
    except Exception as e:
        return False, f'写入 Excel 失败: {str(e)}', changes
    
    if errors:
        error_detail = '; '.join(errors[:5])  # 最多显示5个错误
        if len(errors) > 5:
            error_detail += f' 等{len(errors)}个错误'
        return True, f'已应用 {changes} 项变更到 Excel (部分警告: {error_detail})', changes
    
    return True, f'已应用 {changes} 项变更到 Excel', changes

def regenerate_report() -> tuple[bool, str]:
    """重新生成 HTML 报表"""
    try:
        script_path = os.path.join(BASE_DIR, '更新点检表.py')
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=60
        )
        if result.returncode != 0:
            return False, f'生成报表失败: {result.stderr[:300]}'
        return True, '报表已重新生成'
    except subprocess.TimeoutExpired:
        return False, '生成报表超时'
    except Exception as e:
        return False, f'生成报表异常: {str(e)}'

# ==================== 全量同步入口 ====================

def full_sync(operation: str = '未知操作') -> tuple[bool, str]:
    """
    执行完整同步流程：
    1. 将协作数据应用到 Excel
    2. 重新生成报表
    3. 推送到 GitHub
    4. 【关键修复】只有推送成功后才清空协作数据，防止推送失败时数据丢失
    
    返回: (整体是否成功, 消息)
    - 只有 Excel写入+报表生成+Git推送 全部成功才返回 True
    - 任何一步失败都返回 False，协作数据保留以便下次重试
    """
    messages = []
    
    # 步骤1: 应用协作数据到 Excel（最关键的一步）
    ok, msg, changes = apply_collab_to_excel()
    messages.append(msg)
    if not ok:
        return False, '; '.join(messages)
    
    if changes > 0:
        # 步骤2: 重新生成报表
        ok, msg = regenerate_report()
        messages.append(msg)
        if not ok:
            # 报表生成失败但数据已在Excel中，保留协作数据以便下次重试
            messages.append('警告：报表生成失败，协作数据保留待重试')
            return False, '；'.join(messages)
        
        # 步骤3: 推送到 GitHub
        push_ok, push_msg = git_push(f'{operation}，{changes}项变更')
        messages.append(push_msg)
        if not push_ok:
            # 【关键修复】推送失败时不清空协作数据，保留以便下次重试
            messages.append('错误：GitHub推送失败，协作数据已保留待重试（请勿重启服务器！）')
            return False, '；'.join(messages)
        
        # 【关键修复】只有推送成功后才清空协作数据
        collab = load_collab_data()
        collab['newProjects'] = []
        collab['deletedIds'] = []
        collab['archived'] = {}
        collab['localEdits'] = {}
        collab['notes'] = {}
        collab['checked'] = {}
        save_collab_data(collab)
    else:
        # 【修复】changes=0 但协作数据不为空时，说明是重试场景
        # （Excel已被修改但上次推送失败，协作数据还保留着）
        # 直接检查是否有未提交的变更并尝试推送
        collab = load_collab_data()
        has_pending = (
            collab.get('newProjects') or 
            collab.get('deletedIds') or 
            collab.get('archived') or 
            collab.get('localEdits')
        )
        if has_pending:
            messages.append('检测到待推送的协作数据（重试场景）')
            # 重新生成报表（确保报表是最新的）
            ok, msg = regenerate_report()
            messages.append(msg)
            if not ok:
                messages.append('警告：报表生成失败')
                return False, '；'.join(messages)
            # 尝试推送
            push_ok, push_msg = git_push(f'{operation}，重试推送')
            messages.append(push_msg)
            if not push_ok:
                messages.append('错误：GitHub推送失败，协作数据已保留待重试')
                return False, '；'.join(messages)
            # 推送成功，清空协作数据
            collab['newProjects'] = []
            collab['deletedIds'] = []
            collab['archived'] = {}
            collab['localEdits'] = {}
            collab['notes'] = {}
            collab['checked'] = {}
            save_collab_data(collab)
        else:
            # 【关键修复】即使没有待处理的协作数据，也要检查是否有其他文件变更
            # （比如超声波户表脚本.xlsx 被外部直接修改、报表重新生成等）
            # 确保所有变更都能被推送到 GitHub，避免重新部署后丢失
            has_file_changes = False
            try:
                status = subprocess.run(
                    ['git', 'status', '--porcelain',
                     '超声波户表脚本.xlsx', '用户管理.xlsx',
                     '项目延期点检表.html', 'data/'],
                    capture_output=True, text=True, cwd=BASE_DIR, timeout=10
                )
                if status.stdout.strip():
                    has_file_changes = True
                    print(f"[同步] 检测到非协作数据的文件变更:\n{status.stdout.strip()}")
            except:
                pass
            
            if has_file_changes:
                messages.append('检测到文件变更，正在同步...')
                # 重新生成报表（确保报表是最新的）
                ok, msg = regenerate_report()
                messages.append(msg)
                if not ok:
                    messages.append('警告：报表生成失败')
                # 尝试推送所有变更
                push_ok, push_msg = git_push(f'{operation}，文件变更同步')
                messages.append(push_msg)
                if not push_ok:
                    messages.append('错误：GitHub推送失败')
                    return False, '；'.join(messages)
            else:
                messages.append('无需要同步的变更')
    
    return True, '；'.join(messages)

# ==================== 简化方案：原子操作函数 ====================
# 【简化方案】每个操作直接修改 Excel + 生成报表 + 推送 GitHub
# 不再通过协作数据 JSON 中转，避免多源状态不同步的问题

def _find_row_for_project_id(project_id: int):
    """根据项目ID查找 Excel 中的行号（1-based），返回 (ws, row_num) 或 (None, None)"""
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_FILE):
        return None, None
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        # ID = Excel 行号（idx = row_num - 1）
        row_num = project_id + 1
        if 1 <= row_num <= ws.max_row:
            return wb, row_num
        return None, None
    except Exception:
        return None, None

# ==================== 新增项目辅助函数 ====================

def _normalize_date(val, is_start=True):
    """日期标准化：空值/哨兵值统一处理
    - 开始日期: 空/'/' → '1900-01-01'
    - 结束日期: 空 → '2100-01-01'
    - 其他有效日期: 直接返回
    """
    if is_start:
        if not val or str(val).strip() == '' or str(val).strip() == '/':
            return '1900-01-01'
        return str(val).strip()
    else:
        if not val or str(val).strip() == '':
            return '2100-01-01'
        return str(val).strip()

def _copy_formulas(ws, source_row, target_row):
    """将 source_row 的 O-W 列(15-23)公式复制到 target_row，替换相对引用行号
    - 绝对引用($A$2, C$2 保持不变
    - 相对引用(K4, L4 等替换为 target_row
    """
    import re
    for col in range(15, 24):  # O(15) 到 W(23)
        source_val = ws.cell(row=source_row, column=col).value
        if source_val and isinstance(source_val, str) and source_val.startswith('='):
            # 只替换相对引用的行号: 字母+source_row且不是$source_row
            # 即: 前面没有$的字母后接source_row
            new_val = re.sub(
                r'(?<!\$)([A-Z])' + str(source_row) + r'\b',
                r'\g<1>' + str(target_row),
                source_val
            )
            ws.cell(row=target_row, column=col, value=new_val)

def _format_resource_name(name, ws=None):
    """3层查找: 纯姓名 → Excel格式 " @姓名(工号)(姓名)"
    第1层: 在 getAllEmails()/现有人员列表中查找
    第2层: 在 Excel 已有 K 列中模糊匹配姓名
    第3层: 使用简化格式 " @姓名"
    """
    if not name:
        return ''
    name = str(name).strip()
    if name.startswith(' @'):  # 已经是Excel格式，直接返回
        return name
    
    # 第1层: 从 getAllEmails 查找
    try:
        emails = getAllEmails()
        for email_info in emails:
            display = email_info.get('display', '')
            # display 可能是 "项家祺(80004213)" 或 "项家祺"
            if name in display:
                # 提取工号
                import re
                m = re.search(r'\((\d+)\)', display)
                emp_id = m.group(1) if m else ''
                if emp_id:
                    return f' @{name}({emp_id})({name})'
                else:
                    return f' @{name}'
    except Exception:
        pass
    
    # 第2层: 在 Excel K 列中模糊匹配
    if ws is not None:
        try:
            found_formats = set()
            for row in range(4, ws.max_row + 1):
                val = ws.cell(row=row, column=11).value
                if val and name in str(val):
                    found_formats.add(str(val).strip())
            if found_formats:
                # 返回最长的那个（通常信息最全）
                return max(found_formats, key=len)
        except Exception:
            pass
    
    # 第3层: 简化格式
    return f' @{name}'

def action_add_project(project_data: dict, operator: str = 'unknown') -> tuple[bool, str]:
    """【简化方案】新增项目：直接写入 Excel + 生成报表 + 推送 GitHub
    
    关键：只要 Excel 写入成功就算成功，报表生成和 Git 推送不阻塞主操作
    （避免因环境问题导致新增失败，但数据实际已写入）
    """
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            return False, 'Excel 文件不存在'
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        last_row = ws.max_row + 1
        new_id = last_row - 1  # ID = 行号 - 1
        
        # 写入各列（与 apply_collab_to_excel 中新增项目逻辑一致）
        ws.cell(row=last_row, column=5, value=project_data.get('部门', ''))
        ws.cell(row=last_row, column=6, value=project_data.get('项目', ''))
        # 【修复】日期标准化：空/哨兵值统一处理，直接写入（与读取端对称）
        ws.cell(row=last_row, column=7, value=_normalize_date(project_data.get('项目开始时间', ''), is_start=True))
        ws.cell(row=last_row, column=8, value=_normalize_date(project_data.get('项目结束时间', ''), is_start=False))
        ws.cell(row=last_row, column=9, value=project_data.get('项目描述', ''))
        ws.cell(row=last_row, column=10, value=project_data.get('资源类型', ''))
        # 【修复】资源名称格式转换：纯姓名 → Excel格式
        ws.cell(row=last_row, column=11, value=_format_resource_name(project_data.get('资源名称', ''), ws))
        ws.cell(row=last_row, column=12, value=_normalize_date(project_data.get('资源开始时间', ''), is_start=True))
        ws.cell(row=last_row, column=13, value=_normalize_date(project_data.get('资源结束时间', ''), is_start=False))
        ws.cell(row=last_row, column=14, value=project_data.get('日平均工时', 0) or 0)
        if project_data.get('已归档'):
            ws.cell(row=last_row, column=COL_ARCHIVED + 1, value='已归档')
        
        # 【修复】复制 O-W 列公式（从第4行复制，第4行是有完整公式的参考行）
        ref_row = 4 if last_row > 4 else last_row - 1 if last_row > 1 else 4
        _copy_formulas(ws, ref_row, last_row)
        
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        
        # 【双重保护】同时写入协作数据 newProjects，防止Excel被覆盖时数据丢失
        try:
            collab = load_collab_data()
            existing_ids = {p.get('id') for p in collab.get('newProjects', [])}
            if new_id not in existing_ids:
                all_projects = read_excel_projects()
                for p in all_projects:
                    if p['id'] == new_id:
                        collab.setdefault('newProjects', []).append(p)
                        break
                collab['lastUpdate'] = datetime.now().isoformat()
                save_collab_data(collab)
                print(f'[新增项目] 已写入ID={new_id}到协作数据（双重保护）')
        except Exception as e:
            print(f'[新增项目] 写入协作数据失败（不影响Excel写入）: {e}')
        
        # 【关键】Excel 已写入成功！报表+推送放到后台线程，不阻塞HTTP响应
        messages = [f'新增成功(ID={new_id})']
        commit_msg = f'{operator}新增项目: {project_data.get("项目", "")}'
        thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
        thread.start()
        messages.append('（报表和同步正在后台执行）')
        
        return True, '; '.join(messages)
    except Exception as e:
        return False, f'新增失败: {str(e)}'

def action_add_project_batch(projects_data: list, operator: str = 'unknown') -> dict:
    """批量新增项目：一次打开Excel，写入所有行，复制公式，保存一次
    
    Args:
        projects_data: 项目数据列表，每条格式同 action_add_project
        operator: 操作人
    
    Returns:
        dict: {
            'success': bool,
            'total': int,
            'added': int,
            'skipped': int,
            'errors': [str],
            'ids': [int],
            'message': str
        }
    """
    MAX_BATCH = 50
    result = {
        'success': False,
        'total': len(projects_data),
        'added': 0,
        'skipped': 0,
        'errors': [],
        'ids': [],
        'message': ''
    }
    
    if not projects_data:
        result['message'] = '没有数据需要导入'
        return result
    
    # 50条限制
    if len(projects_data) > MAX_BATCH:
        result['errors'].append(f'超过{MAX_BATCH}条限制，仅处理前{MAX_BATCH}条')
        projects_data = projects_data[:MAX_BATCH]
        result['total'] = MAX_BATCH
    
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            result['message'] = 'Excel 文件不存在'
            return result
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        ref_row = 4  # 公式参考行
        
        for idx, project_data in enumerate(projects_data):
            try:
                # ========== 写入前校验（7道关卡） ==========
                skip_reason = None
                
                # 1. 资源名称非空
                res_name = str(project_data.get('资源名称', '')).strip()
                if not res_name:
                    skip_reason = '资源名称为空'
                
                # 2. 资源结束时间非空
                res_end = str(project_data.get('资源结束时间', '')).strip()
                if not skip_reason and not res_end:
                    skip_reason = '资源结束时间为空'
                
                # 3. 日平均工时 0-24
                try:
                    daily_hours = float(project_data.get('日平均工时', 0) or 0)
                    daily_hours = max(0, min(24, daily_hours))  # 截断到0-24
                except (ValueError, TypeError):
                    daily_hours = 0
                
                # 4. 日期格式校验 & 5. 结束 ≥ 开始
                try:
                    proj_start = _normalize_date(project_data.get('项目开始时间', ''), True)
                    proj_end = _normalize_date(project_data.get('项目结束时间', ''), False)
                    res_start = _normalize_date(project_data.get('资源开始时间', ''), True)
                    res_end_norm = _normalize_date(res_end, False)
                    
                    # 简单日期比较（YYYY-MM-DD字符串比较即可）
                    if res_end_norm < res_start and res_start != '1900-01-01':
                        skip_reason = f'结束时间({res_end_norm})早于开始时间({res_start})'
                except Exception as e:
                    skip_reason = f'日期格式错误: {e}'
                
                if skip_reason:
                    result['skipped'] += 1
                    result['errors'].append(f'第{idx+1}条跳过: {skip_reason}')
                    continue
                
                # ========== 写入 ==========
                last_row = ws.max_row + 1
                new_id = last_row - 1
                
                ws.cell(row=last_row, column=5, value=project_data.get('部门', ''))
                ws.cell(row=last_row, column=6, value=project_data.get('项目', ''))
                ws.cell(row=last_row, column=7, value=proj_start)
                ws.cell(row=last_row, column=8, value=proj_end)
                ws.cell(row=last_row, column=9, value=project_data.get('项目描述', ''))
                ws.cell(row=last_row, column=10, value=project_data.get('资源类型', ''))
                ws.cell(row=last_row, column=11, value=_format_resource_name(res_name, ws))
                ws.cell(row=last_row, column=12, value=res_start)
                ws.cell(row=last_row, column=13, value=res_end_norm)
                ws.cell(row=last_row, column=14, value=daily_hours)
                if project_data.get('已归档'):
                    ws.cell(row=last_row, column=COL_ARCHIVED + 1, value='已归档')
                
                # 复制公式
                _copy_formulas(ws, ref_row, last_row)
                
                result['added'] += 1
                result['ids'].append(new_id)
                
            except Exception as e:
                result['skipped'] += 1
                result['errors'].append(f'第{idx+1}条异常: {str(e)}')
                continue
        
        # 一次保存
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        
        # 【双重保护】同时写入协作数据 newProjects，防止Excel被覆盖时数据丢失
        # （git pull 或其他场景覆盖Excel后，full_sync 可从协作数据恢复这些项目）
        if result['added'] > 0:
            try:
                collab = load_collab_data()
                existing_ids = {p.get('id') for p in collab.get('newProjects', [])}
                # 重新读取刚写入的项目，获取完整数据
                all_projects = read_excel_projects()
                added_ids_set = set(result['ids'])
                for p in all_projects:
                    if p['id'] in added_ids_set and p['id'] not in existing_ids:
                        collab.setdefault('newProjects', []).append(p)
                collab['lastUpdate'] = datetime.now().isoformat()
                save_collab_data(collab)
                print(f'[批量导入] 已写入{result["added"]}个项目到协作数据（双重保护）')
            except Exception as e:
                print(f'[批量导入] 写入协作数据失败（不影响Excel写入）: {e}')
        
        result['success'] = True
        result['message'] = f'批量导入完成: 成功{result["added"]}条，跳过{result["skipped"]}条（共{result["total"]}条）'
        
        # 后台报表+推送
        commit_msg = f'{operator}批量导入{result["added"]}个项目'
        thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
        thread.start()
        result['message'] += '（报表和同步正在后台执行）'
        
        return result
        
    except Exception as e:
        result['message'] = f'批量导入失败: {str(e)}'
        result['errors'].append(str(e))
        return result

def _background_report_and_push(commit_msg: str):
    """【后台线程】生成报表 + 推送 GitHub（并行执行，总耗时 = max(报表, push) 而非 sum）
    
    优化前（串行）：耗时 = 报表(3-8s) + push(2-4s) = 5-12s
    优化后（并行）：耗时 = max(报表(3-8s), push(2-4s)) = 3-8s
    """
    import concurrent.futures
    
    def _do_report():
        try:
            ok, msg = regenerate_report()
            print(f'[后台] 报表生成: {msg}')
            return ok
        except Exception as e:
            print(f'[后台] 报表生成异常: {e}')
            return False
    
    def _do_push():
        try:
            ok, msg = git_push(commit_msg)
            print(f'[后台] Git推送: {msg}')
            return ok
        except Exception as e:
            print(f'[后台] Git推送异常: {e}')
            return False
    
    # 并行执行报表生成和 Git 推送
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f1 = executor.submit(_do_report)
        f2 = executor.submit(_do_push)
        # 等待两个都完成（或超时）
        try:
            concurrent.futures.wait([f1, f2], timeout=60)
        except:
            pass


def _finish_action(excel_written: bool, action_name: str, operator: str, detail: str = '') -> tuple[bool, str]:
    """统一的操作完成处理：Excel写入成功即算成功，报表和推送在后台异步执行
    
    关键：先返回 HTTP 响应，再执行耗时操作，彻底解决 Render 502 超时问题
    
    Args:
        excel_written: Excel是否已成功写入
        action_name: 操作名称（如"新增项目"、"归档"）
        operator: 操作人
        detail: 额外详情
    
    Returns:
        (是否成功, 消息)
    """
    if not excel_written:
        return False, f'{action_name}失败：Excel写入未完成'
    
    messages = [f'{action_name}成功']
    if detail:
        messages.append(detail)
    
    # 【关键】报表生成和 Git 推送放到后台线程，不阻塞 HTTP 响应
    # （Render 免费版 30 秒超时，Git 推送经常超过这个时间）
    commit_msg = f'{operator}{action_name}'
    thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
    thread.start()
    
    messages.append('（报表和同步正在后台执行）')
    return True, '; '.join(messages)


def action_delete_project(project_id: int, operator: str = 'unknown') -> tuple[bool, str]:
    """【简化方案】删除项目：软删除（V列标记）"""
    wb, row_num = _find_row_for_project_id(project_id)
    if wb is None:
        return False, f'项目ID={project_id} 不存在'
    try:
        ws = wb['任务计划表']
        _safe_write_cell(ws, row_num, COL_DELETED + 1, '已删除')
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        return _finish_action(True, '删除项目', operator, f'ID={project_id}')
    except Exception as e:
        return False, f'删除失败: {str(e)}'

# ========== 批量操作（一次打开Excel，处理所有ID，保存一次） ==========

def action_batch_archive(project_ids: list, operator: str = 'unknown') -> dict:
    """批量归档项目：一次打开Excel，标记所有ID为已归档，保存一次
    
    Args:
        project_ids: 项目ID列表
        operator: 操作人
    
    Returns:
        dict: {success, total, done, errors, message}
    """
    result = {'success': False, 'total': len(project_ids), 'done': 0, 'errors': []}
    if not project_ids:
        result['message'] = '没有要归档的项目'
        return result
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            result['message'] = 'Excel 文件不存在'
            return result
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        done_ids = []
        for pid in project_ids:
            try:
                row_num = int(pid) + 1
                if 1 <= row_num <= ws.max_row:
                    ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='已归档')
                    done_ids.append(pid)
                    result['done'] += 1
                else:
                    result['errors'].append(f'ID={pid} 超出范围')
            except Exception as e:
                result['errors'].append(f'ID={pid}: {str(e)}')
        if result['done'] > 0:
            wb.save(EXCEL_FILE)
            invalidate_projects_cache()
            result['success'] = True
            commit_msg = f'{operator}批量归档{result["done"]}个项目'
            thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
            thread.start()
            result['message'] = f'成功归档 {result["done"]}/{result["total"]} 个项目（报表和同步正在后台执行）'
        else:
            result['message'] = '没有项目被成功归档'
        return result
    except Exception as e:
        result['message'] = f'批量归档失败: {str(e)}'
        result['errors'].append(str(e))
        return result

def action_batch_unarchive(project_ids: list, operator: str = 'unknown') -> dict:
    """批量恢复归档项目：一次打开Excel，清空所有ID的归档标记，保存一次"""
    result = {'success': False, 'total': len(project_ids), 'done': 0, 'errors': []}
    if not project_ids:
        result['message'] = '没有要恢复的项目'
        return result
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            result['message'] = 'Excel 文件不存在'
            return result
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        for pid in project_ids:
            try:
                row_num = int(pid) + 1
                if 1 <= row_num <= ws.max_row:
                    ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='')
                    result['done'] += 1
                else:
                    result['errors'].append(f'ID={pid} 超出范围')
            except Exception as e:
                result['errors'].append(f'ID={pid}: {str(e)}')
        if result['done'] > 0:
            wb.save(EXCEL_FILE)
            invalidate_projects_cache()
            result['success'] = True
            commit_msg = f'{operator}批量恢复{result["done"]}个项目归档'
            thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
            thread.start()
            result['message'] = f'成功恢复 {result["done"]}/{result["total"]} 个项目（报表和同步正在后台执行）'
        else:
            result['message'] = '没有项目被成功恢复'
        return result
    except Exception as e:
        result['message'] = f'批量恢复失败: {str(e)}'
        result['errors'].append(str(e))
        return result

def action_batch_delete(project_ids: list, operator: str = 'unknown') -> dict:
    """批量删除项目：一次打开Excel，标记所有ID为已删除，保存一次"""
    result = {'success': False, 'total': len(project_ids), 'done': 0, 'errors': []}
    if not project_ids:
        result['message'] = '没有要删除的项目'
        return result
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            result['message'] = 'Excel 文件不存在'
            return result
        wb = load_workbook(EXCEL_FILE)
        ws = wb['任务计划表']
        for pid in project_ids:
            try:
                row_num = int(pid) + 1
                if 1 <= row_num <= ws.max_row:
                    _safe_write_cell(ws, row_num, COL_DELETED + 1, '已删除')
                    result['done'] += 1
                else:
                    result['errors'].append(f'ID={pid} 超出范围')
            except Exception as e:
                result['errors'].append(f'ID={pid}: {str(e)}')
        if result['done'] > 0:
            wb.save(EXCEL_FILE)
            invalidate_projects_cache()
            result['success'] = True
            commit_msg = f'{operator}批量删除{result["done"]}个项目'
            thread = threading.Thread(target=_background_report_and_push, args=(commit_msg,), daemon=True)
            thread.start()
            result['message'] = f'成功删除 {result["done"]}/{result["total"]} 个项目（报表和同步正在后台执行）'
        else:
            result['message'] = '没有项目被成功删除'
        return result
    except Exception as e:
        result['message'] = f'批量删除失败: {str(e)}'
        result['errors'].append(str(e))
        return result

def action_archive_project(project_id: int, operator: str = 'unknown') -> tuple[bool, str]:
    """【简化方案】归档项目：A列标记"已归档" """
    wb, row_num = _find_row_for_project_id(project_id)
    if wb is None:
        return False, f'项目ID={project_id} 不存在'
    try:
        ws = wb['任务计划表']
        ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='已归档')
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        return _finish_action(True, '归档项目', operator, f'ID={project_id}')
    except Exception as e:
        return False, f'归档失败: {str(e)}'

def action_unarchive_project(project_id: int, operator: str = 'unknown') -> tuple[bool, str]:
    """【简化方案】取消归档：清空A列归档标志"""
    wb, row_num = _find_row_for_project_id(project_id)
    if wb is None:
        return False, f'项目ID={project_id} 不存在'
    try:
        ws = wb['任务计划表']
        ws.cell(row=row_num, column=COL_ARCHIVED + 1, value='')
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        return _finish_action(True, '取消归档项目', operator, f'ID={project_id}')
    except Exception as e:
        return False, f'取消归档失败: {str(e)}'

def action_edit_project(project_id: int, edit_data: dict, operator: str = 'unknown') -> tuple[bool, str]:
    """【简化方案】编辑项目：直接修改Excel单元格"""
    wb, row_num = _find_row_for_project_id(project_id)
    if wb is None:
        return False, f'项目ID={project_id} 不存在'
    try:
        ws = wb['任务计划表']
        col_map = {
            '部门': 5, '项目': 6,
            '项目开始时间': 7, '项目结束时间': 8,
            '项目描述': 9, '资源类型': 10, '资源名称': 11,
            '资源开始时间': 12, '资源结束时间': 13,
            '日平均工时': 14
        }
        for field, value in edit_data.items():
            if field in col_map:
                _safe_write_cell(ws, row_num, col_map[field], value)
        wb.save(EXCEL_FILE)
        invalidate_projects_cache()
        return _finish_action(True, '编辑项目', operator, f'ID={project_id}')
    except Exception as e:
        return False, f'编辑失败: {str(e)}'

def startup_sync() -> tuple[bool, str]:
    """服务器启动时同步：确保Git仓库 + 拉取最新 + 生成报表
    
    【关键修复】Render 部署时 .git 目录可能不存在，
    必须先确保 Git 仓库可用，再进行拉取操作。
    """
    messages = []
    
    # 先确保 Git 仓库存在（Render 部署环境保障）
    ensure_ok, ensure_msg = ensure_git_repo()
    messages.append(ensure_msg)
    if not ensure_ok:
        messages.append('警告：Git仓库不可用，将无法同步数据到GitHub')
    
    ok, msg = git_pull()
    messages.append(msg)
    ok2, msg2 = regenerate_report()
    messages.append(msg2)
    return (ok and ok2), '；'.join(messages)

# ==================== 审批系统 & 操作记录 ====================

def _ensure_operations_sheet():
    """确保「操作记录」Sheet存在，不存在则创建并写入表头"""
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE)
    if OPERATIONS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(OPERATIONS_SHEET)
        for col_idx, header in enumerate(OPERATION_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(EXCEL_FILE)
        print(f'[操作记录] 已创建Sheet: {OPERATIONS_SHEET}')
    else:
        wb.close()


def _load_operations_sheet():
    """读取操作记录Sheet为列表（直接读取Excel，无缓存）
    
    Returns:
        list: 操作记录列表，每条为dict
    """
    _ensure_operations_sheet()
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[OPERATIONS_SHEET]
    
    operations = []
    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    
    import json
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers):
            if header:
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                # JSON字符串解析回list/dict（项目ID列表、项目名列表、变更前/后内容等）
                if isinstance(val, str) and val:
                    stripped = val.strip()
                    if (stripped.startswith('[') and stripped.endswith(']')) or \
                       (stripped.startswith('{') and stripped.endswith('}')):
                        try:
                            val = json.loads(stripped)
                        except (json.JSONDecodeError, ValueError):
                            pass
                row_data[header] = val
        if row_data.get('操作ID'):
            operations.append(row_data)
    
    wb.close()
    # 按操作时间倒序（最新的在前）
    operations.sort(key=lambda x: x.get('操作时间', ''), reverse=True)
    return operations


def _generate_op_id():
    """生成操作ID: OP-YYYYMMDD-NNN"""
    from datetime import datetime
    today = datetime.now().strftime('%Y%m%d')
    ops = _load_operations_sheet()
    today_count = sum(1 for o in ops if o.get('操作ID', '').startswith(f'OP-{today}'))
    return f'OP-{today}-{today_count + 1:03d}'


def _append_operation(operation_data):
    """追加一条操作记录到Excel
    
    Args:
        operation_data: dict，键为OPERATION_HEADERS中的字段名
    
    Returns:
        str: 操作ID
    """
    _ensure_operations_sheet()
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE)
    ws = wb[OPERATIONS_SHEET]
    
    # 确保有操作ID
    if '操作ID' not in operation_data or not operation_data['操作ID']:
        operation_data['操作ID'] = _generate_op_id()
    
    new_row = ws.max_row + 1
    for col_idx, header in enumerate(OPERATION_HEADERS, 1):
        val = operation_data.get(header, '')
        # dict/list转JSON字符串
        if isinstance(val, (dict, list)):
            import json
            val = json.dumps(val, ensure_ascii=False)
        ws.cell(row=new_row, column=col_idx, value=val)
    
    wb.save(EXCEL_FILE)
    wb.close()
    print(f'[操作记录] 已追加: {operation_data["操作ID"]}')
    return operation_data['操作ID']


def _update_operation(op_id, updates):
    """更新操作记录的字段
    
    Args:
        op_id: 操作ID
        updates: dict，要更新的字段
    """
    _ensure_operations_sheet()
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE)
    ws = wb[OPERATIONS_SHEET]
    
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == op_id:
            for key, val in updates.items():
                if key in headers:
                    col_idx = headers.index(key) + 1
                    if isinstance(val, (dict, list)):
                        import json
                        val = json.dumps(val, ensure_ascii=False)
                    ws.cell(row=row_idx, column=col_idx, value=val)
            break
    
    wb.save(EXCEL_FILE)
    wb.close()
    print(f'[操作记录] 已更新: {op_id}')


def _write_approval_row(project_id, status, submitter, op_type, detail=''):
    """写入/清除单条审批信息到任务计划表的X/Y/Z/AA列
    
    Args:
        project_id: 项目ID (int)
        status: 审批状态 (PENDING_* 或 空字符串表示清除)
        submitter: 申请人用户名
        op_type: 操作类型 (archive/unarchive/edit)
        detail: 变更内容(JSON字符串，仅edit用)
    """
    from openpyxl import load_workbook
    from datetime import datetime
    wb = load_workbook(EXCEL_FILE)
    ws = wb['任务计划表']
    
    row_num = int(project_id) + 2  # id从0开始，第1行是表头，id=0对应第2行
    if 2 <= row_num <= ws.max_row:
        ws.cell(row=row_num, column=COL_APPROVAL_STATUS + 1, value=status)
        if status:
            # 写入审批信息
            submitter_info = f"{submitter}|{datetime.now().isoformat()}"
            ws.cell(row=row_num, column=COL_APPROVAL_SUBMITTER + 1, value=submitter_info)
            ws.cell(row=row_num, column=COL_APPROVAL_TYPE + 1, value=op_type)
            if detail:
                ws.cell(row=row_num, column=COL_APPROVAL_DETAIL + 1, value=detail)
        else:
            # 清除审批信息
            ws.cell(row=row_num, column=COL_APPROVAL_SUBMITTER + 1, value='')
            ws.cell(row=row_num, column=COL_APPROVAL_TYPE + 1, value='')
            ws.cell(row=row_num, column=COL_APPROVAL_DETAIL + 1, value='')
    
    wb.save(EXCEL_FILE)
    wb.close()
    invalidate_projects_cache()


def _read_approval_row(project_id):
    """读取单条审批信息
    
    Returns:
        dict or None: {status, submitter, submit_time, op_type, detail}
    """
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['任务计划表']
    
    row_num = int(project_id) + 2
    if row_num < 2 or row_num > ws.max_row:
        wb.close()
        return None
    
    status = ws.cell(row=row_num, column=COL_APPROVAL_STATUS + 1).value or ''
    submitter_raw = ws.cell(row=row_num, column=COL_APPROVAL_SUBMITTER + 1).value or ''
    op_type = ws.cell(row=row_num, column=COL_APPROVAL_TYPE + 1).value or ''
    detail = ws.cell(row=row_num, column=COL_APPROVAL_DETAIL + 1).value or ''
    
    wb.close()
    
    if not status:
        return None
    
    submitter = ''
    submit_time = ''
    if submitter_raw and '|' in submitter_raw:
        parts = submitter_raw.split('|', 1)
        submitter = parts[0]
        submit_time = parts[1] if len(parts) > 1 else ''
    
    return {
        'status': status,
        'submitter': submitter,
        'submit_time': submit_time,
        'op_type': op_type,
        'detail': detail,
    }


def submit_approval(approval_data, operator):
    """提交审批申请（简化版：直接写任务计划表X/Y/Z/AA列）
    
    Args:
        approval_data: dict，包含：
            - operation_type: 操作类型
            - project_ids: 项目ID列表
            - project_names: 项目名列表（保留兼容）
            - before_data: 变更前内容（保留兼容）
            - after_data: 变更后内容（edit操作用）
        operator: 操作人用户名
    """
    try:
        import json
        op_type = approval_data.get('operation_type', '')
        project_ids = approval_data.get('project_ids', [])
        after_data = approval_data.get('after_data', {})
        
        status_map = {
            'archive': 'PENDING_ARCHIVE',
            'unarchive': 'PENDING_UNARCHIVE',
            'edit': 'PENDING_EDIT',
            'batch_archive': 'PENDING_ARCHIVE',
            'batch_unarchive': 'PENDING_UNARCHIVE',
        }
        x_status = status_map.get(op_type, '')
        if not x_status:
            return False, f'不支持的操作类型: {op_type}', ''
        
        # 变更内容转JSON（仅编辑用）
        detail_json = ''
        if op_type == 'edit' and after_data:
            detail_json = json.dumps(after_data, ensure_ascii=False)
        
        # 逐行写入审批信息
        for pid in project_ids:
            _write_approval_row(pid, x_status, operator, op_type, detail_json)
        
        # 同时写操作记录（审计用，不参与核心流程）
        op_data = {
            '操作时间': datetime.now().isoformat(),
            '操作人': operator,
            '操作类型': op_type,
            '项目ID列表': project_ids,
            '项目名列表': approval_data.get('project_names', []),
            '变更前内容': approval_data.get('before_data', {}),
            '变更后内容': after_data,
            '状态': 'pending',
        }
        op_id = _append_operation(op_data)
        
        return True, f'审批申请已提交，等待审批', op_id
    except Exception as e:
        return False, f'提交审批失败: {str(e)}', ''


def approve_operation(op_id, approver, comment=''):
    """通过审批（简化版：从任务计划表读取审批信息，执行操作，清除列）"""
    from datetime import datetime
    try:
        # 先从操作记录找到项目ID
        ops = _load_operations_sheet()
        op = next((o for o in ops if o.get('操作ID') == op_id), None)
        if not op:
            return False, f'操作记录不存在: {op_id}'
        if op.get('状态') != 'pending':
            return False, f'该操作状态不是待审批: {op.get("状态")}'
        
        import json
        op_type = op.get('操作类型', '')
        project_ids = op.get('项目ID列表', [])
        if isinstance(project_ids, str):
            project_ids = json.loads(project_ids)
        
        # 执行实际操作
        ok = True
        msg = ''
        if op_type in ('archive', 'batch_archive'):
            if len(project_ids) == 1:
                ok, msg = action_archive_project(project_ids[0], approver)
            else:
                result = action_batch_archive(project_ids, approver)
                ok = result.get('success', False)
                msg = result.get('message', '')
        elif op_type in ('unarchive', 'batch_unarchive'):
            if len(project_ids) == 1:
                ok, msg = action_unarchive_project(project_ids[0], approver)
            else:
                result = action_batch_unarchive(project_ids, approver)
                ok = result.get('success', False)
                msg = result.get('message', '')
        elif op_type == 'edit':
            after_data = op.get('变更后内容', {})
            if isinstance(after_data, str) and after_data:
                after_data = json.loads(after_data)
            if project_ids and after_data:
                ok, msg = action_edit_project(project_ids[0], after_data, approver)
            else:
                ok, msg = False, '编辑操作缺少项目ID或变更内容'
        else:
            ok, msg = False, f'不支持的操作类型: {op_type}'
        
        if not ok:
            return False, f'执行操作失败: {msg}'
        
        # 清除任务计划表的审批列
        for pid in project_ids:
            _write_approval_row(pid, '', '', '', '')
        
        # 更新操作记录
        _update_operation(op_id, {
            '状态': 'approved',
            '审批人': approver,
            '审批时间': datetime.now().isoformat(),
        })
        
        return True, f'审批通过，操作已生效: {msg}'
    except Exception as e:
        return False, f'审批失败: {str(e)}'


def reject_operation(op_id, approver, comment=''):
    """拒绝审批（简化版）"""
    from datetime import datetime
    try:
        ops = _load_operations_sheet()
        op = next((o for o in ops if o.get('操作ID') == op_id), None)
        if not op:
            return False, f'操作记录不存在: {op_id}'
        if op.get('状态') != 'pending':
            return False, f'该操作状态不是待审批: {op.get("状态")}'
        
        import json
        project_ids = op.get('项目ID列表', [])
        if isinstance(project_ids, str):
            project_ids = json.loads(project_ids)
        
        # 清除任务计划表的审批列
        for pid in project_ids:
            _write_approval_row(pid, '', '', '', '')
        
        # 更新操作记录
        _update_operation(op_id, {
            '状态': 'rejected',
            '审批人': approver,
            '审批时间': datetime.now().isoformat(),
            '变更后内容': {'拒绝原因': comment} if comment else {},
        })
        
        return True, '审批已拒绝'
    except Exception as e:
        return False, f'拒绝审批失败: {str(e)}'


def cancel_operation(op_id, operator):
    """撤回审批申请（简化版）"""
    from datetime import datetime
    try:
        ops = _load_operations_sheet()
        op = next((o for o in ops if o.get('操作ID') == op_id), None)
        if not op:
            return False, f'操作记录不存在: {op_id}'
        if op.get('操作人') != operator:
            return False, '仅申请人本人可撤回'
        if op.get('状态') != 'pending':
            return False, f'该操作状态不是待审批，无法撤回（当前状态: {op.get("状态")}）'
        
        import json
        project_ids = op.get('项目ID列表', [])
        if isinstance(project_ids, str):
            project_ids = json.loads(project_ids)
        
        # 清除任务计划表的审批列
        for pid in project_ids:
            _write_approval_row(pid, '', '', '', '')
        
        # 更新操作记录
        _update_operation(op_id, {
            '状态': 'cancelled',
            '审批人': operator,
            '审批时间': datetime.now().isoformat(),
        })
        
        auth._audit_log('APPROVAL_CANCEL', operator, f'{op_id}: 撤回审批申请')
        return True, '已撤回审批申请'
    except Exception as e:
        return False, f'撤回失败: {str(e)}'


def list_approvals(user, role, permissions):
    """获取审批列表（简化版：从任务计划表X列过滤 + 操作记录补充）
    
    核心逻辑：
    - 待审批：扫描任务计划表X列不为空的行（即状态为PENDING_*）
    - 我发起的：从操作记录过滤操作人=当前用户
    """
    from openpyxl import load_workbook
    import json
    
    can_approve = 'approve' in permissions
    
    # === Part 1: 从任务计划表读取待审批项目 ===
    pending_for_me = []
    if can_approve:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['任务计划表']
        
        # 找项目名列（D列=3，或兼容其他列）
        project_name_col = 3  # 默认D列
        headers = [cell.value for cell in ws[1]]
        for i, h in enumerate(headers):
            if h and ('项目' in str(h) and '名' in str(h)):
                project_name_col = i
                break
        
        # 预加载操作记录，用于匹配真实的操作ID
        all_ops = _load_operations_sheet()
        pending_ops = [o for o in all_ops if o.get('状态') == 'pending']
        
        def _find_op_id(project_id, op_type, submitter):
            """从操作记录中找到匹配的操作ID"""
            import json as _json
            for o in pending_ops:
                if o.get('操作类型') != op_type:
                    continue
                if o.get('操作人') != submitter:
                    continue
                pids = o.get('项目ID列表', [])
                if isinstance(pids, str):
                    try:
                        pids = _json.loads(pids)
                    except:
                        pids = []
                if int(project_id) in [int(p) for p in pids]:
                    return o.get('操作ID', '')
            return ''
        
        for row_num in range(2, ws.max_row + 1):
            status = ws.cell(row=row_num, column=COL_APPROVAL_STATUS + 1).value or ''
            if not status:
                continue
            
            submitter_raw = ws.cell(row=row_num, column=COL_APPROVAL_SUBMITTER + 1).value or ''
            op_type = ws.cell(row=row_num, column=COL_APPROVAL_TYPE + 1).value or ''
            detail = ws.cell(row=row_num, column=COL_APPROVAL_DETAIL + 1).value or ''
            
            submitter = ''
            submit_time = ''
            if submitter_raw and '|' in submitter_raw:
                parts = submitter_raw.split('|', 1)
                submitter = parts[0]
                submit_time = parts[1] if len(parts) > 1 else ''
            
            project_id = row_num - 2
            project_name = ws.cell(row=row_num, column=project_name_col + 1).value or f'项目{project_id}'
            
            # 匹配真实的操作ID
            real_op_id = _find_op_id(project_id, op_type, submitter)
            
            # 解析变更内容
            after_data = {}
            if detail:
                try:
                    after_data = json.loads(detail) if isinstance(detail, str) else detail
                except:
                    pass
            
            pending_for_me.append({
                '操作ID': real_op_id or f'ROW-{project_id}',
                '操作时间': submit_time,
                '操作人': submitter,
                '操作类型': op_type,
                '项目ID列表': [project_id],
                '项目名列表': [str(project_name)],
                '变更前内容': {},
                '变更后内容': after_data,
                '状态': 'pending',
                '审批人': '',
                '审批时间': '',
            })
        
        wb.close()
    
    # === Part 2: 从操作记录读取"我发起的" ===
    ops = _load_operations_sheet()
    my_submissions = [o for o in ops if o.get('操作人') == user]
    
    return {
        'pending_for_me': pending_for_me,
        'my_submissions': my_submissions,
    }


def count_pending_approvals(user, permissions):
    """获取待我审批的数量（简化版：扫描任务计划表X列）"""
    if 'approve' not in permissions:
        return 0
    
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['任务计划表']
    
    count = 0
    for row_num in range(2, ws.max_row + 1):
        status = ws.cell(row=row_num, column=COL_APPROVAL_STATUS + 1).value or ''
        if status:
            count += 1
    
    wb.close()
    return count


def list_operations(user, permissions, filters=None):
    """获取操作记录列表（审计用）
    
    Args:
        user: 当前用户名
        permissions: 当前用户权限
        filters: dict，过滤条件
    
    Returns:
        list: 操作记录列表
    """
    ops = _load_operations_sheet()
    
    # 非admin/无audit_view权限的只能看自己的操作
    if 'audit_view' not in permissions:
        ops = [o for o in ops if o.get('操作人') == user]
    
    # 过滤
    if filters:
        if filters.get('operation_type'):
            ops = [o for o in ops if o.get('操作类型') == filters['operation_type']]
        if filters.get('status'):
            ops = [o for o in ops if o.get('状态') == filters['status']]
        if filters.get('operator'):
            ops = [o for o in ops if filters['operator'] in str(o.get('操作人', ''))]
    
    return ops[:200]  # 最多返回200条


def log_direct_operation(operation_type, project_ids, project_names, operator,
                          before_data=None, after_data=None):
    """记录直接生效的操作（editor/admin的归档/编辑/删除）
    
    Returns:
        str: 操作ID
    """
    op_data = {
        '操作时间': datetime.now().isoformat(),
        '操作人': operator,
        '操作类型': operation_type,
        '项目ID列表': project_ids,
        '项目名列表': project_names,
        '变更前内容': before_data or {},
        '变更后内容': after_data or {},
        '状态': 'direct',
    }
    return _append_operation(op_data)


# ==================== 测试 ====================

if __name__ == '__main__':
    print('=== 测试 Excel 同步模块 ===')
    print(f'Excel 文件: {EXCEL_FILE}')
    print(f'协作数据文件: {COLLAB_FILE}')
    
    projects = read_excel_projects()
    print(f'\\n读取到 {len(projects)} 条资源记录')
    
    collab = load_collab_data()
    print(f'协作数据: newProjects={len(collab.get("newProjects", []))}, '
          f'archived={len(collab.get("archived", {}))}, '
          f'localEdits={len(collab.get("localEdits", {}))}')
    
    if len(sys.argv) > 1 and sys.argv[1] == '--sync':
        print('\\n=== 执行全量同步 ===')
        ok, msg = full_sync('手动触发')
        print(f'结果: {"成功" if ok else "失败"} - {msg}')
