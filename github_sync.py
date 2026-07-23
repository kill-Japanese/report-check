# -*- coding: utf-8 -*-
"""
GitHub API 同步模块 - 当 git 命令不可用时的兜底方案

【零依赖设计】使用 Python 内置 urllib 实现 HTTP 请求，
不依赖 requests 或任何第三方库，确保在任何 Python 环境中都能工作。

功能：
1. 从 GitHub 拉取文件（替代 git pull/checkout）
2. 推送文件变更到 GitHub（替代 git add/commit/push）
3. 自动检测 git 命令是否可用，不可用时自动切换到 API 模式
"""

import os
import sys
import json
import base64
import subprocess
from datetime import datetime

# 【关键】优先用内置 urllib，不依赖 requests
try:
    from urllib.request import Request, urlopen
    from urllib.error import HTTPError, URLError
    from urllib.parse import quote as url_quote
    HAS_URLLIB = True
except ImportError:
    HAS_URLLIB = False
    def url_quote(s, safe=''):
        return s

# requests 作为可选（如果有的话），但不依赖
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GITHUB_API_BASE = 'https://api.github.com'

# GitHub 配置缓存（避免每次都读文件/环境变量）
_github_config_cache = None

# ==================== HTTP 请求封装（零依赖） ====================

def _http_request(url: str, method: str = 'GET', headers: dict = None,
                  data: bytes = None, timeout: int = 30) -> tuple[int, bytes, dict]:
    """
    通用 HTTP 请求（优先用 urllib，requests 兜底）
    返回: (status_code, response_body_bytes, headers_dict)
    """
    if headers is None:
        headers = {}
    
    # 优先用 urllib（零依赖）
    if HAS_URLLIB:
        try:
            req = Request(url, data=data, headers=headers, method=method)
            resp = urlopen(req, timeout=timeout)
            resp_body = resp.read()
            resp_headers = dict(resp.headers.items()) if hasattr(resp, 'headers') else {}
            return resp.getcode(), resp_body, resp_headers
        except HTTPError as e:
            resp_body = e.read() if hasattr(e, 'read') else b''
            resp_headers = dict(e.headers.items()) if hasattr(e, 'headers') else {}
            return e.code, resp_body, resp_headers
        except URLError as e:
            return 0, str(e.reason).encode('utf-8'), {}
        except Exception as e:
            return 0, str(e).encode('utf-8'), {}
    
    # urllib 不可用时才用 requests（极端情况）
    elif HAS_REQUESTS:
        try:
            resp = requests.request(method, url, headers=headers, data=data, timeout=timeout)
            return resp.status_code, resp.content, dict(resp.headers)
        except Exception as e:
            return 0, str(e).encode('utf-8'), {}
    
    return 0, b'No HTTP library available', {}


# ==================== 环境检测 ====================

def has_git_command() -> bool:
    """检测 git 命令是否可用"""
    try:
        result = subprocess.run(
            ['git', '--version'],
            capture_output=True, timeout=5
        )
        return result.returncode == 0
    except (FileNotFoundError, OSError):
        return False


def has_git_repo() -> bool:
    """检测 .git 目录是否存在"""
    return os.path.exists(os.path.join(BASE_DIR, '.git'))


def should_use_github_api() -> bool:
    """判断是否应该使用 GitHub API 模式（git 不可用时）"""
    return not has_git_command() or not has_git_repo()


# ==================== GitHub 配置解析 ====================

def _parse_github_url(url: str) -> tuple[str, str, str]:
    """
    从 GitHub URL 中解析 token, owner, repo
    支持格式：
    - https://token@github.com/owner/repo.git
    - https://github.com/owner/repo.git
    - git@github.com:owner/repo.git
    """
    token = ''
    owner = ''
    repo = ''
    
    url_clean = url.rstrip('.git').rstrip('/').strip()
    
    if '@github.com' in url_clean:
        # https://token@github.com/owner/repo
        token_part = url_clean.split('@github.com')[0]
        if 'https://' in token_part or 'http://' in token_part:
            auth_part = token_part.split('://')[1]
            if ':' in auth_part:
                token = auth_part.split(':')[1]
            else:
                token = auth_part
        path_part = url_clean.split('@github.com/')[1] if '@github.com/' in url_clean else url_clean.split('@github.com:')[1]
        parts = path_part.rstrip('/').split('/')
        if len(parts) >= 2:
            owner = parts[0]
            repo = '/'.join(parts[1:])
    elif 'github.com/' in url_clean:
        # https://github.com/owner/repo
        path_part = url_clean.split('github.com/')[1]
        parts = path_part.rstrip('/').split('/')
        if len(parts) >= 2:
            owner = parts[0]
            repo = '/'.join(parts[1:])
    elif 'github.com:' in url_clean:
        # git@github.com:owner/repo
        path_part = url_clean.split('github.com:')[1]
        parts = path_part.rstrip('/').split('/')
        if len(parts) >= 2:
            owner = parts[0]
            repo = '/'.join(parts[1:])
    
    return token, owner, repo


def _get_github_config() -> tuple[str, str, str]:
    """
    从多种来源获取 GitHub 认证信息（带缓存，避免每次都重读）
    
    优先级：
    1. 环境变量 GITHUB_TOKEN（最高优先级）
    2. .git/config 文件中的 remote URL（即使 git 命令不可用）
    3. git remote 命令（git 命令可用时）
    4. 默认值兜底
    
    返回: (token, owner, repo)
    """
    global _github_config_cache
    if _github_config_cache is not None:
        return _github_config_cache
    
    # 第1优先级：环境变量
    token = os.environ.get('GITHUB_TOKEN', '')
    owner = os.environ.get('GITHUB_OWNER', '')
    repo = os.environ.get('GITHUB_REPO', '')
    
    # 第2优先级：.git/config 文件（即使 git 命令不可用，只要文件存在）
    if not token or not owner or not repo:
        git_config = os.path.join(BASE_DIR, '.git', 'config')
        if os.path.exists(git_config):
            try:
                with open(git_config, 'r') as f:
                    config_content = f.read()
                # 解析 remote origin 的 URL
                in_origin = False
                for line in config_content.split('\n'):
                    line = line.strip()
                    if line == '[remote "origin"]':
                        in_origin = True
                        continue
                    if in_origin and line.startswith('['):
                        in_origin = False
                        continue
                    if in_origin and line.startswith('url ='):
                        url = line.split('url =')[1].strip()
                        t, o, r = _parse_github_url(url)
                        if t and not token:
                            token = t
                        if o and not owner:
                            owner = o
                        if r and not repo:
                            repo = r
                        break
            except Exception as e:
                print(f'[GitHub Config] 读取 .git/config 失败: {e}')
    
    # 第3优先级：git remote 命令（git 命令可用时）
    if not token or not owner or not repo:
        try:
            result = subprocess.run(
                ['git', 'remote', '-v'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=5
            )
            for line in result.stdout.split('\n'):
                if 'origin' in line and 'github.com' in line:
                    url = line.split()[1]
                    t, o, r = _parse_github_url(url)
                    if t and not token:
                        token = t
                    if o and not owner:
                        owner = o
                    if r and not repo:
                        repo = r
                    break
        except (FileNotFoundError, OSError):
            pass  # git 命令不可用
        except Exception:
            pass
    
    # 第4优先级：默认值（本项目）
    if not owner:
        owner = 'kill-Japanese'
    if not repo:
        repo = 'report-check'
    
    # 调试信息（只在首次调用时打印）
    if not token:
        print(f'[GitHub Config] ⚠️  未找到 GitHub Token，请设置 GITHUB_TOKEN 环境变量')
    
    # 缓存结果
    _github_config_cache = (token, owner, repo)
    return token, owner, repo


# ==================== GitHub API 核心操作 ====================

def github_api_get_file(path: str, branch: str = 'main') -> tuple[bool, bytes, str]:
    """
    从 GitHub 获取文件内容
    返回: (成功, 文件内容(bytes), 文件SHA)
    """
    token, owner, repo = _get_github_config()
    if not token:
        return False, b'', ''
    
    # 【关键修复】URL 路径中的中文字符需要百分号编码
    encoded_path = url_quote(path, safe='/')
    url = f'{GITHUB_API_BASE}/repos/{owner}/{repo}/contents/{encoded_path}?ref={branch}'
    headers = {
        'Authorization': f'token {token}',
        'Accept': 'application/vnd.github.v3+json',
        'User-Agent': 'report-check-server'
    }
    
    status_code, resp_body, _ = _http_request(url, 'GET', headers=headers)
    
    if status_code == 200:
        try:
            data = json.loads(resp_body.decode('utf-8'))
            content = base64.b64decode(data['content'])
            sha = data.get('sha', '')
            return True, content, sha
        except Exception as e:
            print(f'[GitHub API] 解析文件内容失败: {e}')
            return False, b'', ''
    elif status_code == 404:
        return True, b'', ''  # 文件不存在不算错误
    else:
        err_msg = resp_body.decode('utf-8', errors='replace')[:300]
        print(f'[GitHub API] 获取文件失败: HTTP {status_code} {err_msg}')
        return False, b'', ''


def github_api_push_file(path: str, content: bytes, message: str,
                         sha: str = '', branch: str = 'main') -> tuple[bool, str]:
    """
    推送文件到 GitHub（创建或更新）
    返回: (成功, 消息)
    """
    token, owner, repo = _get_github_config()
    if not token:
        return False, '缺少 GitHub Token'
    
    # 【关键修复】URL 路径中的中文字符需要百分号编码
    encoded_path = url_quote(path, safe='/')
    url = f'{GITHUB_API_BASE}/repos/{owner}/{repo}/contents/{encoded_path}'
    headers = {
        'Authorization': f'token {token}',
        'Accept': 'application/vnd.github.v3+json',
        'Content-Type': 'application/json',
        'User-Agent': 'report-check-server'
    }
    
    # 如果没有提供 SHA，尝试获取
    if not sha:
        ok, _, existing_sha = github_api_get_file(path, branch)
        if ok:
            sha = existing_sha
    
    payload = {
        'message': f'[API同步] {message} - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        'content': base64.b64encode(content).decode('utf-8'),
        'branch': branch
    }
    if sha:
        payload['sha'] = sha  # 更新文件需要 SHA
    
    data_bytes = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    status_code, resp_body, _ = _http_request(url, 'PUT', headers=headers, data=data_bytes)
    
    if status_code in (200, 201):
        return True, '已同步到 GitHub'
    else:
        err_msg = resp_body.decode('utf-8', errors='replace')[:300]
        print(f'[GitHub API] 推送文件失败: HTTP {status_code} {err_msg}')
        return False, f'API推送失败: HTTP {status_code}'


# ==================== 高级操作（替代 git pull/push） ====================

# 需要同步的关键文件
CRITICAL_FILES = [
    '用户管理.xlsx',
    '超声波户表脚本.xlsx',
]


def _get_excel_row_count(filepath: str) -> int:
    """获取Excel文件的任务计划表行数"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        ws = wb['任务计划表']
        count = ws.max_row
        wb.close()
        return count
    except:
        return 0


def github_api_pull() -> tuple[bool, str]:
    """
    从 GitHub 拉取所有关键文件（替代 git pull）
    
    【关键修复】对于Excel文件，比较行数：
    - 本地行数 > 远程行数 → 不覆盖（本地有新增数据）
    - 本地行数 <= 远程行数 → 正常同步
    """
    messages = []
    restored = []
    import tempfile
    
    for filename in CRITICAL_FILES:
        filepath = os.path.join(BASE_DIR, filename)
        existed_before = os.path.exists(filepath)
        
        ok, content, _ = github_api_get_file(filename)
        if ok and content:
            # 【关键修复】对于Excel文件，比较行数后再决定是否覆盖
            if filename.endswith('.xlsx') and existed_before:
                try:
                    local_rows = _get_excel_row_count(filepath)
                    # 写入临时文件比较远程行数
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
                        tf.write(content)
                        temp_path = tf.name
                    remote_rows = _get_excel_row_count(temp_path)
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                    
                    if local_rows > remote_rows:
                        print(f'[github_api_pull] 跳过{filename}: 本地{local_rows}行 > 远程{remote_rows}行（本地有新增数据）')
                        messages.append(f'{filename}(跳过:本地数据更新)')
                        continue
                except Exception as e:
                    print(f'[github_api_pull] 行数比较失败，正常覆盖: {e}')
            
            with open(filepath, 'wb') as f:
                f.write(content)
            if not existed_before:
                restored.append(f'{filename}(新建)')
            else:
                restored.append(f'{filename}(已同步)')
        elif not ok:
            messages.append(f'{filename}拉取失败')
    
    if restored:
        messages.append(f'已同步 {len(restored)} 个文件: {", ".join(restored)}')
    
    return True, '（拉取成功；' + '；'.join(messages) + '）' if messages else '拉取成功'


def github_api_push(message: str = '同步数据') -> tuple[bool, str]:
    """
    将所有关键文件的变更推送到 GitHub（替代 git push）
    
    【极速优化】
    1. 并行推送文件（ThreadPoolExecutor）而不是串行
    2. 先比较 SHA 跳过未变更的文件，避免不必要的网络请求
    3. 预期：从 10-20 秒降到 3-6 秒
    
    返回: (成功, 消息)
    """
    import hashlib
    import concurrent.futures
    
    # 收集所有需要推送的文件
    files_to_push = []  # (path_in_repo, local_filepath, content_bytes)
    
    for filename in CRITICAL_FILES:
        filepath = os.path.join(BASE_DIR, filename)
        if os.path.exists(filepath):
            try:
                with open(filepath, 'rb') as f:
                    content = f.read()
                files_to_push.append((filename, filepath, content))
            except Exception as e:
                pass
    
    # HTML 报表
    html_file = '项目延期点检表.html'
    html_path = os.path.join(BASE_DIR, html_file)
    if os.path.exists(html_path):
        try:
            with open(html_path, 'rb') as f:
                content = f.read()
            files_to_push.append((html_file, html_path, content))
        except:
            pass
    
    # data 目录中的文件
    data_dir = os.path.join(BASE_DIR, 'data')
    if os.path.exists(data_dir):
        for fname in os.listdir(data_dir):
            if fname.endswith('.json') or fname.endswith('.log'):
                fpath = os.path.join(data_dir, fname)
                try:
                    with open(fpath, 'rb') as f:
                        content = f.read()
                    files_to_push.append((f'data/{fname}', fpath, content))
                except:
                    pass
    
    if not files_to_push:
        return True, '无变更需要推送'
    
    # 获取配置（只用一次）
    token, owner, repo = _get_github_config()
    if not token:
        return False, '缺少 GitHub Token'
    
    # 【并行】检查文件是否有变更（先获取所有 SHA，再比较）
    def _check_and_push(item):
        """检查并推送单个文件：有变更才推送"""
        path_in_repo, local_path, content = item
        try:
            # 先获取远程 SHA
            ok, _, remote_sha = github_api_get_file(path_in_repo)
            if ok and remote_sha:
                # 比较内容（远程 SHA = git SHA，需要用 git blob 方式计算）
                local_git_sha = _git_sha1(content)
                if local_git_sha == remote_sha:
                    return (True, path_in_repo, '未变更，跳过')
            
            # 有变更，推送
            ok, msg = github_api_push_file(path_in_repo, content, message, sha=remote_sha if ok else '')
            return (ok, path_in_repo, msg if not ok else '已推送')
        except Exception as e:
            return (False, path_in_repo, str(e))
    
    pushed = []
    errors = []
    skipped = 0
    
    # 并行执行
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(_check_and_push, item) for item in files_to_push]
        for future in concurrent.futures.as_completed(futures, timeout=60):
            try:
                ok, filename, msg = future.result()
                if ok:
                    if '跳过' in msg:
                        skipped += 1
                    else:
                        pushed.append(filename)
                else:
                    errors.append(f'{filename}: {msg}')
            except:
                pass
    
    if pushed:
        msg = f'已推送 {len(pushed)} 个文件到 GitHub'
        if skipped:
            msg += f'（{skipped} 个未变更已跳过）'
        return True, msg
    elif skipped > 0:
        return True, f'无变更（{skipped} 个文件已跳过）'
    elif errors:
        return False, f'推送失败: {"; ".join(errors[:3])}'
    else:
        return True, '无变更需要推送'


def _git_sha1(content: bytes) -> str:
    """计算 Git blob SHA1（与 GitHub 返回的 SHA 一致）"""
    import hashlib
    header = f'blob {len(content)}\x00'.encode('utf-8')
    return hashlib.sha1(header + content).hexdigest()


# ==================== 统一接口（自动切换模式） ====================

def unified_pull() -> tuple[bool, str]:
    """统一拉取接口：自动选择 git 命令或 GitHub API"""
    if should_use_github_api():
        print('[同步] 使用 GitHub API 模式拉取数据')
        return github_api_pull()
    else:
        try:
            from sync_excel import git_pull
            return git_pull()
        except:
            return github_api_pull()


def unified_push(message: str = '同步数据') -> tuple[bool, str]:
    """统一推送接口：自动选择 git 命令或 GitHub API"""
    if should_use_github_api():
        print('[同步] 使用 GitHub API 模式推送数据')
        return github_api_push(message)
    else:
        try:
            from sync_excel import git_push
            return git_push(message)
        except:
            return github_api_push(message)


if __name__ == '__main__':
    print('=== GitHub API 同步模块测试（零依赖） ===')
    print(f'git 命令可用: {has_git_command()}')
    print(f'.git 目录存在: {has_git_repo()}')
    print(f'使用 API 模式: {should_use_github_api()}')
    print(f'urllib 可用: {HAS_URLLIB}')
    print(f'requests 可用: {HAS_REQUESTS}')
    
    token, owner, repo = _get_github_config()
    print(f'仓库: {owner}/{repo}')
    print(f'Token: {"***" + token[-4:] if token else "未找到"}')
    
    # 测试 HTTP 请求
    print('\n--- 测试 HTTP 请求 ---')
    status, body, _ = _http_request('https://api.github.com', 'GET')
    print(f'  GitHub API 连通性: HTTP {status}')
