# -*- coding: utf-8 -*-
"""
项目延期点检表 - 一键更新脚本

使用方法：
  python 更新点检表.py                    # 自动检测最新上传的Excel
  python 更新点检表.py 你的文件.xlsx       # 指定数据源文件
  python 更新点检表.py --send-email        # 生成后自动发邮件
  python 更新点检表.py --test-email        # 测试邮件连接

数据源文件格式要求：
  - Sheet名称: 任务计划表
  - 包含列: 部门(市场)、项目名称、项目描述、开始时间、结束时间、
           资源类型、资源名称、日平均工时
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import os
import sys
import glob
import re

# ==================== 配置区 ====================
# 输出文件名（一般不需要改）
OUTPUT_HTML = '项目延期点检表.html'
OUTPUT_EXCEL = '项目延期点检表.xlsx'
# 使用上海时区确保日期正确
try:
    from datetime import timezone, timedelta
    shanghai_tz = timezone(timedelta(hours=8))
    TODAY = pd.Timestamp(datetime.now(shanghai_tz)).tz_localize(None).normalize()
except:
    TODAY = pd.Timestamp.now().normalize()

# 数据源文件（优先级：命令行参数 > 自动检测 > 手动指定）
# 如需手动指定，取消下一行注释并修改路径：
# SOURCE_FILE = '你的项目表.xlsx'
# ================================================


def find_source_file():
    """自动查找数据源文件：优先级从高到低
    1. 命令行参数
    2. data_source.json（Skill 记录的数据源）
    3. 当前目录下的 超声波户表脚本.xlsx（GitHub 源文件，数据之源）
    4. .uploads 目录下最新的"超声波户表脚本"文件
    5. .uploads 目录下最新的 xlsx 文件
    6. 当前目录下最新的 xlsx 文件
    """
    import json
    
    # 1. 命令行参数（如果指定了但不存在，直接报错）
    for arg in sys.argv[1:]:
        if arg.endswith('.xlsx') and not arg.startswith('--'):
            if os.path.exists(arg):
                return arg, '命令行指定'
            else:
                print(f"❌ 命令行指定的文件不存在: {arg}")
                print()
                sys.exit(1)
    
    # 2. 读取 data_source.json（Skill 记录的数据源，最高优先级）
    data_source_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.trae', 'data_source.json')
    if os.path.exists(data_source_file):
        try:
            with open(data_source_file, 'r', encoding='utf-8') as f:
                info = json.load(f)
            recorded_path = info.get('file_path', '')
            recorded_name = info.get('file_name', '')
            if recorded_path and os.path.exists(recorded_path):
                return recorded_path, f'Skill记录({recorded_name})'
        except Exception:
            pass  # 读取失败则忽略，继续往下找
    
    # 3. 当前目录下的 超声波户表脚本.xlsx（GitHub 源文件，最高优先级）
    github_source = os.path.join(os.path.dirname(os.path.abspath(__file__)), '超声波户表脚本.xlsx')
    if os.path.exists(github_source):
        return github_source, 'GitHub源文件(超声波户表脚本.xlsx)'
    
    # 4. .uploads 目录下的 xlsx 文件
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.uploads')
    if os.path.isdir(upload_dir):
        xlsx_files = glob.glob(os.path.join(upload_dir, '*.xlsx'))
        if xlsx_files:
            # 优先选择文件名包含"超声波户表脚本"的文件（按修改时间排序取最新）
            script_files = [f for f in xlsx_files if '超声波户表脚本' in os.path.basename(f)]
            if script_files:
                latest = max(script_files, key=os.path.getmtime)
                return latest, '自动检测(超声波户表脚本最新)'
            # 没有的话再按修改时间取最新
            latest = max(xlsx_files, key=os.path.getmtime)
            return latest, '自动检测(.uploads最新)'
    
    # 5. 当前目录下的 xlsx 文件（排除输出文件）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_files = glob.glob(os.path.join(current_dir, '*.xlsx'))
    xlsx_files = [f for f in xlsx_files if os.path.basename(f) != OUTPUT_EXCEL]
    if xlsx_files:
        latest = max(xlsx_files, key=os.path.getmtime)
        return latest, '自动检测(当前目录最新)'
    
    return None, None


# ==================== 1. 读取数据 ====================
# 优先使用脚本顶部手动指定的，否则自动检测
_manual_source = None  # 如果在上方配置区设置了SOURCE_FILE，会覆盖这里

try:
    SOURCE_FILE
except NameError:
    SOURCE_FILE = None

if SOURCE_FILE is None:
    SOURCE_FILE, source_method = find_source_file()
else:
    source_method = '手动指定'

if not SOURCE_FILE or not os.path.exists(SOURCE_FILE):
    print("❌ 找不到数据源Excel文件！")
    print()
    print("请用以下任一方式指定数据源：")
    print("  方式1：命令行指定  →  python 更新点检表.py 你的文件.xlsx")
    print("  方式2：把Excel文件上传后放在当前目录下")
    print("  方式3：在脚本第40行左右手动设置 SOURCE_FILE = '路径/文件名.xlsx'")
    print()
    print("当前目录下的xlsx文件：")
    for f in glob.glob('*.xlsx'):
        print(f"  - {f}")
    sys.exit(1)

print(f"📖 数据源: {SOURCE_FILE}")
print(f"📌 检测方式: {source_method}")

df = pd.read_excel(SOURCE_FILE, sheet_name='任务计划表', header=None)
today = TODAY.to_pydatetime()

# ==================== 2. 提取项目数据 ====================
projects = []
current_dept = None
current_project = None
current_start = None
current_end = None
current_desc = None

for idx in range(3, len(df)):
    row = df.iloc[idx]
    
    if pd.notna(row[4]):
        val = row[4]
        current_dept = str(val) if not isinstance(val, float) else val
    
    if pd.notna(row[5]):
        current_project = str(row[5])
        current_start = row[6] if pd.notna(row[6]) else None
        current_end = row[7] if pd.notna(row[7]) else None
        current_desc = str(row[8]) if pd.notna(row[8]) else ''
    
    resource_type = str(row[9]) if pd.notna(row[9]) else ''
    resource_name = str(row[10]) if pd.notna(row[10]) else ''
    # 解析资源名称: 支持 "@姓名(工号)(姓名)" 或 "@姓名(工号)" 格式，提取纯姓名
    def clean_name(name):
        if not name:
            return name
        name = name.strip()
        # 去掉开头的 @ 或空格 @
        if name.startswith('@'):
            name = name[1:]
        elif name.startswith(' @'):
            name = name[2:]
        name = name.strip()
        # 提取括号前的部分 (姓名)
        if '(' in name:
            name = name[:name.index('(')].strip()
        return name
    
    resource_name = clean_name(resource_name)
    res_start = row[11] if pd.notna(row[11]) else None
    res_end = row[12] if pd.notna(row[12]) else None
    hours = row[13] if pd.notna(row[13]) else None
    
    if current_project:
        def fmt_date(d, is_start=True):
            if d is None or (isinstance(d, float) and pd.isna(d)):
                return ''
            # 通配符处理: / 在开始代表1900-01-01，在结束代表2100-01-01
            if d == '/' or str(d).strip() == '/':
                return '1900-01-01' if is_start else '2100-01-01'
            if isinstance(d, pd.Timestamp):
                return d.strftime('%Y-%m-%d')
            if isinstance(d, datetime):
                return d.strftime('%Y-%m-%d')
            return str(d)
        
        def calc_days(d, is_end=True):
            if d is None or (isinstance(d, float) and pd.isna(d)):
                return None
            # 通配符处理: / 在结束代表2100-01-01（只对结束时间计算剩余天数）
            if d == '/' or str(d).strip() == '/':
                if is_end:
                    return (datetime(2100, 1, 1) - today).days
                return None
            if isinstance(d, pd.Timestamp):
                d = d.to_pydatetime()
            if isinstance(d, datetime):
                return (d - today).days
            return None
        
        proj_days = calc_days(current_end, True)
        res_days = calc_days(res_end, True)
        
        # 【严重修复】读取归档标志（第1列，A列）—— 原使用U列(20)会覆盖延期计算公式
        archived_flag = ''
        if 0 < len(row) and pd.notna(row[0]):
            archived_flag = str(row[0]).strip()
        is_archived = archived_flag in ('已归档', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y')
        
        # 【严重修复】读取删除标志（第2列，B列）—— 原使用V列(21)会覆盖延期计算公式
        # 软删除，已删除的项目不加入列表
        deleted_flag = ''
        if 1 < len(row) and pd.notna(row[1]):
            deleted_flag = str(row[1]).strip()
        is_deleted = deleted_flag in ('已删除', '1', 'true', 'True', 'YES', 'yes', 'Y', 'y')
        if is_deleted:
            continue  # 跳过已软删除的项目
        
        # 跳过空行：没有资源类型和资源名称的行不处理
        has_resource = resource_type.strip() or resource_name.strip()
        if has_resource:
            projects.append({
                'id': idx,
                '部门': current_dept if current_dept and not (isinstance(current_dept, float) and pd.isna(current_dept)) else '',
                '项目': current_project,
            '项目开始时间': fmt_date(current_start, True),
            '项目结束时间': fmt_date(current_end, False),
            '项目剩余天数': proj_days,
            '项目描述': current_desc,
            '资源类型': resource_type,
            '资源名称': resource_name,
            '资源开始时间': fmt_date(res_start, True),
            '资源结束时间': fmt_date(res_end, False),
            '资源剩余天数': res_days,
                '日平均工时': hours if hours else 0,
                '已归档': is_archived,
            })

print(f"✅ 提取到 {len(projects)} 条资源记录")

# ==================== 2.5 加载协作数据（新增/删除/归档）====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
COLLAB_FILE = os.path.join(BASE_DIR, 'data', '协作数据.json')
collab_data = {
    'newProjects': [], 'deletedIds': [], 'archived': {},
    'localEdits': {}, 'notes': {}, 'checked': {}, 'customEmails': {}
}
if os.path.exists(COLLAB_FILE):
    try:
        with open(COLLAB_FILE, 'r', encoding='utf-8') as f:
            collab_data.update(json.load(f))
        print(f"📂 已加载协作数据: 新增{len(collab_data.get('newProjects', []))}条, "
              f"删除{len(collab_data.get('deletedIds', []))}条, "
              f"归档{len(collab_data.get('archived', {}))}条")
    except Exception as e:
        print(f"⚠️  加载协作数据失败: {e}")

# 应用删除
deleted_ids = set(collab_data.get('deletedIds', []))
if deleted_ids:
    before = len(projects)
    projects = [p for p in projects if str(p['id']) not in deleted_ids and p['id'] not in deleted_ids]
    print(f"🗑️  已过滤删除的项目: {before - len(projects)} 条")

# 加载新增项目（从协作数据 + html_new_projects.json）
NEW_PROJECTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'html_new_projects.json')
all_new_projects = list(collab_data.get('newProjects', []))

if os.path.exists(NEW_PROJECTS_FILE):
    try:
        with open(NEW_PROJECTS_FILE, 'r', encoding='utf-8') as f:
            html_new = json.load(f)
        if isinstance(html_new, list):
            existing_ids = {p.get('id') for p in all_new_projects}
            for np in html_new:
                if np.get('id') not in existing_ids:
                    all_new_projects.append(np)
    except Exception as e:
        print(f"⚠️  加载HTML新增项目失败: {e}")

if all_new_projects:
    new_id_start = max([p['id'] for p in projects]) + 1 if projects else 10000
    for np in all_new_projects:
        if np.get('id') is None or not any(p['id'] == np.get('id') for p in projects):
            np['id'] = new_id_start
            # 【修复】先设置默认值，确保字段始终存在（防止KeyError）
            np['资源剩余天数'] = None
            np['项目剩余天数'] = None
            # 重新计算剩余天数
            if np.get('资源结束时间'):
                try:
                    res_end = datetime.strptime(np['资源结束时间'], '%Y-%m-%d')
                    np['资源剩余天数'] = (res_end - today).days
                except:
                    np['资源剩余天数'] = None
            if np.get('项目结束时间') and str(np['项目结束时间']) not in ['2100-01-01', '']:
                try:
                    proj_end = datetime.strptime(str(np['项目结束时间']), '%Y-%m-%d')
                    np['项目剩余天数'] = (proj_end - today).days
                except:
                    np['项目剩余天数'] = None
            # 确保其他必要字段也存在
            for field in ['部门', '项目', '项目描述', '资源类型', '资源名称', '日平均工时', '已归档',
                          '项目开始时间', '项目结束时间', '资源开始时间', '资源结束时间']:
                if field not in np:
                    if field == '日平均工时':
                        np[field] = 0
                    elif field == '已归档':
                        np[field] = False
                    elif field == '项目开始时间' or field == '资源开始时间':
                        np[field] = '1900-01-01'
                    elif field == '项目结束时间' or field == '资源结束时间':
                        np[field] = '2100-01-01'
                    else:
                        np[field] = ''
            projects.append(np)
            new_id_start += 1
    print(f"📥 已加载 {len(all_new_projects)} 条新增项目")

# ==================== 2.6 全局字段完整性检查 ====================
# 确保所有项目都有必要的字段，防止后续 KeyError
REQUIRED_FIELDS = [
    'id', '部门', '项目', '项目开始时间', '项目结束时间',
    '项目剩余天数', '项目描述', '资源类型', '资源名称',
    '资源开始时间', '资源结束时间', '资源剩余天数',
    '日平均工时', '已归档'
]
for p in projects:
    for field in REQUIRED_FIELDS:
        if field not in p:
            if field == '日平均工时':
                p[field] = 0
            elif field == '已归档':
                p[field] = False
            elif field == '项目剩余天数' or field == '资源剩余天数':
                p[field] = None
            elif field == '项目开始时间' or field == '资源开始时间':
                p[field] = '1900-01-01'
            elif field == '项目结束时间' or field == '资源结束时间':
                p[field] = '2100-01-01'
            else:
                p[field] = ''

# ==================== 3. 分类统计（排除已归档项目）====================
# 过滤出未归档的项目用于统计
active_projects = [p for p in projects if not p.get('已归档', False)]

# 统一规则：优先使用资源结束时间，没有则使用项目结束时间（与前端 recalcDays 保持一致）
def get_effective_days(p):
    return p['资源剩余天数'] if p['资源剩余天数'] is not None else p['项目剩余天数']

delayed_projects = []
for p in active_projects:
    days = get_effective_days(p)
    if days is not None and days <= 3:
        delayed_projects.append(p)

delayed_projects.sort(key=lambda x: (
    get_effective_days(x) if get_effective_days(x) is not None else 999,
))

depts = {}
for p in active_projects:
    d = p['部门'] or '未分配'
    depts.setdefault(d, []).append(p)

stats = {
    'total': len(active_projects),
    'delayed': len([p for p in active_projects if get_effective_days(p) is not None and get_effective_days(p) < 0]),
    'urgent': len([p for p in active_projects if get_effective_days(p) is not None and 0 <= get_effective_days(p) <= 3]),
    'warning': len([p for p in active_projects if get_effective_days(p) is not None and 3 < get_effective_days(p) <= 7]),
}

data_json = {
    'today': today.strftime('%Y-%m-%d'),
    'threeDaysLater': (today + timedelta(days=3)).strftime('%Y-%m-%d'),
    'stats': stats,
    'allProjects': projects,
    'delayedProjects': delayed_projects,
    'depts': depts,
    'sourceFile': os.path.basename(SOURCE_FILE),
    'sourceMethod': source_method
}

# 清理显示用的文件名（去掉上传时的哈希前缀，如 711e68fe-..._户表7-3日工作.xlsx → 户表7-3日工作.xlsx）
display_source_name = os.path.basename(SOURCE_FILE)
_clean_match = re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_(.+)$', display_source_name, re.IGNORECASE)
if _clean_match:
    display_source_name = _clean_match.group(1)
data_json['displaySourceName'] = display_source_name

print(f"⚠️  延期预警: {len(delayed_projects)} 条")
print(f"   - 已延期: {len([p for p in delayed_projects if get_effective_days(p) is not None and get_effective_days(p) < 0])} 条")
print(f"   - 3天内到期: {len([p for p in delayed_projects if get_effective_days(p) is not None and 0 <= get_effective_days(p) <= 3])} 条")

# ==================== 4. 生成HTML ====================
print(f"\n🌐 生成HTML: {OUTPUT_HTML}")

html_template = '''<!-- Generated by Trae Work -->
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>项目延期点检表 - 三天后到期项目</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    font-family: 'Noto Sans CJK SC', 'Microsoft YaHei', Arial, sans-serif;
    background: #f0f2f5;
    color: #1f2937;
    padding: 20px;
  }
  .container { max-width: 1400px; margin: 0 auto; }
  
  .header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 30px;
    border-radius: 12px;
    margin-bottom: 20px;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
  }
  .header h1 { font-size: 24px; margin-bottom: 8px; }
  .header .subtitle { opacity: 0.9; font-size: 14px; }
  .header .date-info { margin-top: 12px; display: flex; gap: 20px; flex-wrap: wrap; }
  .header .date-info span {
    background: rgba(255,255,255,0.2);
    padding: 6px 12px;
    border-radius: 20px;
    font-size: 13px;
  }
  .update-tip {
    background: #ECFDF5;
    border: 1px solid #6EE7B7;
    color: #065F46;
    padding: 12px 16px;
    border-radius: 8px;
    margin-bottom: 16px;
    font-size: 13px;
  }
  .update-tip code {
    background: #065F46;
    color: #ECFDF5;
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 12px;
  }

  .section-title {
    font-size: 18px;
    font-weight: 600;
    color: #1f2937;
    margin: 20px 0 12px 0;
    padding-left: 12px;
    border-left: 4px solid #667eea;
  }

  .stats-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 20px;
  }
  .stat-card {
    background: white;
    padding: 20px;
    border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    text-align: center;
  }
  .stat-card .number { font-size: 32px; font-weight: 700; margin-bottom: 6px; }
  .stat-card .label { color: #6b7280; font-size: 13px; }
  .stat-card.delayed .number { color: #ef4444; }
  .stat-card.urgent .number { color: #f97316; }
  .stat-card.warning .number { color: #eab308; }
  .stat-card.normal .number { color: #22c55e; }

  .toolbar {
    background: white;
    padding: 16px;
    border-radius: 10px;
    margin-bottom: 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    display: flex;
    gap: 12px;
    flex-wrap: wrap;
    align-items: center;
  }
  .toolbar input, .toolbar select {
    padding: 8px 12px;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    font-size: 13px;
    font-family: inherit;
  }
  .toolbar input:focus, .toolbar select:focus { outline: none; border-color: #667eea; }
  .toolbar .search { flex: 1; min-width: 200px; }
  .btn {
    padding: 8px 16px;
    border: none;
    border-radius: 6px;
    cursor: pointer;
    font-size: 13px;
    font-family: inherit;
    transition: all 0.2s;
  }
  .btn-primary { background: #667eea; color: white; }
  .btn-primary:hover { background: #5a67d8; }
  .btn-success { background: #22c55e; color: white; }
  .btn-success:hover { background: #16a34a; }
  .btn-secondary { background: #e5e7eb; color: #374151; }
  .btn-secondary:hover { background: #d1d5db; }
  .btn-warning { background: #f59e0b; color: white; }
  .btn-warning:hover { background: #d97706; }

  .tabs {
    display: flex;
    gap: 4px;
    margin-bottom: 16px;
    background: white;
    padding: 6px;
    border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    flex-wrap: wrap;
  }
  .tab {
    padding: 10px 20px;
    border-radius: 8px;
    cursor: pointer;
    font-size: 13px;
    transition: all 0.2s;
    white-space: nowrap;
  }
  .tab:hover { background: #f3f4f6; }
  .tab.active { background: #667eea; color: white; }
  .tab .badge {
    background: rgba(0,0,0,0.1);
    padding: 2px 8px;
    border-radius: 10px;
    font-size: 11px;
    margin-left: 6px;
  }
  .tab.active .badge { background: rgba(255,255,255,0.25); }

  table {
    width: 100%;
    border-collapse: collapse;
    background: white;
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }
  thead {
    background: #f9fafb;
    border-bottom: 2px solid #e5e7eb;
  }
  th {
    padding: 12px;
    text-align: left;
    font-size: 12px;
    font-weight: 600;
    color: #374151;
    white-space: nowrap;
  }
  td {
    padding: 10px 12px;
    font-size: 13px;
    border-bottom: 1px solid #f3f4f6;
    vertical-align: middle;
  }
  tr:hover { background: #f9fafb; }
  tr.delayed-row { background: #fef2f2; }
  tr.delayed-row:hover { background: #fee2e2; }
  tr.urgent-row { background: #fff7ed; }
  tr.urgent-row:hover { background: #ffedd5; }

  .status-tag {
    display: inline-flex;
    align-items: center;
    gap: 4px;
    padding: 4px 10px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 500;
  }
  .status-delayed { background: #fee2e2; color: #dc2626; }
  .status-urgent { background: #ffedd5; color: #ea580c; }
  .status-warning { background: #fef3c7; color: #d97706; }
  .status-normal { background: #dcfce7; color: #16a34a; }
  .status-none { background: #f3f4f6; color: #6b7280; }

  .editable {
    cursor: text;
    padding: 4px 8px;
    border-radius: 4px;
    transition: all 0.2s;
    border: 1px solid transparent;
  }
  .editable:hover { background: #eef2ff; border-color: #c7d2fe; }
  .editable:focus {
    outline: none;
    background: white;
    border-color: #667eea;
    box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
  }
  .editable.date-input {
    font-family: inherit;
    font-size: 13px;
    width: 120px;
  }
  .note-cell { min-width: 150px; max-width: 300px; }
  .note-text { color: #6b7280; font-style: italic; }

  .dept-tag {
    display: inline-block;
    padding: 2px 8px;
    background: #eef2ff;
    color: #4f46e5;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 500;
  }

  .save-indicator {
    position: fixed;
    bottom: 20px;
    right: 20px;
    background: #22c55e;
    color: white;
    padding: 10px 20px;
    border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    display: none;
    z-index: 1000;
  }
  .save-indicator.show { display: block; animation: fadeInOut 2s; }
  @keyframes fadeInOut {
    0% { opacity: 0; transform: translateY(10px); }
    15% { opacity: 1; transform: translateY(0); }
    85% { opacity: 1; transform: translateY(0); }
    100% { opacity: 0; transform: translateY(-10px); }
  }

  .edit-mode-bar {
    background: #FEF3C7;
    border: 1px solid #FCD34D;
    padding: 10px 16px;
    border-radius: 8px;
    margin-bottom: 12px;
    font-size: 13px;
    color: #92400E;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .edit-mode-bar.active { background: #DBEAFE; border-color: #93C5FD; color: #1E40AF; }

  .empty-state {
    text-align: center;
    padding: 60px 20px;
    color: #9ca3af;
  }
  .empty-state .icon { font-size: 48px; margin-bottom: 16px; }

  .archive-btn {
    background: none;
    border: 1px solid #d1d5db;
    border-radius: 4px;
    padding: 4px 8px;
    cursor: pointer;
    font-size: 12px;
    color: #6b7280;
    transition: all 0.2s;
    white-space: nowrap;
  }
  .archive-btn:hover {
    background: #fef3c7;
    border-color: #f59e0b;
    color: #92400e;
  }
  .archive-btn.restore {
    color: #059669;
    border-color: #6ee7b7;
  }
  .archive-btn.restore:hover {
    background: #d1fae5;
    border-color: #10b981;
    color: #065f46;
  }
  tr.archived-row {
    opacity: 0.6;
    background: #f9fafb !important;
  }
  /* 看板视图样式 */
  .board-container {
    display: flex;
    flex-direction: column;
    gap: 24px;
  }
  .board-market-group {
    background: white;
    border-radius: 12px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    overflow: hidden;
  }
  .board-market-header {
    background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%);
    color: white;
    padding: 14px 20px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .board-market-title {
    font-size: 16px;
    font-weight: 700;
  }
  .board-market-stats {
    font-size: 12px;
    opacity: 0.9;
    background: rgba(255,255,255,0.2);
    padding: 4px 10px;
    border-radius: 12px;
  }
  .board-project-list {
    padding: 16px;
    display: flex;
    flex-direction: column;
    gap: 16px;
  }
  .board-project-group {
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    overflow: hidden;
  }
  .board-project-name-group {
    border: 2px solid #6366f1;
    border-radius: 12px;
    overflow: hidden;
    background: #fafbff;
  }
  .board-project-name-header {
    background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
    padding: 12px 16px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    cursor: pointer;
  }
  .board-project-name-header:hover {
    background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
  }
  .board-project-name-header .board-project-title {
    color: white;
    font-size: 15px;
  }
  .board-project-name-header .board-project-meta span {
    color: rgba(255,255,255,0.9);
  }
  .board-project-name-header .dept-tag {
    background: rgba(255,255,255,0.2);
    color: white;
  }
  .board-tr-container {
    padding: 12px;
    display: flex;
    flex-direction: column;
    gap: 12px;
  }
  .board-tr-group {
    border: 1px solid #c7d2fe;
    border-radius: 8px;
    overflow: hidden;
    background: white;
  }
  .board-tr-header {
    background: #eef2ff;
    padding: 8px 14px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1px solid #c7d2fe;
    cursor: pointer;
  }
  .board-tr-header:hover {
    background: #e0e7ff;
  }
  .board-tr-title {
    font-weight: 600;
    color: #4338ca;
    font-size: 13px;
  }
  .board-project-header {
    background: #f8fafc;
    padding: 10px 16px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1px solid #e5e7eb;
    cursor: pointer;
  }
  .board-project-header:hover {
    background: #f1f5f9;
  }
  .board-project-title {
    font-weight: 600;
    color: #1e293b;
    font-size: 14px;
  }
  .board-project-desc {
    font-size: 12px;
    color: #64748b;
    margin-top: 2px;
  }
  .board-project-meta {
    display: flex;
    gap: 8px;
    align-items: center;
  }
  .board-resource-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
  }
  .board-resource-table th {
    background: #fafafa;
    padding: 8px 12px;
    text-align: left;
    font-weight: 600;
    color: #475569;
    border-bottom: 1px solid #e5e7eb;
    font-size: 11px;
  }
  .board-resource-table td {
    padding: 8px 12px;
    border-bottom: 1px solid #f1f5f9;
  }
  .board-resource-table tr:last-child td {
    border-bottom: none;
  }
  .board-resource-table tr:hover {
    background: #fafafa;
  }
  .board-time-range {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: #eff6ff;
    color: #1d4ed8;
    padding: 3px 8px;
    border-radius: 6px;
    font-size: 11px;
    font-weight: 500;
  }
  .board-empty {
    text-align: center;
    padding: 40px;
    color: #94a3b8;
    font-size: 14px;
  }

  .archived-tag {
    display: inline-block;
    background: #e5e7eb;
    color: #6b7280;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 11px;
    margin-left: 6px;
  }

  .modal-overlay {
    position: fixed; top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(0,0,0,0.5); z-index: 10000;
    display: flex; align-items: center; justify-content: center; padding: 20px;
  }
  .modal-box {
    background: white; border-radius: 12px; max-width: 600px; width: 100%;
    max-height: 85vh; overflow: auto; padding: 24px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
  }
  .modal-box h3 { margin: 0 0 16px 0; color: #1e293b; }
  .modal-close {
    float: right; background: none; border: none; font-size: 24px;
    cursor: pointer; color: #64748b; line-height: 1;
  }
  .form-group { margin-bottom: 14px; }
  .form-group label {
    display: block; font-size: 13px; font-weight: 600; color: #374151;
    margin-bottom: 6px;
  }
  .form-group input, .form-group select, .form-group textarea {
    width: 100%; padding: 8px 12px; border: 1px solid #d1d5db;
    border-radius: 6px; font-size: 14px; font-family: inherit;
    box-sizing: border-box;
  }
  .form-group input:focus, .form-group select:focus, .form-group textarea:focus {
    outline: none; border-color: #667eea;
  }
  .form-row { display: flex; gap: 12px; }
  .form-row .form-group { flex: 1; }
  .form-hint { font-size: 12px; color: #6b7280; margin-top: 4px; }
  .form-actions {
    display: flex; gap: 8px; justify-content: flex-end; margin-top: 20px;
    padding-top: 16px; border-top: 1px solid #e5e7eb;
  }
  .member-list { max-height: 300px; overflow-y: auto; border: 1px solid #e5e7eb; border-radius: 8px; }
  .member-item {
    display: flex; align-items: center; justify-content: space-between;
    padding: 10px 14px; border-bottom: 1px solid #f3f4f6;
  }
  .member-item:last-child { border-bottom: none; }
  .member-item .member-info { flex: 1; }
  .member-item .member-name { font-weight: 600; color: #1e293b; }
  .member-item .member-email { font-size: 12px; color: #6b7280; }
  .member-item .member-source { font-size: 11px; color: #9ca3af; margin-left: 8px; }
  .tag-builtin { background: #e0e7ff; color: #4338ca; padding: 1px 6px; border-radius: 4px; font-size: 10px; }
  .tag-custom { background: #d1fae5; color: #065f46; padding: 1px 6px; border-radius: 4px; font-size: 10px; }
  .email-suggest {
    background: #f0f9ff; border: 1px dashed #0ea5e9; color: #0369a1;
    padding: 6px 10px; border-radius: 6px; font-size: 12px; margin-top: 6px;
    cursor: pointer;
  }
  .email-suggest:hover { background: #e0f2fe; }

  .hours-panel {
    background: white;
    border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    margin-bottom: 16px;
    overflow: hidden;
  }
  .hours-panel-header {
    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
    color: white;
    padding: 12px 16px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .hours-panel-header span:first-child {
    font-weight: 600;
    font-size: 14px;
  }
  .hours-panel-hint {
    font-size: 12px;
    opacity: 0.9;
  }
  .hours-panel-body {
    padding: 16px;
  }
  .hours-row {
    display: flex;
    align-items: center;
    gap: 8px;
    flex-wrap: wrap;
    margin-bottom: 16px;
  }
  .hours-row label {
    font-size: 13px;
    color: #374151;
    font-weight: 500;
  }
  .hours-row input, .hours-row select {
    padding: 6px 10px;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    font-size: 13px;
    font-family: inherit;
  }
  .hours-results {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 12px;
  }
  .hours-stat {
    background: #F0FDF4;
    border: 1px solid #BBF7D0;
    border-radius: 8px;
    padding: 14px;
    text-align: center;
  }
  .hours-label {
    font-size: 12px;
    color: #065F46;
    margin-bottom: 6px;
  }
  .hours-value {
    font-size: 24px;
    font-weight: 700;
    color: #047857;
  }

  .resource-panel {
    background: white;
    border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    margin-bottom: 16px;
    overflow: hidden;
  }
  .resource-panel-header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 12px 16px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .resource-panel-header span:first-child {
    font-weight: 600;
    font-size: 14px;
  }
  .resource-panel-body {
    padding: 16px;
  }
  .resource-input-row {
    display: flex;
    align-items: center;
    gap: 12px;
    flex-wrap: wrap;
    margin-bottom: 16px;
  }
  .resource-input-row label {
    font-size: 13px;
    color: #374151;
    font-weight: 500;
  }
  .resource-input-row input, .resource-input-row select {
    padding: 8px 12px;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    font-size: 13px;
    font-family: inherit;
  }
  .resource-input-row input:focus, .resource-input-row select:focus {
    outline: none;
    border-color: #667eea;
  }
  .resource-summary {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
    gap: 12px;
    margin-bottom: 16px;
  }
  .resource-stat {
    background: #EEF2FF;
    border: 1px solid #C7D2FE;
    border-radius: 8px;
    padding: 12px;
    text-align: center;
  }
  .resource-stat-label {
    font-size: 12px;
    color: #4338CA;
    margin-bottom: 4px;
  }
  .resource-stat-value {
    font-size: 20px;
    font-weight: 700;
    color: #4F46E5;
  }
  .resource-detail-title {
    font-size: 13px;
    font-weight: 600;
    color: #374151;
    margin-bottom: 8px;
    padding-bottom: 8px;
    border-bottom: 1px solid #e5e7eb;
  }
  .resource-detail-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
  }
  .resource-detail-table th {
    background: #F9FAFB;
    padding: 8px 10px;
    text-align: left;
    font-weight: 600;
    color: #374151;
    border-bottom: 2px solid #e5e7eb;
    white-space: nowrap;
  }
  .resource-detail-table td {
    padding: 8px 10px;
    border-bottom: 1px solid #f3f4f6;
  }
  .resource-detail-table tr:hover {
    background: #f9fafb;
  }
  .overlap-badge {
    display: inline-block;
    background: #DBEAFE;
    color: #1E40AF;
    padding: 2px 8px;
    border-radius: 10px;
    font-size: 11px;
    font-weight: 500;
  }
  .calc-detail {
    background: #F9FAFB;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 10px 12px;
    margin-top: 8px;
    font-size: 11px;
    color: #6b7280;
    font-family: 'Consolas', monospace;
    line-height: 1.6;
  }

  @media (max-width: 768px) {
    body { padding: 10px; }
    .header { padding: 20px; }
    .header h1 { font-size: 18px; }
    table { font-size: 12px; }
    th, td { padding: 8px; }
  }
</style>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>📋 项目延期点检表</h1>
    <div class="subtitle">自动检测三天内即将到期或已延期的项目与任务资源</div>
    <div class="date-info">
      <span>📅 点检日期：<strong id="todayDate"></strong></span>
      <span>⏰ 预警截止：<strong id="alertDate"></strong></span>
      <span>📝 数据来源：<span id="sourceFileName">__SOURCE_FILE__</span></span>
    </div>
  </div>

  <div class="update-tip">
    💡 <strong>更新计划方式：</strong>
    方式A：修改原始Excel后运行 <code>python 更新点检表.py</code> 重新生成
    |
    方式B：点击下方「✏️ 编辑模式」直接在页面修改（协作模式下实时同步到服务器）
  </div>

  <div class="section-title">📊 统计看板</div>
  <div class="stats-row">
    <div class="stat-card delayed">
      <div class="number" id="statDelayed">0</div>
      <div class="label">已延期</div>
    </div>
    <div class="stat-card urgent">
      <div class="number" id="statUrgent">0</div>
      <div class="label">3天内到期</div>
    </div>
    <div class="stat-card warning">
      <div class="number" id="statWarning">0</div>
      <div class="label">7天内到期</div>
    </div>
    <div class="stat-card normal">
      <div class="number" id="statTotal">0</div>
      <div class="label">总资源数</div>
    </div>
  </div>

  <div class="toolbar">
    <input type="text" class="search" id="searchInput" placeholder="🔍 搜索项目名称、部门、资源...">
    <select id="deptFilter">
      <option value="">全部部门</option>
    </select>
    <select id="statusFilter">
      <option value="">全部状态</option>
      <option value="delayed">🔴 已延期</option>
      <option value="urgent">🟠 3天内到期</option>
      <option value="warning">🟡 7天内到期</option>
      <option value="normal">🟢 正常</option>
    </select>
    <button class="btn btn-warning" id="editModeBtn" onclick="toggleEditMode()">✏️ 编辑模式</button>
    <button class="btn btn-secondary" onclick="exportCSV()">📥 导出CSV</button>
    <button class="btn" style="background:#8b5cf6;color:white" onclick="exportShareData()">📤 导出共享数据</button>
    <button class="btn" style="background:#06b6d4;color:white" onclick="document.getElementById('importShareFile').click()">📥 导入共享数据</button>
    <button class="btn" style="background:#10b981;color:white" onclick="generateReportEmail()">📧 整理报表邮件</button>
    <input type="file" id="importShareFile" accept=".json" style="display:none" onchange="importShareData(event)">
    <button class="btn btn-success" onclick="saveData()">💾 保存修改</button>
    <button class="btn btn-secondary" onclick="resetAll()">🔄 重置</button>
    <button class="btn" style="background:#667eea;color:white" onclick="generateAllEmails()">📧 延期催办邮件生成</button>
    <button class="btn" style="background:#f59e0b;color:white" onclick="openEmailClient()">📨 邮件客户端发送</button>
    <button class="btn" style="background:#10b981;color:white" onclick="sendAllEmails()">📤 自动发送</button>
    <button class="btn" style="background:#ec4899;color:white" onclick="openAddProjectModal()">➕ 添加项目</button>
    <button class="btn" style="background:#8b5cf6;color:white" onclick="openImportModal()">📥 从Project导入</button>
    <button class="btn" style="background:#14b8a6;color:white" onclick="openMemberManager()">👥 成员管理</button>
    <button class="btn" style="background:#f97316;color:white" onclick="syncProjectsToExcel()">🔄 同步到Excel</button>
    <button class="btn" style="background:#6366f1;color:white" onclick="selectAllVisible()">☑️ 全选当前</button>
    <button class="btn" style="background:#8b5cf6;color:white" onclick="clearSelection()">⬜ 清空选择</button>
    <button class="btn" style="background:#10b981;color:white" onclick="batchArchive()">📦 批量归档</button>
    <button class="btn" style="background:#06b6d4;color:white" onclick="batchRestore()">↩️ 批量恢复</button>
    <button class="btn" style="background:#ef4444;color:white" onclick="batchDelete()">🗑️ 批量删除</button>
    <span id="selectionCount" style="color:#6b7280;font-size:13px;margin-left:8px"></span>
  </div>

  <div class="hours-panel" id="hoursPanel">
    <div class="hours-panel-header">
      <span>⏱️ 工时统计</span>
      <span class="hours-panel-hint">选择时间段，自动计算该范围内的平均工时投入</span>
    </div>
    <div class="hours-panel-body">
      <div class="hours-row">
        <label>统计范围：</label>
        <select id="hoursRangePreset" onchange="applyHoursPreset()">
          <option value="week">本周</option>
          <option value="nextweek">下周</option>
          <option value="month">本月</option>
          <option value="custom">自定义</option>
        </select>
        <input type="date" id="hoursStartDate" onchange="calcHours()">
        <span style="margin:0 6px">至</span>
        <input type="date" id="hoursEndDate" onchange="calcHours()">
        <button class="btn btn-primary" onclick="calcHours()">📊 计算</button>
      </div>
      <div class="hours-results" id="hoursResults">
        <div class="hours-stat">
          <div class="hours-label">参与资源数</div>
          <div class="hours-value" id="hoursCount">0</div>
        </div>
        <div class="hours-stat">
          <div class="hours-label">总工时 (h)</div>
          <div class="hours-value" id="hoursTotal">0</div>
        </div>
        <div class="hours-stat">
          <div class="hours-label">日均总工时 (h/天)</div>
          <div class="hours-value" id="hoursAvgPerDay">0</div>
        </div>
        <div class="hours-stat">
          <div class="hours-label">单资源日均 (h/人/天)</div>
          <div class="hours-value" id="hoursAvgPerPerson">0</div>
        </div>
      </div>
    </div>
  </div>

  <div class="resource-panel" id="resourcePanel">
    <div class="resource-panel-header">
      <span>👤 人员资源占用检索</span>
      <span style="font-size:12px;opacity:0.9">复刻Excel：人员资源情况!A2(开始) B2(结束) C2(人员)</span>
    </div>
    <div class="resource-panel-body">
      <div class="resource-input-row">
        <label>📅 开始时间：</label>
        <input type="date" id="resSearchStart">
        <label>📅 结束时间：</label>
        <input type="date" id="resSearchEnd">
        <label>👤 人员：</label>
        <select id="resSearchPerson">
          <option value="">-- 选择人员 --</option>
        </select>
        <button class="btn btn-primary" onclick="searchResource()">🔍 检索</button>
        <button class="btn btn-secondary" onclick="resetResourceSearch()">🔄 重置</button>
      </div>
      <div id="resourceResults" style="display:none">
        <div class="resource-summary">
          <div class="resource-stat">
            <div class="resource-stat-label">占用任务数</div>
            <div class="resource-stat-value" id="resTaskCount">0</div>
          </div>
          <div class="resource-stat">
            <div class="resource-stat-label">占用天数</div>
            <div class="resource-stat-value" id="resTotalDays">0</div>
          </div>
          <div class="resource-stat">
            <div class="resource-stat-label">总工时 (h)</div>
            <div class="resource-stat-value" id="resTotalHours">0</div>
          </div>
          <div class="resource-stat">
            <div class="resource-stat-label">日均工时 (h/天)</div>
            <div class="resource-stat-value" id="resAvgHours">0</div>
          </div>
        </div>
        <div class="resource-detail-title">📋 资源占用明细</div>
        <div id="resourceDetailTable"></div>
        <div class="calc-detail" id="calcDetail"></div>
      </div>
    </div>
  </div>

  <div class="edit-mode-bar" id="editModeBar">
    <span>📝 编辑模式：<strong id="editModeStatus">关闭</strong>（开启后可直接点击单元格修改计划数据）</span>
    <span id="editModeHint" style="font-size:12px;opacity:0.8">修改后点击「保存修改」或点击其他区域自动保存</span>
  </div>

  <div class="tabs" id="tabs">
    <div class="tab" data-tab="delayed">⚠️ 延期预警 <span class="badge" id="tabDelayedCount">0</span></div>
    <div class="tab active" data-tab="all">📊 全部项目 <span class="badge" id="tabAllCount">0</span></div>
    <div class="tab" data-tab="byDept">🏢 按市场 <span class="badge" id="tabDeptCount">0</span></div>
    <div class="tab" data-tab="archived">📦 已归档 <span class="badge" id="tabArchivedCount">0</span></div>
  </div>

  <div id="tableContainer"></div>
</div>

<div class="save-indicator" id="saveIndicator">✓ 已保存</div>

<script>
// 协作模式变量（必须在最前面！）
let COLLAB_ENABLED = window.COLLAB_MODE === true || false;
let collabLastUpdate = '';
let collabSyncTimer = null;
let collabDirty = false;
// 【死循环防护】
let _collabLastRefreshTime = 0;    // 上次刷新时间戳（防止刷太频繁）
let _collabLastProjectCount = -1;   // 上次项目数（项目数没变就不重绘）
let _collabConsecutiveSame = 0;      // 连续相同 lastUpdate 的次数

function collabIsEnabled() {
  return COLLAB_ENABLED;
}

// ==================== 统一的归档判断函数 ====================
// 【简化方案】协作模式下只使用服务器返回的 p['已归档'] 字段
// 非协作模式下使用本地 archived 对象（兼容）
function isProjectArchived(p) {
  if (!p) return false;
  // 协作模式：只使用服务器返回的数据
  if (collabIsEnabled()) {
    return p['已归档'] === true || p['已归档'] === '已归档' || p['已归档'] === '是';
  }
  // 非协作模式：使用本地 archived 对象
  return !!archived[p.id];
}

const RAW_DATA = __DATA_PLACEHOLDER__;

let currentTab = 'all';
let editMode = false;
let localEdits = JSON.parse(localStorage.getItem('projectEdits') || '{}');
let notes = JSON.parse(localStorage.getItem('projectNotes') || '{}');
let checked = JSON.parse(localStorage.getItem('projectChecked') || '{}');
let archived = JSON.parse(localStorage.getItem('projectArchived') || '{}');
let deletedIds = JSON.parse(localStorage.getItem('deletedIds') || '[]');
let customEmails = JSON.parse(localStorage.getItem('customEmails') || '{}');
let newProjects = JSON.parse(localStorage.getItem('newProjects') || '[]');

// 【关键修复】从服务器获取最新项目数据（刷新 RAW_DATA）
// 同步操作后必须调用此函数，否则客户端的 RAW_DATA 是过时的
async function refreshRawData() {
  if (!collabIsEnabled()) return false;
  try {
    const resp = await fetch('/api/projects');
    if (resp.ok) {
      const data = await resp.json();
      if (data.allProjects) {
        RAW_DATA.allProjects = data.allProjects;
        // 【修复】更新当前日期，避免点检日期停留在HTML生成时的旧日期
        if (data.today) RAW_DATA.today = data.today;
        if (data.threeDaysLater) RAW_DATA.threeDaysLater = data.threeDaysLater;
        // 更新页面显示
        const todayEl = document.getElementById('todayDate');
        if (todayEl) todayEl.textContent = RAW_DATA.today;
        const alertEl = document.getElementById('alertDate');
        if (alertEl) alertEl.textContent = RAW_DATA.threeDaysLater;
        // 重建部门索引
        const depts = {};
        data.allProjects.forEach(function(p) {
          if (p['已归档']) return;  // 已归档的不计入活跃部门
          const d = p['部门'] || '未分配';
          if (!depts[d]) depts[d] = [];
          depts[d].push(p);
        });
        RAW_DATA.depts = depts;
        // 重新计算统计
        const activeProjects = data.allProjects.filter(p => !p['已归档']);
        RAW_DATA.stats = RAW_DATA.stats || {};
        RAW_DATA.stats.total = activeProjects.length;
        console.log('[同步] 已从服务器刷新 ' + data.allProjects.length + ' 个项目');
        
        // 【关键修复】刷新后重新合并 localStorage 中的新增项目
        // 防止 refreshRawData 覆盖了还未同步成功的本地新增项目
        const merged = mergeLocalNewProjects();
        if (merged > 0) {
          console.log('[同步] 刷新后恢复了 ' + merged + ' 个本地新增项目');
        }
        
        return true;
      }
    }
  } catch (e) {
    console.warn('[同步] 刷新项目数据失败:', e);
  }
  return false;
}

// 【修复】从 RAW_DATA 同步 Excel 中的归档标志到 archived 对象（双向同步）
// 后端写入Excel的归档/删除状态需要同步到前端，否则归档看板看不到、删除项目会恢复
async function syncFromExcel() {
  if (typeof RAW_DATA === 'undefined' || !RAW_DATA.allProjects) return;
  
  // 【关键修复】协作模式下先从服务器获取最新数据，避免使用过时的 RAW_DATA
  if (collabIsEnabled()) {
    await refreshRawData();
  }
  
  let syncedCount = 0;
  let removedCount = 0;
  
  // 构建 Excel 中已归档项目的 ID 集合
  const excelArchivedIds = new Set();
  RAW_DATA.allProjects.forEach(function(p) {
    if (p['已归档']) {
      excelArchivedIds.add(p.id);
      // 正向同步：Excel中已归档但前端archived没有的，补充进去
      if (!archived[p.id]) {
        archived[p.id] = { time: new Date().toISOString(), project: p['项目'], fromExcel: true };
        syncedCount++;
      }
    }
  });
  
  // 反向同步：Excel中未归档但前端archived中有的（且来源是Excel的），从archived中移除
  // （用户在当前会话中新归档的项目没有 fromExcel 标记，不会被误删）
  Object.keys(archived).forEach(function(pid) {
    const numPid = parseInt(pid);
    if (!excelArchivedIds.has(numPid) && !excelArchivedIds.has(pid)) {
      // Excel中没有这个归档，检查是否是从Excel同步来的
      if (archived[pid].fromExcel) {
        delete archived[pid];
        removedCount++;
      }
    }
  });
  
  // 同步回 localStorage
  localStorage.setItem('projectArchived', JSON.stringify(archived));
  if (syncedCount > 0 || removedCount > 0) {
    console.log('[同步] 从Excel同步归档状态: 新增' + syncedCount + '个, 移除' + removedCount + '个');
  }
}
syncFromExcel();

// 将本地新增的项目合并到原始数据中（页面刷新后恢复新增项目）
function mergeLocalNewProjects() {
  const existingIds = new Set(RAW_DATA.allProjects.map(p => p.id));
  let addedCount = 0;

  // 1. 从 newProjects（新增项目专用数组）恢复
  newProjects.forEach(np => {
    if (!existingIds.has(np.id)) {
      RAW_DATA.allProjects.push(np);
      const dept = np.部门 || '未分配';
      if (!RAW_DATA.depts[dept]) RAW_DATA.depts[dept] = [];
      RAW_DATA.depts[dept].push(np);
      existingIds.add(np.id);
      addedCount++;
    }
  });

  // 2. 从 localEdits 中恢复（兼容旧数据，只恢复完整的新增项目，不恢复局部编辑）
  Object.keys(localEdits).forEach(id => {
    const numId = parseInt(id);
    if (!existingIds.has(numId)) {
      const edit = localEdits[id];
      // 只有完整的项目数据才合并（不是局部编辑）
      if (edit && edit.项目 && edit.资源名称) {
        RAW_DATA.allProjects.push(edit);
        const dept = edit.部门 || '未分配';
        if (!RAW_DATA.depts[dept]) RAW_DATA.depts[dept] = [];
        RAW_DATA.depts[dept].push(edit);
        addedCount++;
      }
    }
  });

  if (addedCount > 0) {
    console.log(`已从本地存储恢复 ${addedCount} 个新增项目`);
  }
  return addedCount;
}

// 【简化方案】协作模式下不使用本地 localStorage 数据
// 只有非协作模式才从本地恢复数据
if (!collabIsEnabled()) {
  mergeLocalNewProjects();
  syncFromExcel();
}

// 【简化方案】从服务器加载完整项目数据
// 所有状态以服务器为准，不再维护复杂的本地/服务器合并逻辑
async function collabLoadData() {
  try {
    // 直接从 /api/projects 获取完整项目数据（包含归档/删除状态）
    const resp = await fetch('/api/projects');
    if (!resp.ok) return;
    const data = await resp.json();
    
    if (data.allProjects) {
      // 【关键】用服务器返回的最新数据替换 RAW_DATA
      RAW_DATA.allProjects = data.allProjects;
      // 【修复】更新当前日期
      if (data.today) RAW_DATA.today = data.today;
      if (data.threeDaysLater) RAW_DATA.threeDaysLater = data.threeDaysLater;
      // 更新页面显示
      const todayEl = document.getElementById('todayDate');
      if (todayEl) todayEl.textContent = RAW_DATA.today;
      const alertEl = document.getElementById('alertDate');
      if (alertEl) alertEl.textContent = RAW_DATA.threeDaysLater;
      
      // 重建部门索引（只包含未归档的项目）
      const depts = {};
      data.allProjects.forEach(function(p) {
        if (p['已归档']) return;
        const d = p['部门'] || '未分配';
        if (!depts[d]) depts[d] = [];
        depts[d].push(p);
      });
      RAW_DATA.depts = depts;
      
      // 更新统计
      const active = data.allProjects.filter(p => !p['已归档']);
      RAW_DATA.stats = RAW_DATA.stats || {};
      RAW_DATA.stats.total = active.length;
    }
    
    // 【简化】备注、勾选、自定义邮箱保留在本地（这些是用户个人偏好，不需要同步到Excel）
    // 归档和删除状态完全由服务器返回的数据决定，不再使用本地 archived/deletedIds
    
    collabLastUpdate = data.lastUpdate || '';
    
    console.log('[协作] 已从服务器加载数据（简化方案），项目数:', RAW_DATA.allProjects.length);
  } catch (e) {
    console.warn('[协作] 加载数据失败:', e);
  }
}

// 提交本地修改到服务器
async function collabSyncToServer() {
  if (!collabDirty) return;
  
  try {
    const resp = await fetch('/api/sync', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        localEdits: localEdits,
        notes: notes,
        checked: checked,
        archived: archived,
        customEmails: customEmails,
        newProjects: newProjects,
        deletedIds: deletedIds
      })
    });
    
    if (resp.ok) {
      const result = await resp.json();
      collabLastUpdate = result.lastUpdate || collabLastUpdate;
      collabDirty = false;
      // 【关键修复】同步成功后从服务器刷新 RAW_DATA
      // 否则客户端的归档/删除状态是过时的，重置时会出错
      await refreshRawData();
      
      // 【关键修复】只移除已经确认写入服务器的项目
      // 防止因任何原因未写入成功的项目被误删
      if (newProjects.length > 0) {
        const serverIds = new Set(RAW_DATA.allProjects.map(p => p.id));
        const pendingProjects = newProjects.filter(np => !serverIds.has(np.id));
        const confirmedCount = newProjects.length - pendingProjects.length;
        newProjects = pendingProjects;
        localStorage.setItem('newProjects', JSON.stringify(newProjects));
        if (confirmedCount > 0) {
          console.log('[协作] 已确认 ' + confirmedCount + ' 个新增项目写入服务器');
        }
        if (pendingProjects.length > 0) {
          console.log('[协作] ' + pendingProjects.length + ' 个项目待重试同步');
        }
      }
      if (deletedIds.length > 0) {
        const serverIds = new Set(RAW_DATA.allProjects.map(p => p.id));
        const stillDeleted = deletedIds.filter(id => !serverIds.has(id));
        deletedIds = stillDeleted;
        localStorage.setItem('deletedIds', JSON.stringify(deletedIds));
      }
      console.log('[协作] 已同步到服务器');
    }
  } catch (e) {
    console.warn('[协作] 同步失败:', e);
  }
}

// 标记需要同步
function collabMarkDirty() {
  if (collabIsEnabled()) {
    collabDirty = true;
    // 延迟同步（避免频繁请求）
    clearTimeout(collabSyncTimer);
    collabSyncTimer = setTimeout(collabSyncToServer, 500);
  }
}

// 检查服务器是否有更新
async function collabCheckUpdates() {
  if (!collabIsEnabled()) return;
  
  try {
    const resp = await fetch('/api/version', { credentials: 'same-origin' });
    // 401/403 静默处理（session 过期），不打印红色错误
    if (resp.status === 401 || resp.status === 403) return;
    if (!resp.ok) return;
    const data = await resp.json();
    
    // 【死循环防护层1】lastUpdate 为空就跳过
    if (!data.lastUpdate) return;
    
    // 【死循环防护层2】lastUpdate 没变就跳过（最关键！）
    if (data.lastUpdate === collabLastUpdate) {
      _collabConsecutiveSame++;
      return;
    }
    
    // 【死循环防护层3】最小刷新间隔 15 秒（防止刷太猛）
    const now = Date.now();
    if (now - _collabLastRefreshTime < 15000) {
      return;
    }
    
    // 检测到有更新，执行刷新
    console.log('[协作] 检测到服务器有更新，正在刷新...');
    _collabLastRefreshTime = now;
    
    const oldCount = RAW_DATA.allProjects ? RAW_DATA.allProjects.length : 0;
    await collabLoadData();
    
    // 【死循环防护层4】项目数没变就不重绘界面（省资源 + 防止循环）
    const newCount = RAW_DATA.allProjects ? RAW_DATA.allProjects.length : 0;
    if (newCount === oldCount && newCount === _collabLastProjectCount) {
      // 项目数和上次一样，只更新统计不重绘
      return;
    }
    _collabLastProjectCount = newCount;
    _collabConsecutiveSame = 0;
    
    updateStats();
    renderTable();
    initResourceSearch();
  } catch (e) {
    // 忽略网络错误
  }
}

// 显示协作状态
function collabShowStatus() {
  if (!collabIsEnabled()) return;
  
  // 检查是否已有状态栏
  let bar = document.getElementById('collabStatusBar');
  if (bar) return;
  
  bar = document.createElement('div');
  bar.id = 'collabStatusBar';
  bar.style.cssText = 'position:fixed;top:10px;right:10px;background:#10b981;color:white;padding:8px 16px;border-radius:20px;font-size:13px;z-index:9999;box-shadow:0 2px 10px rgba(0,0,0,0.2)';
  bar.innerHTML = '🤝 协作模式 - 已连接';
  document.body.appendChild(bar);
}

function getProject(p) {
  // 【简化方案】协作模式下：所有编辑都通过API同步到服务器了
  // 直接使用服务器返回的数据，忽略本地 localEdits
  if (collabIsEnabled()) {
    return p;
  }
  // 非协作模式：使用本地编辑（兼容）
  const edit = localEdits[p.id];
  if (edit) return { ...p, ...edit };
  return p;
}

function recalcDays(p) {
  const today = new Date(RAW_DATA.today);
  const endDateStr = p.资源结束时间 || p.项目结束时间;
  if (!endDateStr) return null;
  const endDate = new Date(endDateStr);
  if (isNaN(endDate.getTime())) return null;
  const diff = Math.ceil((endDate - today) / (1000 * 60 * 60 * 24));
  return diff;
}

function init() {
  document.getElementById('todayDate').textContent = RAW_DATA.today;
  document.getElementById('alertDate').textContent = RAW_DATA.threeDaysLater;
  
  // 协作模式初始化
  if (collabIsEnabled()) {
    collabShowStatus();
    // 先加载服务器数据
    collabLoadData().then(() => {
      updateStats();
      renderTable();
      initHoursPanel();
      initResourceSearch();
      initDeptFilter();
      bindEvents();
      // 每5秒检查一次更新
      setInterval(collabCheckUpdates, 5000);
    });
  } else {
    updateStats();
    initDeptFilter();
    bindEvents();
    renderTable();
    initHoursPanel();
    initResourceSearch();
  }
}

function initHoursPanel() {
  const today = new Date();
  const startOfWeek = new Date(today);
  const day = today.getDay();
  const diff = day === 0 ? -6 : 1 - day;
  startOfWeek.setDate(today.getDate() + diff);
  const endOfWeek = new Date(startOfWeek);
  endOfWeek.setDate(startOfWeek.getDate() + 6);
  
  document.getElementById('hoursStartDate').value = formatDate(startOfWeek);
  document.getElementById('hoursEndDate').value = formatDate(endOfWeek);
  calcHours();
}

function formatDate(d) {
  return d.toISOString().split('T')[0];
}

// 显示友好日期：1900-01-01 显示为 "开始不限"，2100-01-01 显示为 "结束不限"
function displayDate(dateStr) {
  if (!dateStr || dateStr === '1900-01-01') return '—';
  if (dateStr === '2100-01-01') return '∞';
  return dateStr;
}

// 获取实际计算用日期（处理通配符）
function getActualDate(dateStr, isStart) {
  if (!dateStr) return null;
  if (dateStr === '1900-01-01' || dateStr === '∞') return new Date('1900-01-01');
  if (dateStr === '2100-01-01' || dateStr === '—') return new Date('2100-01-01');
  return new Date(dateStr);
}

function applyHoursPreset() {
  const preset = document.getElementById('hoursRangePreset').value;
  const today = new Date();
  let start, end;
  
  if (preset === 'week') {
    const day = today.getDay();
    const diff = day === 0 ? -6 : 1 - day;
    start = new Date(today);
    start.setDate(today.getDate() + diff);
    end = new Date(start);
    end.setDate(start.getDate() + 6);
  } else if (preset === 'nextweek') {
    const day = today.getDay();
    const diff = day === 0 ? -6 : 1 - day;
    start = new Date(today);
    start.setDate(today.getDate() + diff + 7);
    end = new Date(start);
    end.setDate(start.getDate() + 6);
  } else if (preset === 'month') {
    start = new Date(today.getFullYear(), today.getMonth(), 1);
    end = new Date(today.getFullYear(), today.getMonth() + 1, 0);
  } else {
    return;
  }
  
  document.getElementById('hoursStartDate').value = formatDate(start);
  document.getElementById('hoursEndDate').value = formatDate(end);
  calcHours();
}

function calcHours() {
  const startStr = document.getElementById('hoursStartDate').value;
  const endStr = document.getElementById('hoursEndDate').value;
  if (!startStr || !endStr) return;
  
  const startDate = new Date(startStr);
  const endDate = new Date(endStr);
  if (startDate > endDate) return;
  
  const totalDays = Math.ceil((endDate - startDate) / (1000 * 60 * 60 * 24)) + 1;
  
  const data = RAW_DATA.allProjects.map(getProject).filter(p => !isProjectArchived(p));
  
  let totalHours = 0;
  const uniqueResources = new Set();
  
  data.forEach(p => {
    const hoursPerDay = parseFloat(p.日平均工时) || 0;
    if (hoursPerDay <= 0) return;
    
    const pStart = p.资源开始时间 ? new Date(p.资源开始时间) : null;
    const pEnd = p.资源结束时间 ? new Date(p.资源结束时间) : null;
    
    if (!pStart && !pEnd) return;
    
    const overlapStart = pStart ? (pStart > startDate ? pStart : startDate) : startDate;
    const overlapEnd = pEnd ? (pEnd < endDate ? pEnd : endDate) : endDate;
    
    if (overlapStart > overlapEnd) return;
    
    const overlapDays = Math.ceil((overlapEnd - overlapStart) / (1000 * 60 * 60 * 24)) + 1;
    totalHours += overlapDays * hoursPerDay;
    const resName = p.资源名称 || '未分配';
    uniqueResources.add(resName);
  });
  
  const resourceCount = uniqueResources.size;
  const avgPerDay = totalDays > 0 ? (totalHours / totalDays) : 0;
  const avgPerPerson = resourceCount > 0 && totalDays > 0 ? (totalHours / resourceCount / totalDays) : 0;
  
  document.getElementById('hoursCount').textContent = resourceCount;
  document.getElementById('hoursTotal').textContent = totalHours.toFixed(1);
  document.getElementById('hoursAvgPerDay').textContent = avgPerDay.toFixed(1);
  document.getElementById('hoursAvgPerPerson').textContent = avgPerPerson.toFixed(1);
}

function initResourceSearch() {
  const personSet = new Set();
  // 1. 从项目数据中提取人员（包括本地编辑的）
  RAW_DATA.allProjects.forEach(p => {
    const actual = localEdits[p.id] || p;
    if (actual.资源名称) personSet.add(actual.资源名称);
  });
  // 2. 添加自定义邮箱列表中的所有人员
  const allEmails = getAllEmails();
  Object.keys(allEmails).forEach(name => personSet.add(name));
  
  const sel = document.getElementById('resSearchPerson');
  // 清空旧选项
  sel.innerHTML = '<option value="">-- 选择人员 --</option>';
  Array.from(personSet).sort().forEach(name => {
    const opt = document.createElement('option');
    opt.value = name;
    opt.textContent = name;
    sel.appendChild(opt);
  });
  
  const today = new Date();
  const nextMonth = new Date(today);
  nextMonth.setMonth(today.getMonth() + 1);
  document.getElementById('resSearchStart').value = formatDate(today);
  document.getElementById('resSearchEnd').value = formatDate(nextMonth);
}

function resetResourceSearch() {
  document.getElementById('resSearchPerson').value = '';
  document.getElementById('resourceResults').style.display = 'none';
}

function searchResource() {
  const startStr = document.getElementById('resSearchStart').value;
  const endStr = document.getElementById('resSearchEnd').value;
  const person = document.getElementById('resSearchPerson').value;
  
  if (!startStr || !endStr) {
    alert('请选择开始时间和结束时间');
    return;
  }
  if (!person) {
    alert('请选择人员');
    return;
  }
  
  const searchStart = new Date(startStr);
  const searchEnd = new Date(endStr);
  if (searchStart > searchEnd) {
    alert('开始时间不能晚于结束时间');
    return;
  }
  
  const searchDays = Math.ceil((searchEnd - searchStart) / (1000 * 60 * 60 * 24)) + 1;
  
  // 检索该人员的所有资源占用（包括本地编辑的）
  const personProjects = RAW_DATA.allProjects.filter(p => {
    const actual = localEdits[p.id] || p;
    return actual.资源名称 === person && !isProjectArchived(p);
  });
  
  const matches = [];
  let totalOverlapDays = 0;
  let totalHours = 0;
  let calcLines = [];
  
  calcLines.push('【计算过程】');
  calcLines.push(`检索条件: ${startStr} ~ ${endStr} (共${searchDays}天), 人员: ${person}`);
  calcLines.push('');
  
  personProjects.forEach((p, idx) => {
    const hoursPerDay = parseFloat(p.日平均工时) || 0;
    const pStart = p.资源开始时间 ? new Date(p.资源开始时间) : null;
    const pEnd = p.资源结束时间 ? new Date(p.资源结束时间) : null;
    
    if (!pStart || !pEnd) {
      calcLines.push(`[${idx+1}] ${p.项目} - ${p.资源类型}: 无有效时间，跳过`);
      return;
    }
    
    // 计算重叠时间
    const overlapStart = pStart > searchStart ? pStart : searchStart;
    const overlapEnd = pEnd < searchEnd ? pEnd : searchEnd;
    
    if (overlapStart > overlapEnd) {
      calcLines.push(`[${idx+1}] ${p.项目} - ${p.资源类型}: ${formatDate(pStart)}~${formatDate(pEnd)} 与检索范围无重叠，跳过`);
      return;
    }
    
    const overlapDays = Math.ceil((overlapEnd - overlapStart) / (1000 * 60 * 60 * 24)) + 1;
    const overlapHours = overlapDays * hoursPerDay;
    
    totalOverlapDays += overlapDays;
    totalHours += overlapHours;
    
    matches.push({
      ...p,
      overlapStart: formatDate(overlapStart),
      overlapEnd: formatDate(overlapEnd),
      overlapDays: overlapDays,
      overlapHours: overlapHours,
      pStartOrig: formatDate(pStart),
      pEndOrig: formatDate(pEnd)
    });
    
    calcLines.push(`[${idx+1}] ${p.项目} - ${p.资源类型}: 占用${formatDate(pStart)}~${formatDate(pEnd)} × ${hoursPerDay}h/天`);
    calcLines.push(`     重叠: ${formatDate(overlapStart)}~${formatDate(overlapEnd)} = ${overlapDays}天 × ${hoursPerDay}h = ${overlapHours.toFixed(1)}h`);
  });
  
  calcLines.push('');
  calcLines.push(`【汇总】`);
  calcLines.push(`任务数: ${matches.length} 个`);
  calcLines.push(`总重叠天数: ${totalOverlapDays} 天`);
  calcLines.push(`总工时: ${totalHours.toFixed(1)} h`);
  calcLines.push(`检索区间日均: ${(totalHours / searchDays).toFixed(2)} h/天 (基于检索范围${searchDays}天)`);
  calcLines.push(`实际占用日均: ${totalOverlapDays > 0 ? (totalHours / totalOverlapDays).toFixed(2) : 0} h/天 (基于实际占用${totalOverlapDays}天)`);
  
  // 显示结果
  document.getElementById('resourceResults').style.display = 'block';
  document.getElementById('resTaskCount').textContent = matches.length;
  document.getElementById('resTotalDays').textContent = totalOverlapDays;
  document.getElementById('resTotalHours').textContent = totalHours.toFixed(1);
  document.getElementById('resAvgHours').textContent = (totalHours / searchDays).toFixed(2);
  
  // 明细表
  let tableHtml = '<table class="resource-detail-table"><thead><tr>';
  tableHtml += '<th>序号</th><th>部门</th><th>项目名称</th><th>资源类型</th>';
  tableHtml += '<th>原占用开始</th><th>原占用结束</th>';
  tableHtml += '<th>重叠开始</th><th>重叠结束</th><th>重叠天数</th>';
  tableHtml += '<th>工时(h/天)</th><th>本段工时(h)</th>';
  tableHtml += '</tr></thead><tbody>';
  
  if (matches.length === 0) {
    tableHtml += '<tr><td colspan="11" style="text-align:center;padding:20px;color:#9ca3af">该人员在此时间段内无资源占用</td></tr>';
  } else {
    matches.forEach((m, idx) => {
      tableHtml += `<tr>
        <td>${idx+1}</td>
        <td><span class="dept-tag">${m.部门 || '-'}</span></td>
        <td><strong>${m.项目}</strong></td>
        <td>${m.资源类型 || '-'}</td>
        <td>${m.pStartOrig}</td>
        <td>${m.pEndOrig}</td>
        <td><span class="overlap-badge">${m.overlapStart}</span></td>
        <td><span class="overlap-badge">${m.overlapEnd}</span></td>
        <td style="text-align:center;font-weight:600">${m.overlapDays}</td>
        <td style="text-align:center">${m.日平均工时 || 0}</td>
        <td style="text-align:center;font-weight:600;color:#4f46e5">${m.overlapHours.toFixed(1)}</td>
      </tr>`;
    });
  }
  
  tableHtml += '</tbody></table>';
  document.getElementById('resourceDetailTable').innerHTML = tableHtml;
  document.getElementById('calcDetail').textContent = calcLines.join('\\n');
}

function getStatus(daysLeft) {
  if (daysLeft === null || daysLeft === undefined) return { label: '未设置', class: 'status-none', icon: '⚪' };
  if (daysLeft < 0) return { label: `已延期${-daysLeft}天`, class: 'status-delayed', icon: '🔴' };
  if (daysLeft <= 3) return { label: `${daysLeft}天后到期`, class: 'status-urgent', icon: '🟠' };
  if (daysLeft <= 7) return { label: `${daysLeft}天后到期`, class: 'status-warning', icon: '🟡' };
  return { label: `${daysLeft}天后到期`, class: 'status-normal', icon: '🟢' };
}

function updateStats() {
  // 统计全部项目（排除已归档的项目）
  // 【关键修复】使用统一的 isProjectArchived() 函数判断归档状态
  const all = RAW_DATA.allProjects.map(getProject).map(p => ({ ...p, _days: recalcDays(p) }));
  const active = all.filter(p => !isProjectArchived(p));
  let delayed = 0, urgent = 0, warning = 0;
  active.forEach(p => {
    const days = p._days;
    if (days !== null && days !== undefined) {
      if (days < 0) delayed++;
      else if (days <= 3) urgent++;
      else if (days <= 7) warning++;
    }
  });
  document.getElementById('statDelayed').textContent = delayed;
  document.getElementById('statUrgent').textContent = urgent;
  document.getElementById('statWarning').textContent = warning;
  document.getElementById('statTotal').textContent = active.length;
  
  // Tab计数（排除归档，复用上面的 active）
  const delayedCount = active.filter(p => {
    const d = recalcDays(p);
    return d !== null && d <= 3;
  }).length;
  // 【简化方案】归档数量：协作模式下只使用服务器返回的已归档数据
  // 不再混用本地 localStorage 的 archived 对象（每个浏览器不同导致计数不一致）
  let archivedCount;
  if (collabIsEnabled()) {
    archivedCount = RAW_DATA.allProjects.filter(p => p['已归档']).length;
  } else {
    // 非协作模式：合并本地和服务器数据（兼容）
    const mergedArchivedIds = new Set();
    RAW_DATA.allProjects.forEach(p => { if (p['已归档']) mergedArchivedIds.add(p.id); });
    Object.keys(archived).forEach(id => mergedArchivedIds.add(parseInt(id)));
    archivedCount = mergedArchivedIds.size;
  }
  document.getElementById('tabDelayedCount').textContent = delayedCount;
  document.getElementById('tabAllCount').textContent = active.length;
  document.getElementById('tabDeptCount').textContent = Object.keys(RAW_DATA.depts).length;
  document.getElementById('tabArchivedCount').textContent = archivedCount;
}

function initDeptFilter() {
  const sel = document.getElementById('deptFilter');
  Object.keys(RAW_DATA.depts).sort().forEach(d => {
    const opt = document.createElement('option');
    opt.value = d;
    opt.textContent = d;
    sel.appendChild(opt);
  });
}

function bindEvents() {
  document.getElementById('searchInput').addEventListener('input', renderTable);
  document.getElementById('deptFilter').addEventListener('change', renderTable);
  document.getElementById('statusFilter').addEventListener('change', renderTable);
  
  document.querySelectorAll('.tab').forEach(tab => {
    tab.addEventListener('click', () => {
      document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
      tab.classList.add('active');
      currentTab = tab.dataset.tab;
      renderTable();
    });
  });
}

function toggleEditMode() {
  editMode = !editMode;
  const btn = document.getElementById('editModeBtn');
  const bar = document.getElementById('editModeBar');
  const status = document.getElementById('editModeStatus');
  if (editMode) {
    btn.textContent = '✅ 完成编辑';
    btn.classList.remove('btn-warning');
    btn.classList.add('btn-success');
    bar.classList.add('active');
    status.textContent = '开启';
  } else {
    btn.textContent = '✏️ 编辑模式';
    btn.classList.remove('btn-success');
    btn.classList.add('btn-warning');
    bar.classList.remove('active');
    status.textContent = '关闭';
  }
  renderTable();
  initHoursPanel();
  initResourceSearch();
}


function getFilteredData() {
  let data = RAW_DATA.allProjects.map(getProject).map(p => ({ ...p, _days: recalcDays(p) }));
  
  if (currentTab === 'archived') {
    data = data.filter(p => isProjectArchived(p));
  } else {
    data = data.filter(p => !isProjectArchived(p));
    if (currentTab === 'delayed') {
      data = data.filter(p => p._days !== null && p._days <= 3);
    }
  }
  
  data.sort((a, b) => {
    const da = a._days ?? 999;
    const db = b._days ?? 999;
    return da - db;
  });
  
  const search = document.getElementById('searchInput').value.toLowerCase();
  const dept = document.getElementById('deptFilter').value;
  const status = document.getElementById('statusFilter').value;
  
  return data.filter(p => {
    if (search) {
      const text = `${p.部门} ${p.项目} ${p.项目描述} ${p.资源类型} ${p.资源名称}`.toLowerCase();
      if (!text.includes(search)) return false;
    }
    if (dept && p.部门 !== dept) return false;
    if (status) {
      const days = p._days;
      if (status === 'delayed' && !(days !== null && days < 0)) return false;
      if (status === 'urgent' && !(days !== null && days >= 0 && days <= 3)) return false;
      if (status === 'warning' && !(days !== null && days > 3 && days <= 7)) return false;
      if (status === 'normal' && !(days !== null && days > 7)) return false;
    }
    return true;
  });
}

function renderBoardView(data, container) {
  if (!data || data.length === 0) {
    container.innerHTML = '<div class="empty-state"><div class="icon">🎉</div><div>暂无数据</div></div>';
    return;
  }

  // TR排序辅助函数：提取TR后的数字和字母
  // TR1 → {num:1, suffix:''}
  // TR4A → {num:4, suffix:'A'}
  // TR-001 → {num:1, suffix:''}
  function parseTRSortKey(trStr) {
    if (!trStr) return { num: 99999, suffix: 'ZZZ' };
    const s = String(trStr).trim().toUpperCase();
    // 匹配 TR 后面的内容（忽略空格、横线、大小写）
    const m = s.match(/^TR[\s\-]*(\d+)([A-Z]?)/i);
    if (m) {
      return { num: parseInt(m[1], 10), suffix: m[2] || '' };
    }
    // 没有TR前缀的，放到最后
    return { num: 99999, suffix: s };
  }
  function compareTR(a, b) {
    const ka = parseTRSortKey(a);
    const kb = parseTRSortKey(b);
    if (ka.num !== kb.num) return ka.num - kb.num;
    return ka.suffix.localeCompare(kb.suffix);
  }

  // 按市场分组
  const byMarket = {};
  data.forEach(p => {
    const market = p.部门 || '未分配';
    if (!byMarket[market]) byMarket[market] = [];
    byMarket[market].push(p);
  });

  // 按市场名称排序
  const sortedMarkets = Object.keys(byMarket).sort();

  let html = '<div class="board-container">';

  sortedMarkets.forEach(market => {
    const marketProjects = byMarket[market];

    // 按项目名分组（大框）
    const byProjectName = {};
    marketProjects.forEach(p => {
      const name = p.项目 || '未命名项目';
      if (!byProjectName[name]) byProjectName[name] = [];
      byProjectName[name].push(p);
    });

    // 统计市场级别
    const projCount = Object.keys(byProjectName).length;
    const resCount = marketProjects.length;
    const delayedCount = marketProjects.filter(p => {
      const days = p._days;
      return days !== null && days !== undefined && days <= 3;
    }).length;

    html += '<div class="board-market-group">';
    html += '<div class="board-market-header">';
    html += '<div class="board-market-title">🏢 ' + market + '</div>';
    html += '<div class="board-market-stats">';
    html += projCount + ' 个项目 · ' + resCount + ' 项资源';
    if (delayedCount > 0) html += ' · ⚠️ ' + delayedCount + ' 延期';
    html += '</div>';
    html += '</div>';
    html += '<div class="board-project-list">';

    // 按项目名排序（有延期的排前面）
    const sortedProjNames = Object.keys(byProjectName).sort((a, b) => {
      const aMin = Math.min(...byProjectName[a].map(p => p._days !== null && p._days !== undefined ? p._days : 999));
      const bMin = Math.min(...byProjectName[b].map(p => p._days !== null && p._days !== undefined ? p._days : 999));
      return aMin - bMin;
    });

    sortedProjNames.forEach((projName, pIdx) => {
      const projResources = byProjectName[projName];

      // 在项目名内，按项目描述（TR）分组（小框）
      const byTR = {};
      projResources.forEach(p => {
        const tr = p.项目描述 || '未分配TR';
        if (!byTR[tr]) byTR[tr] = [];
        byTR[tr].push(p);
      });

      // 按TR规则排序（TR1→TR2→TR3→TR4→TR4A→TR5）
      const sortedTRs = Object.keys(byTR).sort(compareTR);

      // 项目级统计
      const projDelayed = projResources.filter(p => {
        const days = p._days;
        return days !== null && days !== undefined && days <= 3;
      }).length;

      const projGroupId = 'board_projname_' + market.replace(/[^a-zA-Z0-9]/g, '') + '_' + pIdx;

      html += '<div class="board-project-name-group">';
      // 项目名大框标题
      html += '<div class="board-project-name-header" onclick="toggleBoardGroup(\\'' + projGroupId + '\\')">';
      html += '<div>';
      html += '<div class="board-project-title">📁 ' + projName + '</div>';
      html += '</div>';
      html += '<div class="board-project-meta">';
      html += '<span class="dept-tag" style="font-size:11px">' + projResources.length + ' 项资源 · ' + sortedTRs.length + ' 个TR</span>';
      if (projDelayed > 0) html += '<span class="status-tag status-delayed" style="font-size:11px">⚠️ ' + projDelayed + ' 延期</span>';
      html += '<span style="color:#94a3b8;font-size:11px">▼</span>';
      html += '</div>';
      html += '</div>';
      html += '<div id="' + projGroupId + '" style="display:block" class="board-tr-container">';

      // 每个TR小框
      sortedTRs.forEach((trName, trIdx) => {
        const resources = byTR[trName];
        const trGroupId = projGroupId + '_tr_' + trIdx;

        // 按开始时间排序
        resources.sort((a, b) => {
          const sa = a.资源开始时间 || '9999-12-31';
          const sb = b.资源开始时间 || '9999-12-31';
          return sa.localeCompare(sb);
        });

        const trDelayed = resources.filter(p => {
          const days = p._days;
          return days !== null && days !== undefined && days <= 3;
        }).length;

        html += '<div class="board-tr-group">';
        // TR小框标题
        html += '<div class="board-tr-header" onclick="toggleBoardGroup(\\'' + trGroupId + '\\')">';
        html += '<div class="board-tr-title">🏷️ ' + (trName === '未分配TR' ? '未分配TR' : trName) + '</div>';
        html += '<div class="board-project-meta">';
        html += '<span style="font-size:11px;color:#6b7280">' + resources.length + ' 项</span>';
        if (trDelayed > 0) html += '<span class="status-tag status-delayed" style="font-size:11px">⚠️ ' + trDelayed + '</span>';
        html += '<span style="color:#94a3b8;font-size:11px">▼</span>';
        html += '</div>';
        html += '</div>';
        html += '<div id="' + trGroupId + '" style="display:block">';

        html += '<table class="board-resource-table"><thead><tr>';
        html += '<th>✓</th><th>负责人</th><th>资源类型</th><th>开始时间</th><th>结束时间</th><th>剩余</th><th>工时</th><th>状态</th><th>备注</th><th>操作</th>';
        html += '</tr></thead><tbody>';

        resources.forEach(r => {
          const days = r._days;
          const status = getStatus(days);
          const noteId = 'note_' + r.id;
          const noteText = notes[noteId] || '';
          const isChecked = checked[r.id] ? 'checked' : '';
          const rowClass = days < 0 ? 'delayed-row' : (days <= 3 ? 'urgent-row' : '');

          html += '<tr class="' + rowClass + '">';
          html += '<td><input type="checkbox" ' + isChecked + ' onchange="toggleCheck(' + r.id + ', this.checked)"></td>';
          html += '<td><strong>' + (r.资源名称 || '-') + '</strong></td>';
          html += '<td>' + (r.资源类型 || '-') + '</td>';
          html += '<td><span class="board-time-range">' + displayDate(r.资源开始时间) + '</span></td>';
          html += '<td><span class="board-time-range">' + displayDate(r.资源结束时间) + '</span></td>';
          html += '<td style="text-align:center;font-weight:600">' + (days !== null ? days : '-') + '</td>';
          html += '<td style="text-align:center;font-weight:600;color:#4f46e5">' + (r.日平均工时 || 0) + 'h</td>';
          html += '<td><span class="status-tag ' + status.class + '" style="font-size:11px">' + status.icon + ' ' + status.label + '</span></td>';
          html += '<td class="note-cell"><div class="editable ' + (noteText ? '' : 'note-text') + '" contenteditable="true" onblur="saveNote(' + r.id + ', this.innerText)" onfocus="if(this.innerText===\\'点击添加备注...\\'){this.innerText=\\'\\';this.classList.remove(\\'note-text\\')}" data-id="' + r.id + '">' + (noteText || '点击添加备注...') + '</div></td>';
          if (archived[r.id]) {
            html += '<td><button class="archive-btn restore" onclick="restoreProject(' + r.id + ')">↩️ 恢复</button> <button class="archive-btn" style="background:#ef4444" onclick="deleteProject(' + r.id + ')">🗑️ 删除</button></td>';
          } else {
            html += '<td><button class="archive-btn" onclick="archiveProject(' + r.id + ')">📦 归档</button> <button class="archive-btn" style="background:#ef4444" onclick="deleteProject(' + r.id + ')">🗑️ 删除</button></td>';
          }
          html += '</tr>';
        });

        html += '</tbody></table></div></div>';
      });

      html += '</div></div>';
    });

    html += '</div></div>';
  });

  html += '</div>';
  container.innerHTML = html;
}

function toggleBoardGroup(id) {
  const el = document.getElementById(id);
  if (el) {
    el.style.display = el.style.display === 'none' ? 'block' : 'none';
  }
}

// 工程师邮箱映射表（姓名拼音@hxgroup.com）
const ENGINEER_EMAILS = {
  "危才文": "caiwen.wei@hxgroup.com",
  "朱葵阳": "kuiyang.zhu@hxgroup.com",
  "李昊龙": "haolong.li@hxgroup.com",
  "毛文豪": "wenhao.mao@hxgroup.com",
  "童景顺": "jingshun.tong@hxgroup.com",
  "答金雨": "jinyu.da@hxgroup.com",
  "舒杰": "jie.shu@hxgroup.com",
  "袁琮欣": "congxin.yuan@hxgroup.com",
  "项家祺": "jiaqi.xiang@hxgroup.com",
};

// 邮件抄送名单（点检报表自动抄送给这些人）
let CC_EMAILS = JSON.parse(localStorage.getItem('ccEmails') || '[]');

function saveCCEmails() {
  localStorage.setItem('ccEmails', JSON.stringify(CC_EMAILS));
}

// 获取所有人员（包括资源人员和成员管理中的人员）
function getAllPersons() {
  const personSet = new Set();
  const allEmails = getAllEmails();
  Object.keys(allEmails).forEach(n => personSet.add(n));
  
  // 添加所有资源人员
  RAW_DATA.allProjects.forEach(p => {
    const actual = getProject(p);
    if (actual.资源名称) personSet.add(actual.资源名称);
  });
  
  return Array.from(personSet).sort();
}

// 获取或生成人员邮箱
function getOrCreateEmail(name) {
  const allEmails = getAllEmails();
  if (allEmails[name]) return allEmails[name];
  
  // 自动生成邮箱并保存
  const email = generateEmailFromName(name);
  if (email && email.includes('@')) {
    customEmails[name] = email;
    localStorage.setItem('customEmails', JSON.stringify(customEmails));
    collabMarkDirty();
    return email;
  }
  return null;
}

function toggleCCEmail(name) {
  const email = getOrCreateEmail(name);
  if (!email) return;
  
  const idx = CC_EMAILS.indexOf(email);
  if (idx >= 0) {
    CC_EMAILS.splice(idx, 1);
  } else {
    CC_EMAILS.push(email);
  }
  saveCCEmails();
  renderCCSelector();
}

function renderCCSelector() {
  // 渲染所有抄送选择器容器
  const containers = [
    document.getElementById('ccSelectorContainer'),
    document.getElementById('emailClientCCSelector')
  ].filter(c => c);
  
  if (containers.length === 0) return;
  
  const names = getAllPersons();
  
  let html = '<div style="display:flex;flex-wrap:wrap;gap:6px;max-height:180px;overflow:auto;padding:8px;background:#f9fafb;border-radius:6px;border:1px solid #e5e7eb">';
  names.forEach(name => {
    const email = getOrCreateEmail(name);
    const isSelected = email && CC_EMAILS.includes(email);
    const btnStyle = isSelected 
      ? 'background:#3b82f6;color:white;border-color:#3b82f6' 
      : 'background:white;color:#374151;border-color:#d1d5db';
    html += `<button type="button" onclick="toggleCCEmail('${name}')" style="padding:4px 10px;border:1px solid;border-radius:16px;font-size:12px;cursor:pointer;transition:all 0.15s;${btnStyle}">${isSelected ? '✓ ' : ''}${name}</button>`;
  });
  html += '</div>';
  html += `<div style="margin-top:6px;font-size:12px;color:#6b7280">已选择 ${CC_EMAILS.length} 位抄送人员</div>`;
  
  containers.forEach(c => { c.innerHTML = html; });
  
  // 更新抄送邮箱显示区域
  const emailDisplay = document.getElementById('ccEmailDisplay');
  if (emailDisplay) {
    if (CC_EMAILS.length > 0) {
      emailDisplay.style.display = 'block';
      emailDisplay.innerHTML = `<strong>抄送邮箱（${CC_EMAILS.length}人）：</strong><br>${CC_EMAILS.join(', ')}<br><br><button type="button" onclick="navigator.clipboard.writeText('${CC_EMAILS.join(',')}').then(()=>alert('✅ 已复制到剪贴板'))" style="padding:4px 12px;background:#3b82f6;color:white;border:none;border-radius:4px;cursor:pointer;font-size:12px">📋 复制抄送邮箱</button>`;
    } else {
      emailDisplay.style.display = 'none';
    }
  }
}

// ==================== 邮件生成通用函数 ====================

// 获取所有人员及其所有任务（用于邮件客户端发送的完整清单）
function getAllPersonsWithTasks() {
  const allProjects = RAW_DATA.allProjects
    .map(getProject)
    .map(p => ({ ...p, _days: recalcDays(p) }))
    .filter(p => !isProjectArchived(p));
  
  const byPerson = {};
  allProjects.forEach(p => {
    const name = p.资源名称 || '未分配';
    if (!byPerson[name]) byPerson[name] = [];
    byPerson[name].push(p);
  });
  
  return byPerson;
}

// 获取延期人员及其延期任务
function getDelayedPersonsWithTasks() {
  const delayedProjects = RAW_DATA.allProjects
    .map(getProject)
    .map(p => ({ ...p, _days: recalcDays(p) }))
    .filter(p => !isProjectArchived(p))
    .filter(p => {
      const days = p._days;
      return days !== null && days !== undefined && days <= 3;
    });
  
  const byPerson = {};
  delayedProjects.forEach(p => {
    const name = p.资源名称 || '未分配';
    if (!byPerson[name]) byPerson[name] = [];
    byPerson[name].push(p);
  });
  
  return byPerson;
}

// 获取指定人员最近一周（今天起7天）的工作安排 - 按天分组
function getWeeklyWorkSchedule(name, tasks) {
  const today = new Date(RAW_DATA.today);
  today.setHours(0, 0, 0, 0);
  const weekEnd = new Date(today);
  weekEnd.setDate(today.getDate() + 6);
  weekEnd.setHours(23, 59, 59, 999);
  
  const weekDays = [];
  const weekDayNames = ['周日', '周一', '周二', '周三', '周四', '周五', '周六'];
  let totalWeeklyHours = 0;
  
  // 生成一周7天的日期
  for (let d = 0; d < 7; d++) {
    const dayDate = new Date(today);
    dayDate.setDate(today.getDate() + d);
    const dayStr = formatDate(dayDate);
    weekDays.push({
      date: dayDate,
      dateStr: dayStr,
      dayName: weekDayNames[dayDate.getDay()],
      dayLabel: d === 0 ? '今天' : (d === 1 ? '明天' : weekDayNames[dayDate.getDay()]),
      tasks: [],
      totalHours: 0
    });
  }
  
  // 遍历所有任务，分配到每一天
  tasks.forEach(p => {
    const hoursPerDay = parseFloat(p.日平均工时) || 0;
    if (hoursPerDay <= 0) return;
    
    const pStart = p.资源开始时间 ? new Date(p.资源开始时间) : null;
    const pEnd = p.资源结束时间 ? new Date(p.资源结束时间) : null;
    
    if (!pStart || !pEnd) return;
    pStart.setHours(0, 0, 0, 0);
    pEnd.setHours(23, 59, 59, 999);
    
    // 如果任务与本周无重叠，跳过
    if (pEnd < today || pStart > weekEnd) return;
    
    const taskInfo = {
      部门: p.部门 || '-',
      项目: p.项目,
      项目描述: p.项目描述 || '',
      资源类型: p.资源类型 || '-',
      日工时: hoursPerDay,
      原开始: formatDate(pStart),
      原结束: formatDate(pEnd),
      _days: p._days
    };
    
    // 将任务分配到本周内的每一天
    weekDays.forEach(day => {
      const dayStart = new Date(day.date);
      dayStart.setHours(0, 0, 0, 0);
      const dayEnd = new Date(day.date);
      dayEnd.setHours(23, 59, 59, 999);
      
      // 判断该任务在当天是否有效
      if (pStart <= dayEnd && pEnd >= dayStart) {
        day.tasks.push({ ...taskInfo });
        day.totalHours += hoursPerDay;
        totalWeeklyHours += hoursPerDay;
      }
    });
  });
  
  // 对每天的任务进行排序：项目经理优先 > 已延期 > 即将到期 > 日工时从大到小
  weekDays.forEach(day => {
    day.tasks.sort((a, b) => {
      // 1. 项目经理优先
      const aPM = a.资源类型 && a.资源类型.indexOf('项目经理') >= 0 ? 0 : 1;
      const bPM = b.资源类型 && b.资源类型.indexOf('项目经理') >= 0 ? 0 : 1;
      if (aPM !== bPM) return aPM - bPM;
      
      // 2. 按优先级：已延期 > 即将到期 > 其他
      const pa = a._days !== null && a._days !== undefined ? (a._days < 0 ? 0 : a._days <= 3 ? 1 : 2) : 3;
      const pb = b._days !== null && b._days !== undefined ? (b._days < 0 ? 0 : b._days <= 3 ? 1 : 2) : 3;
      if (pa !== pb) return pa - pb;
      
      // 3. 日工时从大到小
      return b.日工时 - a.日工时;
    });
  });
  
  return { weekDays, totalWeeklyHours, weekStart: formatDate(today), weekEnd: formatDate(weekEnd) };
}

// 根据工时和上下班时间生成推荐工作时间表（按天）
// 工作时间：早8:30-12:00，午13:00-17:30，加班18:30起
// 排序规则：项目经理优先排在每天最前面
function generateRecommendedSchedule(name, tasks) {
  const { weekDays, totalWeeklyHours } = getWeeklyWorkSchedule(name, tasks);
  
  // 检查是否有任何一天有工作
  const hasAnyWork = weekDays.some(d => d.tasks.length > 0);
  if (!hasAnyWork) {
    return { hasWork: false, text: '本周暂无工作安排', hasOvertime: false, days: [] };
  }
  
  // 时间槽定义（单位：小时，从0点开始计算）
  const timeSlots = [
    { start: 8.5, end: 12.0, label: '早上', type: 'normal' },    // 8:30-12:00 = 3.5h
    { start: 13.0, end: 17.5, label: '下午', type: 'normal' },   // 13:00-17:30 = 4.5h
    { start: 18.5, end: 24.0, label: '下午', type: 'overtime' }  // 18:30-24:00 = 5.5h
  ];
  
  // 格式化小时为时间字符串
  function formatTime(hours) {
    const h = Math.floor(hours);
    const m = Math.round((hours - h) * 60);
    return `${h.toString().padStart(2,'0')}:${m.toString().padStart(2,'0')}`;
  }
  
  // 为某一天生成推荐安排（支持跨日到次日凌晨）
  function generateDaySchedule(dayInfo) {
    if (dayInfo.tasks.length === 0) {
      return { dateStr: dayInfo.dateStr, dayLabel: dayInfo.dayLabel, dayName: dayInfo.dayName, lines: [], hasOvertime: false, totalHours: 0, hasNextDay: false };
    }
    
    let lines = [];
    let hasOvertime = false;
    let hasNextDay = false;
    let currentTime = 8.5; // 从当天8:30开始
    let dayOffset = 0; // 0=当天, 1=次日, 2=后日...
    const MAX_DAY_OFFSET = 2; // 最多跨到后日
    
    // 扩展的时间槽（包含跨日的凌晨加班时间）
    function getExtendedSlots(offset) {
      if (offset === 0) {
        // 当天：正常工作时间 + 加班
        return [
          { start: 8.5, end: 12.0, label: '早上', type: 'normal', dayPrefix: '' },
          { start: 13.0, end: 17.5, label: '下午', type: 'normal', dayPrefix: '' },
          { start: 18.5, end: 24.0, label: '下午', type: 'overtime', dayPrefix: '' }
        ];
      } else {
        // 次日及以后：凌晨加班（0:00-8:30）+ 正常工作时间 + 加班
        const prefix = offset === 1 ? '次日：' : `第${offset+1}天：`;
        return [
          { start: 0.0, end: 8.5, label: '凌晨', type: 'overtime', dayPrefix: prefix },
          { start: 8.5, end: 12.0, label: '早上', type: 'normal', dayPrefix: prefix },
          { start: 13.0, end: 17.5, label: '下午', type: 'normal', dayPrefix: prefix },
          { start: 18.5, end: 24.0, label: '下午', type: 'overtime', dayPrefix: prefix }
        ];
      }
    }
    
    // 任务已经在 getWeeklyWorkSchedule 中排好序了（项目经理优先 > 延期 > 工时）
    for (let taskIdx = 0; taskIdx < dayInfo.tasks.length; taskIdx++) {
      const task = dayInfo.tasks[taskIdx];
      let taskHoursLeft = task.日工时;
      
      while (taskHoursLeft > 0.001 && dayOffset <= MAX_DAY_OFFSET) {
        const slots = getExtendedSlots(dayOffset);
        
        // 找到当前时间所在的时间段
        let slot = null;
        for (let si = 0; si < slots.length; si++) {
          if (currentTime < slots[si].end - 0.001) {
            slot = slots[si];
            if (currentTime < slot.start) currentTime = slot.start;
            break;
          }
        }
        
        if (!slot) {
          // 当前日期的时间已用完，进入下一天
          dayOffset++;
          if (dayOffset > 1) hasNextDay = true;
          currentTime = 0.0; // 下一天从0点开始
          continue;
        }
        
        if (slot.type === 'overtime') hasOvertime = true;
        if (slot.dayPrefix !== '') hasNextDay = true;
        
        const slotHoursLeft = slot.end - currentTime;
        const allocHours = Math.min(taskHoursLeft, slotHoursLeft);
        const endTime = currentTime + allocHours;
        
        lines.push(
          `${slot.dayPrefix}${slot.label}${formatTime(currentTime)}-${formatTime(endTime)}    ${task.项目}项目 ${task.资源类型} 工时${allocHours}`
        );
        
        currentTime = endTime;
        taskHoursLeft -= allocHours;
        
        // 如果刚好到了午休或下班时间，跳到下一个时间段
        if (Math.abs(currentTime - 12.0) < 0.01) currentTime = 13.0;
        if (Math.abs(currentTime - 17.5) < 0.01) currentTime = 18.5;
      }
    }
    
    return {
      dateStr: dayInfo.dateStr,
      dayLabel: dayInfo.dayLabel,
      dayName: dayInfo.dayName,
      lines: lines,
      hasOvertime: hasOvertime,
      hasNextDay: hasNextDay,
      totalHours: dayInfo.totalHours
    };
  }
  
  // 生成每天的安排
  const daySchedules = weekDays.map(day => generateDaySchedule(day));
  const anyOvertime = daySchedules.some(d => d.hasOvertime);
  const avgDailyHours = (totalWeeklyHours / 7).toFixed(1);
  
  return {
    hasWork: true,
    days: daySchedules,
    hasOvertime: anyOvertime,
    dailyHours: avgDailyHours,
    totalWeeklyHours: totalWeeklyHours.toFixed(1)
  };
}

// 为指定人员生成邮件内容
// delayedOnly: true=只包含延期任务(延期催办邮件用), false=包含所有任务(邮件客户端发送用)
function generateEmailContent(name, tasks, todayStr, delayedOnly) {
  const email = getAllEmails()[name] || name + '@hxgroup.com';
  
  // 分离延期任务和其他任务
  const delayedTasks = tasks.filter(t => t._days !== null && t._days !== undefined && t._days <= 3);
  const otherTasks = tasks.filter(t => delayedTasks.indexOf(t) === -1);
  
  const allDelayedCount = delayedTasks.length;
  const displayTasks = delayedOnly ? delayedTasks : tasks;
  
  let subject = delayedOnly 
    ? `【延期催办】${name} - 您有${allDelayedCount}项任务已延期或即将到期 (${todayStr})`
    : `【任务提醒】${name} - 您有${tasks.length}项任务待跟进 (${todayStr})`;
  
  let body = `${name} 您好：\n\n`;
  if (delayedOnly) {
    body += `以下是您负责的延期或即将到期任务，请参考超声波户用水表产品线研发群内钉钉文档「超声波户表脚本」，请及时跟进处理：\n\n`;
  } else {
    body += `以下是您负责的所有任务，请参考超声波户用水表产品线研发群内钉钉文档「超声波户表脚本」，请及时跟进处理：\n\n`;
  }
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  
  if (delayedOnly) {
    // 延期催办模式：只显示延期任务
    if (delayedTasks.length > 0) {
      body += `⚠️  已延期 / 3天内即将到期任务（${delayedTasks.length}项）：\n\n`;
      delayedTasks.forEach((p, idx) => {
        const days = p._days;
        const statusLabel = days < 0 ? `🔴 已延期${Math.abs(days)}天` : `🟠 ${days}天后到期`;
        body += `【${idx+1}】${statusLabel}\n`;
        body += `    项目：${p.项目}\n`;
        if (p.项目描述) body += `    描述：${p.项目描述.substring(0, 60)}\n`;
        body += `    市场：${p.部门 || '-'}\n`;
        body += `    资源类型：${p.资源类型 || '-'}\n`;
        body += `    开始时间：${displayDate(p.资源开始时间)}\n`;
        body += `    结束时间：${displayDate(p.资源结束时间)}\n`;
        body += `    日工时：${p.日平均工时 || 0}h/天\n`;
        if (notes['note_' + p.id]) body += `    点检备注：${notes['note_' + p.id]}\n`;
        body += `\n`;
      });
    }
  } else {
    // 完整模式：先显示延期/即将到期任务，再显示其他任务
    if (delayedTasks.length > 0) {
      body += `⚠️  已延期 / 3天内即将到期任务（${delayedTasks.length}项）：\n\n`;
      delayedTasks.forEach((p, idx) => {
        const days = p._days;
        const statusLabel = days < 0 ? `🔴 已延期${Math.abs(days)}天` : `🟠 ${days}天后到期`;
        body += `【${idx+1}】${statusLabel}\n`;
        body += `    项目：${p.项目}\n`;
        if (p.项目描述) body += `    描述：${p.项目描述.substring(0, 60)}\n`;
        body += `    市场：${p.部门 || '-'}\n`;
        body += `    资源类型：${p.资源类型 || '-'}\n`;
        body += `    开始时间：${displayDate(p.资源开始时间)}\n`;
        body += `    结束时间：${displayDate(p.资源结束时间)}\n`;
        body += `    日工时：${p.日平均工时 || 0}h/天\n`;
        if (notes['note_' + p.id]) body += `    点检备注：${notes['note_' + p.id]}\n`;
        body += `\n`;
      });
    }
    
    if (otherTasks.length > 0) {
      body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
      body += `📋 其他进行中任务（${otherTasks.length}项）：\n\n`;
      otherTasks.forEach((p, idx) => {
        const days = p._days;
        const statusLabel = days !== null && days !== undefined ? `🟢 剩余${days}天` : '⚪ 未设置';
        body += `【${idx+1}】${statusLabel}\n`;
        body += `    项目：${p.项目}\n`;
        if (p.项目描述) body += `    描述：${p.项目描述.substring(0, 60)}\n`;
        body += `    市场：${p.部门 || '-'}\n`;
        body += `    资源类型：${p.资源类型 || '-'}\n`;
        body += `    开始时间：${displayDate(p.资源开始时间)}\n`;
        body += `    结束时间：${displayDate(p.资源结束时间)}\n`;
        body += `    日工时：${p.日平均工时 || 0}h/天\n`;
        if (notes['note_' + p.id]) body += `    点检备注：${notes['note_' + p.id]}\n`;
        body += `\n`;
      });
    }
    
    // ============ 最近一周工作安排表（仅邮件客户端发送模式） ============
    const weeklyData = getWeeklyWorkSchedule(name, tasks);
    const hasWorkDays = weeklyData.weekDays.filter(d => d.tasks.length > 0);
    if (hasWorkDays.length > 0) {
      body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
      body += `📅 最近一周工作安排（${weeklyData.weekStart} ~ ${weeklyData.weekEnd}）\n\n`;
      body += `本周预估总工时：${weeklyData.totalWeeklyHours.toFixed(1)} h\n\n`;
      
      // 按天显示工作安排表格
      weeklyData.weekDays.forEach(day => {
        if (day.tasks.length === 0) return;
        body += `【${day.dayLabel}】${day.dateStr} (${day.dayName})  当日工时：${day.totalHours.toFixed(1)}h\n`;
        day.tasks.forEach((t, idx) => {
          const statusTag = t._days !== null && t._days !== undefined 
            ? (t._days < 0 ? '🔴' : t._days <= 3 ? '🟠' : '') 
            : '';
          const pmTag = t.资源类型 && t.资源类型.indexOf('项目经理') >= 0 ? '👔' : '';
          body += `  ${idx+1}. ${pmTag}${statusTag} ${t.部门} | ${t.项目} | ${t.资源类型} | ${t.日工时}h/天\n`;
        });
        body += `\n`;
      });
      
      // ============ 推荐工作时间表（按天） ============
      const rec = generateRecommendedSchedule(name, tasks);
      if (rec.hasWork && rec.days && rec.days.length > 0) {
        body += `⏰ 推荐工作时间表（日均${rec.dailyHours}h，按8:30-12:00/13:00-17:30/18:30起加班规划）\n\n`;
        
        rec.days.forEach(day => {
          if (day.lines.length === 0) return;
          body += `【${day.dayLabel}】${day.dateStr} (${day.dayName})  ${day.hasOvertime ? ' ⚠️含加班' : ''}\n`;
          day.lines.forEach(line => {
            body += `  ${line}\n`;
          });
          body += `\n`;
        });
        
        // 加班提醒
        if (rec.hasOvertime) {
          body += `⚠️  【重要提醒】根据当前工作安排，您在 17:30 之后仍有工作任务需要处理。\n`;
          body += `   工作内容繁多，请与上级主管沟通，合理安排工作优先级和时间分配。\n\n`;
        }
      }
    }
  }
  
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  body += `请您：\n`;
  body += `  1. 确认以上任务的当前状态\n`;
  body += `  2. 如任务已完成，请及时反馈，并提交给对应项目经理与职能经理\n`;
  body += `  3. 如需调整时间，请更新项目计划，并提交给对应项目经理与职能经理\n`;
  body += `  4. 参考「推荐工作时间表」合理安排每日工作，如有加班请与主管沟通\n\n`;
  body += `---\n`;
  body += `此邮件由项目点检系统自动生成\n`;
  body += `点检日期：${todayStr}\n`;
  
  return { name, email, subject, body, count: displayTasks.length, delayedCount: allDelayedCount };
}

// 打开邮件客户端发送邮件
function openMailClientForPerson(name, tasks, todayStr, delayedOnly) {
  const { email, subject, body } = generateEmailContent(name, tasks, todayStr, delayedOnly);
  
  const ccList = (CC_EMAILS && CC_EMAILS.length > 0) ? CC_EMAILS.join(',') : '';
  
  let mailtoUrl = 'mailto:' + encodeURIComponent(email);
  const params = [];
  if (ccList) params.push('cc=' + encodeURIComponent(ccList));
  params.push('subject=' + encodeURIComponent(subject));
  params.push('body=' + encodeURIComponent(body));
  mailtoUrl += '?' + params.join('&');
  
  window.location.href = mailtoUrl;
}

function generateAllEmails() {
  const today = new Date();
  const todayStr = today.toISOString().split('T')[0];
  
  // 获取所有人员及其所有任务
  const allPersons = getAllPersonsWithTasks();
  
  // 筛选出有延期任务的人员
  const delayedPersons = {};
  Object.keys(allPersons).forEach(name => {
    const tasks = allPersons[name];
    const hasDelayed = tasks.some(t => t._days !== null && t._days !== undefined && t._days <= 3);
    if (hasDelayed) delayedPersons[name] = tasks;
  });
  
  const personNames = Object.keys(delayedPersons).sort();
  
  if (personNames.length === 0) {
    alert('🎉 恭喜！当前没有延期或即将到期的任务，无需发送催办邮件。');
    return;
  }
  
  // 生成每个人的邮件内容（只包含延期任务）
  const emails = [];
  personNames.forEach(name => {
    const emailData = generateEmailContent(name, delayedPersons[name], todayStr, true);
    emails.push(emailData);
  });
  
  // 生成汇总内容并复制到剪贴板
  let summary = `📋 延期催办邮件汇总 (${todayStr})\n`;
  summary += `共 ${emails.length} 位同事，${emails.reduce((s,e)=>s+e.delayedCount,0)} 项延期任务\n\n`;
  summary += `═══════════════════════════════════════════\n\n`;
  
  emails.forEach(e => {
    summary += `👤 ${e.name} (${e.email})\n`;
    summary += `📧 主题：${e.subject}\n`;
    summary += `📊 延期任务数：${e.delayedCount} 项 / 总任务数：${e.count} 项\n\n`;
    summary += `${e.body}\n\n`;
    summary += `═══════════════════════════════════════════\n\n`;
  });
  
  // 保存到全局变量供打开邮件客户端使用
  window._allEmailData = { emails: emails, allPersons: allPersons, todayStr: todayStr };
  
  // 复制到剪贴板并显示结果
  if (navigator.clipboard) {
    navigator.clipboard.writeText(summary).then(() => {
      showEmailResult(emails, summary, true);
    }).catch(() => {
      showEmailResult(emails, summary, false);
    });
  } else {
    showEmailResult(emails, summary, false);
  }
}

function showEmailResult(emails, summary, copied) {
  const { todayStr, allPersons } = window._allEmailData || { todayStr: new Date().toISOString().split('T')[0], allPersons: {} };
  
  // 创建弹窗显示结果
  const modal = document.createElement('div');
  modal.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:10000;display:flex;align-items:center;justify-content:center;padding:20px';
  modal.innerHTML = `
    <div style="background:white;border-radius:12px;max-width:600px;width:100%;max-height:80vh;overflow:auto;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,0.3)">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <h3 style="margin:0;color:#1e293b">📧 延期催办邮件已生成</h3>
        <button onclick="this.closest('div[style*=fixed]').remove()" style="border:none;background:none;font-size:24px;cursor:pointer;color:#64748b">×</button>
      </div>
      <div style="background:${copied ? '#ecfdf5' : '#fef3c7'};padding:12px 16px;border-radius:8px;margin-bottom:16px">
        ${copied ? '✅ 邮件内容已复制到剪贴板，可直接粘贴发送' : '⚠️ 请手动复制下方邮件内容'}
      </div>
      <div style="margin-bottom:16px">
        <div style="font-weight:600;margin-bottom:8px;color:#334155">涉及人员：${emails.length} 人</div>
        ${emails.map((e, idx) => `
          <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 12px;background:#f8fafc;border-radius:6px;margin-bottom:4px">
            <span><strong>${e.name}</strong> <span style="color:#64748b;font-size:12px">${e.email}</span></span>
            <div style="display:flex;gap:6px;align-items:center">
              <span style="background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:500">${e.delayedCount}项延期</span>
              <span style="color:#64748b;font-size:11px">共${e.count}项</span>
              <button onclick="openSingleEmail('${e.name}')" style="padding:4px 10px;background:#f59e0b;color:white;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600">📨 发送</button>
            </div>
          </div>
        `).join('')}
      </div>
      <div style="border-top:1px solid #e2e8f0;padding-top:16px">
        <div style="font-weight:600;margin-bottom:8px;color:#334155">邮件内容预览（已按人员分组）：</div>
        <textarea id="emailContentText" readonly style="width:100%;height:200px;padding:12px;border:1px solid #e2e8f0;border-radius:8px;font-family:monospace;font-size:12px;resize:vertical"></textarea>
      </div>
      <div style="display:flex;gap:8px;margin-top:16px;justify-content:space-between;flex-wrap:wrap">
        <div style="display:flex;gap:8px">
          <button onclick="document.getElementById('emailContentText').select();document.execCommand('copy');this.textContent='✅ 已复制';setTimeout(()=>{this.textContent='📋 复制全部内容'},2000)" style="padding:10px 20px;background:#667eea;color:white;border:none;border-radius:8px;cursor:pointer;font-weight:600">📋 复制全部内容</button>
          <button onclick="openAllEmailsClient()" style="padding:10px 20px;background:#f59e0b;color:white;border:none;border-radius:8px;cursor:pointer;font-weight:600">📨 全部打开邮件客户端</button>
        </div>
        <button onclick="this.closest('div[style*=fixed]').remove()" style="padding:10px 20px;background:#e2e8f0;color:#475569;border:none;border-radius:8px;cursor:pointer;font-weight:600">关闭</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  document.getElementById('emailContentText').value = summary;
}

// 打开单个人员的邮件客户端（延期催办模式）
function openSingleEmail(name) {
  const { todayStr, allPersons } = window._allEmailData || {};
  const tasks = allPersons[name] || [];
  const today = todayStr || new Date().toISOString().split('T')[0];
  openMailClientForPerson(name, tasks, today, true);
}

// 打开所有人员的邮件客户端（延期催办模式）
function openAllEmailsClient() {
  const { emails, allPersons, todayStr } = window._allEmailData || {};
  if (!emails || emails.length === 0) return;
  
  const today = todayStr || new Date().toISOString().split('T')[0];
  emails.forEach((e, idx) => {
    setTimeout(() => {
      const tasks = allPersons[e.name] || [];
      openMailClientForPerson(e.name, tasks, today, true);
    }, idx * 500);
  });
  
  alert('✅ 已为 ' + emails.length + ' 位同事打开邮件客户端，请在邮件窗口中点击发送。');
}

function sendAllEmails() {
  // 获取所有延期/3天内到期的任务
  const delayedProjects = RAW_DATA.allProjects
    .map(getProject)
    .map(p => ({ ...p, _days: recalcDays(p) }))
    .filter(p => !isProjectArchived(p))
    .filter(p => {
      const days = p._days;
      return days !== null && days !== undefined && days <= 3;
    });
  
  if (delayedProjects.length === 0) {
    alert('🎉 恭喜！当前没有延期或即将到期的任务，无需发送邮件。');
    return;
  }
  
  // 按人员分组
  const byPerson = {};
  delayedProjects.forEach(p => {
    const name = p.资源名称 || '未分配';
    if (!byPerson[name]) byPerson[name] = [];
    byPerson[name].push(p);
  });
  
  const personNames = Object.keys(byPerson).sort();
  
  // 创建选择弹窗
  const modal = document.createElement('div');
  modal.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:10000;display:flex;align-items:center;justify-content:center;padding:20px';
  modal.innerHTML = `
    <div style="background:white;border-radius:12px;max-width:520px;width:100%;max-height:80vh;overflow:auto;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,0.3)">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <h3 style="margin:0;color:#1e293b">📧 选择要发送邮件的人员</h3>
        <button onclick="this.closest('div[style*=fixed]').remove()" style="border:none;background:none;font-size:24px;cursor:pointer;color:#64748b">×</button>
      </div>
      <div style="background:#eff6ff;padding:10px 14px;border-radius:8px;margin-bottom:16px;color:#1e40af;font-size:13px">
        共 ${personNames.length} 位同事有延期任务（${delayedProjects.length} 项），请勾选要发送邮件的人员
      </div>
      <div style="display:flex;gap:8px;margin-bottom:12px">
        <button onclick="document.querySelectorAll('.email-person-checkbox').forEach(c => c.checked = true)" style="padding:6px 14px;background:#e2e8f0;border:none;border-radius:6px;cursor:pointer;font-size:12px;color:#475569">全选</button>
        <button onclick="document.querySelectorAll('.email-person-checkbox').forEach(c => c.checked = false)" style="padding:6px 14px;background:#e2e8f0;border:none;border-radius:6px;cursor:pointer;font-size:12px;color:#475569">取消全选</button>
      </div>
      <div id="emailPersonList" style="max-height:300px;overflow-y:auto;border:1px solid #e2e8f0;border-radius:8px;padding:8px">
        ${personNames.map((name, idx) => {
          const email = getAllEmails()[name] || name + '@hxgroup.com';
          const tasks = byPerson[name];
          const delayedCount = tasks.filter(t => t._days < 0).length;
          const urgentCount = tasks.filter(t => t._days >= 0 && t._days <= 3).length;
          return `
            <label style="display:flex;align-items:center;padding:10px 12px;border-radius:8px;cursor:pointer;transition:background 0.15s" onmouseover="this.style.background='#f8fafc'" onmouseout="this.style.background='transparent'">
              <input type="checkbox" class="email-person-checkbox" checked value="${name}" style="width:18px;height:18px;margin-right:12px;cursor:pointer" ${idx < 3 ? '' : ''}>
              <div style="flex:1">
                <div style="font-weight:600;color:#1e293b">${name}</div>
                <div style="font-size:12px;color:#64748b">${email}</div>
              </div>
              <div style="display:flex;gap:6px">
                ${delayedCount > 0 ? `<span style="background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:500">🔴 ${delayedCount}已延期</span>` : ''}
                ${urgentCount > 0 ? `<span style="background:#ffedd5;color:#9a3412;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:500">🟠 ${urgentCount}即将到期</span>` : ''}
              </div>
            </label>
          `;
        }).join('')}
      </div>
      <div id="emailSendStatus" style="margin-top:12px;padding:10px;border-radius:8px;display:none"></div>
      <div style="display:flex;gap:8px;margin-top:16px;justify-content:flex-end">
        <button onclick="this.closest('div[style*=fixed]').remove()" style="padding:10px 20px;background:#e2e8f0;color:#475569;border:none;border-radius:8px;cursor:pointer;font-weight:600">取消</button>
        <button id="confirmSendEmailBtn" onclick="confirmSendEmails()" style="padding:10px 24px;background:#10b981;color:white;border:none;border-radius:8px;cursor:pointer;font-weight:600">📤 确认发送</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  
  // 保存数据到全局变量供确认函数使用
  window._emailData = { byPerson: byPerson, modal: modal };
}

function confirmSendEmails() {
  const { byPerson, modal } = window._emailData;
  
  // 获取勾选的人员
  const checkedNames = Array.from(document.querySelectorAll('.email-person-checkbox:checked')).map(c => c.value);
  
  if (checkedNames.length === 0) {
    const statusEl = document.getElementById('emailSendStatus');
    statusEl.style.display = 'block';
    statusEl.style.background = '#fef2f2';
    statusEl.style.color = '#991b1b';
    statusEl.textContent = '⚠️ 请至少选择一位人员';
    return;
  }
  
  // 筛选选中人员的任务
  const selectedProjects = [];
  checkedNames.forEach(name => {
    if (byPerson[name]) {
      selectedProjects.push(...byPerson[name]);
    }
  });
  
  // 显示发送中状态
  const statusEl = document.getElementById('emailSendStatus');
  const confirmBtn = document.getElementById('confirmSendEmailBtn');
  statusEl.style.display = 'block';
  statusEl.style.background = '#eff6ff';
  statusEl.style.color = '#1e40af';
  statusEl.textContent = '📧 正在向 ' + checkedNames.length + ' 位同事发送邮件...';
  confirmBtn.disabled = true;
  confirmBtn.style.opacity = '0.6';
  confirmBtn.style.cursor = 'not-allowed';
  
  // 调用本地邮件服务
  fetch('http://localhost:8765/api/send-emails', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ delayed_projects: selectedProjects })
  })
  .then(response => response.json())
  .then(data => {
    if (data.success) {
      statusEl.style.background = '#ecfdf5';
      statusEl.style.color = '#065f46';
      statusEl.textContent = '✅ 邮件发送成功！';
      setTimeout(() => modal.remove(), 1500);
    } else {
      statusEl.style.background = '#fef2f2';
      statusEl.style.color = '#991b1b';
      statusEl.textContent = '❌ 发送失败：' + data.message;
      confirmBtn.disabled = false;
      confirmBtn.style.opacity = '1';
      confirmBtn.style.cursor = 'pointer';
    }
  })
  .catch(error => {
    statusEl.style.background = '#fef2f2';
    statusEl.style.color = '#991b1b';
    statusEl.innerHTML = '❌ 无法连接到邮件服务<br><br>请先启动邮件服务：<code style="background:#f1f5f9;padding:2px 6px;border-radius:4px">python mail_server.py</code><br>然后再点击发送邮件按钮。';
    confirmBtn.disabled = false;
    confirmBtn.style.opacity = '1';
    confirmBtn.style.cursor = 'pointer';
  });
}

function openEmailClient() {
  // 获取所有资源人员及其所有任务
  const allPersons = getAllPersonsWithTasks();
  const personNames = Object.keys(allPersons).sort();
  
  if (personNames.length === 0) {
    alert('⚠️ 当前没有可发送邮件的人员。');
    return;
  }
  
  const today = new Date();
  const todayStr = today.toISOString().split('T')[0];
  
  const modal = document.createElement('div');
  modal.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:10000;display:flex;align-items:center;justify-content:center;padding:20px';
  modal.innerHTML = `
    <div style="background:white;border-radius:12px;max-width:520px;width:100%;max-height:80vh;overflow:auto;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,0.3)">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <h3 style="margin:0;color:#1e293b">📨 选择邮件接收人</h3>
        <button onclick="this.closest('div[style*=fixed]').remove()" style="border:none;background:none;font-size:24px;cursor:pointer;color:#64748b">×</button>
      </div>
      <div style="background:#dbeafe;padding:10px 14px;border-radius:8px;margin-bottom:16px;color:#1e40af;font-size:13px">
        所有资源人员共 <strong>${personNames.length}</strong> 人，将通过您的默认邮件客户端（如Outlook、Foxmail等）发送邮件，无需配置服务器。
      </div>
      <div style="margin-bottom:12px">
        <div style="color:#64748b;font-size:12px;margin-bottom:4px">抄送（点击选择成员）</div>
        <div id="emailClientCCSelector"></div>
      </div>
      <div style="display:flex;gap:8px;margin-bottom:12px">
        <button onclick="document.querySelectorAll('.email-client-checkbox').forEach(c => c.checked = true)" style="padding:6px 14px;background:#e2e8f0;border:none;border-radius:6px;cursor:pointer;font-size:12px;color:#475569">全选</button>
        <button onclick="document.querySelectorAll('.email-client-checkbox').forEach(c => c.checked = false)" style="padding:6px 14px;background:#e2e8f0;border:none;border-radius:6px;cursor:pointer;font-size:12px;color:#475569">取消全选</button>
        <button onclick="selectDelayedOnly()" style="padding:6px 14px;background:#fee2e2;color:#991b1b;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:500">只选延期</button>
      </div>
      <div id="emailClientPersonList" style="max-height:300px;overflow-y:auto;border:1px solid #e2e8f0;border-radius:8px;padding:8px">
        ${personNames.map((name, idx) => {
          const tasks = allPersons[name];
          const delayedCount = tasks.filter(t => t._days !== null && t._days !== undefined && t._days < 0).length;
          const urgentCount = tasks.filter(t => t._days !== null && t._days !== undefined && t._days >= 0 && t._days <= 3).length;
          const totalCount = tasks.length;
          const hasDelayed = delayedCount + urgentCount > 0;
          return `
            <label style="display:flex;align-items:center;padding:10px 12px;border-radius:8px;cursor:pointer;transition:background 0.15s" onmouseover="this.style.background='#f8fafc'" onmouseout="this.style.background='transparent'">
              <input type="checkbox" class="email-client-checkbox" data-delayed="${hasDelayed}" ${hasDelayed ? 'checked' : ''} value="${name}" style="width:18px;height:18px;margin-right:12px;cursor:pointer">
              <div style="flex:1">
                <div style="font-weight:600;color:#1e293b">${name}</div>
                <div style="color:#64748b;font-size:11px">共 ${totalCount} 项任务</div>
              </div>
              <div style="display:flex;gap:6px">
                ${delayedCount > 0 ? `<span style="background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:500">🔴 ${delayedCount}已延期</span>` : ''}
                ${urgentCount > 0 ? `<span style="background:#ffedd5;color:#9a3412;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:500">🟠 ${urgentCount}即将到期</span>` : ''}
              </div>
            </label>
          `;
        }).join('')}
      </div>
      <div style="display:flex;gap:8px;margin-top:16px;justify-content:flex-end">
        <button onclick="this.closest('div[style*=fixed]').remove()" style="padding:10px 20px;background:#e2e8f0;color:#475569;border:none;border-radius:8px;cursor:pointer;font-weight:600">取消</button>
        <button onclick="confirmOpenEmailClient()" style="padding:10px 24px;background:#f59e0b;color:white;border:none;border-radius:8px;cursor:pointer;font-weight:600">📨 打开邮件客户端</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  
  // 渲染抄送选择器
  renderCCSelector();
  
  window._emailClientData = { allPersons: allPersons, modal: modal, todayStr: todayStr };
}

// 只选中有延期任务的人员
function selectDelayedOnly() {
  document.querySelectorAll('.email-client-checkbox').forEach(c => {
    c.checked = c.getAttribute('data-delayed') === 'true';
  });
}

function confirmOpenEmailClient() {
  const { allPersons, modal, todayStr } = window._emailClientData;
  
  const checkedNames = Array.from(document.querySelectorAll('.email-client-checkbox:checked')).map(c => c.value);
  
  if (checkedNames.length === 0) {
    alert('⚠️ 请至少选择一位人员');
    return;
  }
  
  checkedNames.forEach((name, index) => {
    const tasks = allPersons[name];
    if (!tasks || tasks.length === 0) return;
    
    setTimeout(() => {
      openMailClientForPerson(name, tasks, todayStr, false);
    }, index * 500);
  });
  
  modal.remove();
  alert('✅ 已为选中的 ' + checkedNames.length + ' 位同事打开邮件客户端，请在邮件窗口中点击发送。');
}

function renderTable() {
  const data = getFilteredData();
  const container = document.getElementById('tableContainer');
  
  if (data.length === 0) {
    container.innerHTML = '<div class="empty-state"><div class="icon">🎉</div><div>暂无符合条件的项目</div></div>';
    return;
  }
  
  // 全部项目：按市场-项目分组看板视图
  if (currentTab === 'all') {
    renderBoardView(data, container);
    return;
  }
  
  let html = '<table><thead><tr>';
  html += '<th style="width:40px">✓</th>';
  html += '<th>市场</th>';
  html += '<th>项目名称</th>';
  html += '<th>开始时间</th>';
  if (editMode) {
    html += '<th>结束时间</th>';
  } else {
    html += '<th>结束时间</th>';
  }
  html += '<th>负责人</th>';
  html += '<th>资源类型</th>';
  html += '<th>剩余</th>';
  html += '<th>工时(h/天)</th>';
  html += '<th>状态</th>';
  html += '<th class="note-cell">点检备注</th>';
  html += '<th style="width:80px">操作</th>';
  html += '</tr></thead><tbody>';
  
  data.forEach(p => {
    const days = p._days;
    const status = getStatus(days);
    const rowClass = days < 0 ? 'delayed-row' : (days <= 3 ? 'urgent-row' : '');
    const noteId = `note_${p.id}`;
    const noteText = notes[noteId] || '';
    const isChecked = checked[p.id] ? 'checked' : '';
    
    html += `<tr class="${rowClass}">`;
    html += `<td><input type="checkbox" ${isChecked} onchange="toggleCheck(${p.id}, this.checked)"></td>`;
    html += `<td><span class="dept-tag">${p.部门 || '-'}</span></td>`;
    html += `<td><strong>${p.项目}</strong>${p.项目描述 ? `<br><span style="color:#6b7280;font-size:12px">${p.项目描述.substring(0,50)}</span>` : ''}</td>`;
    html += `<td>${displayDate(p.资源开始时间)}</td>`;
    if (editMode) {
      html += `<td><input type="date" class="editable date-input" value="${p.资源结束时间 || ''}" onchange="saveDate(${p.id}, this.value)"></td>`;
    } else {
      html += `<td>${displayDate(p.资源结束时间)}</td>`;
    }
    
    if (editMode) {
      html += `<td><span class="editable" contenteditable="true" onblur="saveField(${p.id}, '资源名称', this.innerText)" data-field="资源名称">${p.资源名称 || '-'}</span></td>`;
    } else {
      html += `<td>${p.资源名称 || '-'}</td>`;
    }
    html += `<td>${p.资源类型 || '-'}</td>`;
    html += `<td style="text-align:center;font-weight:600">${days !== null ? days : '-'}</td>`;
    html += `<td style="text-align:center;font-weight:600;color:#4f46e5">${p.日平均工时 || 0}</td>`;
    html += `<td><span class="status-tag ${status.class}">${status.icon} ${status.label}</span></td>`;
    html += `<td class="note-cell"><div class="editable ${noteText ? '' : 'note-text'}" contenteditable="true" 
              onblur="saveNote(${p.id}, this.innerText)" 
              onfocus="if(this.innerText==='点击添加备注...'){this.innerText='';this.classList.remove('note-text')}"
              data-id="${p.id}">${noteText || '点击添加备注...'}</div></td>`;
    if (isProjectArchived(p)) {
      html += `<td><button class="archive-btn restore" onclick="restoreProject(${p.id})">↩️ 恢复</button> <button class="archive-btn" style="background:#ef4444" onclick="deleteProject(${p.id})">🗑️ 删除</button></td>`;
    } else {
      html += `<td><button class="archive-btn" onclick="archiveProject(${p.id})">📦 归档</button> <button class="archive-btn" style="background:#ef4444" onclick="deleteProject(${p.id})">🗑️ 删除</button></td>`;
    }
    html += '</tr>';
  });
  
  html += '</tbody></table>';
  container.innerHTML = html;
  updateSelectionCount();
}

// 【简化方案】保存字段编辑：协作模式下直接调用API
async function saveField(id, field, value) {
  const cleanValue = value.trim();
  if (collabIsEnabled()) {
    // 协作模式：直接调用API
    const editData = {};
    if (cleanValue === '' || cleanValue === '-') {
      editData[field] = '';
    } else {
      editData[field] = cleanValue;
    }
    const result = await callActionApi('edit', { id: id, fields: editData });
    if (result && result.success) {
      updateStats();
      renderTable();
      initResourceSearch();
      showSaved();
    }
  } else {
    // 非协作模式：本地逻辑（兼容）
    if (!localEdits[id]) localEdits[id] = {};
    if (cleanValue === '' || cleanValue === '-') {
      delete localEdits[id][field];
      if (Object.keys(localEdits[id]).length === 0) delete localEdits[id];
    } else {
      localEdits[id][field] = cleanValue;
    }
    localStorage.setItem('projectEdits', JSON.stringify(localEdits));
    updateStats();
    showSaved();
  }
}

// 【简化方案】保存日期编辑：协作模式下直接调用API
async function saveDate(id, value) {
  if (collabIsEnabled()) {
    // 协作模式：直接调用API
    const editData = { '资源结束时间': value || '' };
    const result = await callActionApi('edit', { id: id, fields: editData });
    if (result && result.success) {
      updateStats();
      renderTable();
      initResourceSearch();
      showSaved();
    }
  } else {
    // 非协作模式：本地逻辑（兼容）
    if (!localEdits[id]) localEdits[id] = {};
    if (!value) {
      delete localEdits[id]['资源结束时间'];
      if (Object.keys(localEdits[id]).length === 0) delete localEdits[id];
    } else {
      localEdits[id]['资源结束时间'] = value;
    }
    localStorage.setItem('projectEdits', JSON.stringify(localEdits));
    updateStats();
    renderTable();
    initResourceSearch();
    showSaved();
  }
}

function saveNote(id, text) {
  const noteId = `note_${id}`;
  if (text.trim() === '' || text === '点击添加备注...') {
    delete notes[noteId];
  } else {
    notes[noteId] = text.trim();
  }
  localStorage.setItem('projectNotes', JSON.stringify(notes));
  showSaved();
}

// ==================== 简化方案：统一的API调用函数 ====================
// 【简化方案】所有操作直接调用API端点，不再维护本地状态
async function callActionApi(endpoint, data) {
  if (!collabIsEnabled()) {
    alert('协作模式未启用，无法执行此操作');
    return { success: false, message: '协作模式未启用' };
  }
  try {
    const resp = await fetch(`/api/action/${endpoint}`, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(data)
    });
    
    // 先处理 HTTP 级别错误
    if (!resp.ok) {
      let errMsg = 'HTTP ' + resp.status;
      try {
        const errData = await resp.json();
        if (errData.error) errMsg = errData.error;
        if (errData.message) errMsg = errData.message;
      } catch (_) {}
      if (resp.status === 401) errMsg = '登录已过期，请重新登录';
      if (resp.status === 403) errMsg = '权限不足，无法执行此操作';
      return { success: false, message: errMsg };
    }
    
    const result = await resp.json();
    
    // 兼容旧格式：只有 error 字段的响应
    if (result.error && result.success === undefined) {
      return { success: false, message: result.error };
    }
    
    // 【关键优化】操作成功后自动从服务器拉取最新项目数据
    // 服务器为了避免502超时，不再在操作响应中返回全量数据
    if (result.success) {
      try {
        const projResp = await fetch('/api/projects');
        if (projResp.ok) {
          const projData = await projResp.json();
          if (projData.allProjects) {
            // 用服务器返回的最新数据替换 RAW_DATA
            RAW_DATA.allProjects = projData.allProjects;
            // 【修复】更新当前日期
            if (projData.today) RAW_DATA.today = projData.today;
            if (projData.threeDaysLater) RAW_DATA.threeDaysLater = projData.threeDaysLater;
            // 更新页面显示
            const todayEl = document.getElementById('todayDate');
            if (todayEl) todayEl.textContent = RAW_DATA.today;
            const alertEl = document.getElementById('alertDate');
            if (alertEl) alertEl.textContent = RAW_DATA.threeDaysLater;
            // 重建部门索引
            const depts = {};
            projData.allProjects.forEach(function(p) {
              if (p['已归档']) return;
              const d = p['部门'] || '未分配';
              if (!depts[d]) depts[d] = [];
              depts[d].push(p);
            });
            RAW_DATA.depts = depts;
            // 更新统计
            const active = projData.allProjects.filter(p => !p['已归档']);
            RAW_DATA.stats = RAW_DATA.stats || {};
            RAW_DATA.stats.total = active.length;
          }
          // 【关键修复】必须更新 collabLastUpdate，否则下一次版本检查会认为有更新
          // 导致不必要的刷新，甚至可能拉到旧数据覆盖当前状态
          if (projData.lastUpdate) {
            collabLastUpdate = projData.lastUpdate;
          }
        }
      } catch (refreshErr) {
        console.warn('[API] 刷新项目数据失败:', refreshErr);
      }
    }
    return result;
  } catch (e) {
    console.warn('[API] 调用失败:', e);
    return { success: false, message: '网络错误：' + (e.message || e) };
  }
}

function toggleCheck(id, isChecked) {
  if (isChecked) checked[id] = true;
  else delete checked[id];
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  updateSelectionCount();
  // 勾选是本地操作，不需要同步到服务器
}

function updateSelectionCount() {
  const el = document.getElementById('selectionCount');
  if (el) {
    const count = Object.keys(checked).length;
    el.textContent = count > 0 ? `已选 ${count} 项` : '';
  }
}

function selectAllVisible() {
  const data = getFilteredData();
  let added = 0;
  data.forEach(p => {
    if (!checked[p.id]) {
      checked[p.id] = true;
      added++;
    }
  });
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  renderTable();
  updateSelectionCount();
  // 全选是本地操作，不需要同步到服务器
}

function clearSelection() {
  if (Object.keys(checked).length === 0) return;
  checked = {};
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  renderTable();
  updateSelectionCount();
}

// 批量归档（一次调用后端批量接口）
async function batchArchive() {
  const ids = Object.keys(checked).map(Number).filter(id => {
    const p = RAW_DATA.allProjects.find(x => x.id === id);
    return p && !isProjectArchived(p);
  });
  if (ids.length === 0) {
    alert('请先勾选要归档的项目（未归档状态）');
    return;
  }
  if (!confirm(`确定要批量归档选中的 ${ids.length} 个项目吗？`)) return;
  
  let done = 0;
  if (collabIsEnabled()) {
    const r = await callActionApi('batch-archive', { ids: ids });
    if (r && r.success) {
      done = r.done || ids.length;
      ids.forEach(id => { archived[id] = { time: new Date().toISOString(), project: '批量归档' }; delete checked[id]; });
    } else {
      alert('❌ 批量归档失败：' + (r ? r.message : '未知错误'));
      return;
    }
  } else {
    ids.forEach(id => { archived[id] = { time: new Date().toISOString(), project: '批量归档' }; delete checked[id]; });
    done = ids.length;
  }
  localStorage.setItem('projectArchived', JSON.stringify(archived));
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  updateStats();
  renderTable();
  updateSelectionCount();
  showSaved();
  alert(`✅ 成功归档 ${done}/${ids.length} 个项目`);
}

// 批量恢复归档（一次调用后端批量接口）
async function batchRestore() {
  const ids = Object.keys(checked).map(Number).filter(id => {
    const p = RAW_DATA.allProjects.find(x => x.id === id);
    return p && isProjectArchived(p);
  });
  if (ids.length === 0) {
    alert('请先勾选要恢复的已归档项目');
    return;
  }
  if (!confirm(`确定要批量恢复选中的 ${ids.length} 个已归档项目吗？`)) return;
  
  let done = 0;
  if (collabIsEnabled()) {
    const r = await callActionApi('batch-unarchive', { ids: ids });
    if (r && r.success) {
      done = r.done || ids.length;
      ids.forEach(id => { delete archived[id]; delete checked[id]; });
    } else {
      alert('❌ 批量恢复失败：' + (r ? r.message : '未知错误'));
      return;
    }
  } else {
    ids.forEach(id => { delete archived[id]; delete checked[id]; });
    done = ids.length;
  }
  localStorage.setItem('projectArchived', JSON.stringify(archived));
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  updateStats();
  renderTable();
  updateSelectionCount();
  showSaved();
  alert(`✅ 成功恢复 ${done}/${ids.length} 个项目`);
}

// 批量删除（一次调用后端批量接口）
async function batchDelete() {
  const ids = Object.keys(checked).map(Number).filter(id => {
    return RAW_DATA.allProjects.find(x => x.id === id);
  });
  if (ids.length === 0) {
    alert('请先勾选要删除的项目');
    return;
  }
  if (!confirm(`⚠️ 确定要彻底删除选中的 ${ids.length} 个项目吗？\\n\\n此操作不可恢复！`)) return;
  
  let done = 0;
  if (collabIsEnabled()) {
    const r = await callActionApi('batch-delete', { ids: ids });
    if (r && r.success) {
      done = r.done || ids.length;
    } else {
      alert('❌ 批量删除失败：' + (r ? r.message : '未知错误'));
      return;
    }
  } else {
    done = ids.length;
  }
  // 清理本地状态
  ids.forEach(id => {
    const idx = RAW_DATA.allProjects.findIndex(x => x.id === id);
    if (idx !== -1) RAW_DATA.allProjects.splice(idx, 1);
    Object.keys(RAW_DATA.depts).forEach(d => {
      RAW_DATA.depts[d] = RAW_DATA.depts[d].filter(x => x.id !== id);
    });
    RAW_DATA.delayedProjects = RAW_DATA.delayedProjects.filter(x => x.id !== id);
    delete localEdits[id];
    delete archived[id];
    delete notes[`note_${id}`];
    delete checked[id];
    if (!deletedIds.includes(id)) deletedIds.push(id);
  });
  localStorage.setItem('projectEdits', JSON.stringify(localEdits));
  localStorage.setItem('projectArchived', JSON.stringify(archived));
  localStorage.setItem('projectNotes', JSON.stringify(notes));
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  localStorage.setItem('deletedIds', JSON.stringify(deletedIds));
  updateStats();
  renderTable();
  initResourceSearch();
  updateSelectionCount();
  showSaved();
  alert(`✅ 成功删除 ${done}/${ids.length} 个项目`);
}

// 【简化方案】归档：直接调用API
async function archiveProject(id) {
  const p = RAW_DATA.allProjects.find(x => x.id === id);
  const name = p ? p.项目 : '此项目';
  if (confirm(`确定要归档「${name}」吗？\n归档后将不在点检列表中显示，可在「已归档」标签页恢复。`)) {
    if (collabIsEnabled()) {
      const result = await callActionApi('archive', { id: id });
      if (result && result.success) {
        // 【关键修复】协作模式下也必须同步本地 archived 对象，
        // 否则后续 collabSyncToServer() 会发送过时的 archived 数据，
        // 后端全量替换后导致归档状态丢失
        archived[id] = { time: new Date().toISOString(), project: name };
        localStorage.setItem('projectArchived', JSON.stringify(archived));
        updateStats();
        renderTable();
        showSaved();
      }
    } else {
      // 非协作模式：本地逻辑（兼容）
      archived[id] = { time: new Date().toISOString(), project: name };
      localStorage.setItem('projectArchived', JSON.stringify(archived));
      updateStats();
      renderTable();
      showSaved();
    }
  }
}

// 【简化方案】恢复归档：直接调用API
async function restoreProject(id) {
  if (collabIsEnabled()) {
    const result = await callActionApi('unarchive', { id: id });
    if (result && result.success) {
      // 【关键修复】协作模式下也必须同步本地 archived 对象，
      // 否则后续 collabSyncToServer() 会发送过时的 archived 数据
      delete archived[id];
      localStorage.setItem('projectArchived', JSON.stringify(archived));
      updateStats();
      renderTable();
      showSaved();
    }
  } else {
    delete archived[id];
    localStorage.setItem('projectArchived', JSON.stringify(archived));
    updateStats();
    renderTable();
    showSaved();
  }
}

// 【简化方案】删除项目：直接调用API
async function deleteProject(id) {
  const p = RAW_DATA.allProjects.find(x => x.id === id);
  const name = p ? p.项目 : '此项目';
  const dept = p ? p.部门 : '';
  if (!confirm(`确定要彻底删除「${dept ? dept + ' - ' : ''}${name}」吗？\\n\\n⚠️ 此操作不可恢复！删除后将从所有视图中移除。`)) {
    return;
  }
  
  if (collabIsEnabled()) {
    const result = await callActionApi('delete', { id: id });
    if (result && result.success) {
      // 【关键修复】协作模式下也必须同步本地状态，
      // 否则后续 collabSyncToServer() 会发送过时的状态数据
      const idx = RAW_DATA.allProjects.findIndex(x => x.id === id);
      if (idx !== -1) RAW_DATA.allProjects.splice(idx, 1);
      Object.keys(RAW_DATA.depts).forEach(d => {
        RAW_DATA.depts[d] = RAW_DATA.depts[d].filter(x => x.id !== id);
      });
      delete localEdits[id];
      delete archived[id];
      delete notes[`note_${id}`];
      delete checked[id];
      if (!deletedIds.includes(id)) deletedIds.push(id);
      localStorage.setItem('projectEdits', JSON.stringify(localEdits));
      localStorage.setItem('projectArchived', JSON.stringify(archived));
      localStorage.setItem('projectNotes', JSON.stringify(notes));
      localStorage.setItem('projectChecked', JSON.stringify(checked));
      localStorage.setItem('deletedIds', JSON.stringify(deletedIds));
      updateStats();
      renderTable();
      initResourceSearch();
      showSaved();
    }
  } else {
    // 非协作模式：本地逻辑（兼容）
    const idx = RAW_DATA.allProjects.findIndex(x => x.id === id);
    if (idx !== -1) RAW_DATA.allProjects.splice(idx, 1);
    Object.keys(RAW_DATA.depts).forEach(d => {
      RAW_DATA.depts[d] = RAW_DATA.depts[d].filter(x => x.id !== id);
    });
    RAW_DATA.delayedProjects = RAW_DATA.delayedProjects.filter(x => x.id !== id);
    delete localEdits[id];
    delete archived[id];
    delete notes[`note_${id}`];
    delete checked[id];
    if (!deletedIds.includes(id)) deletedIds.push(id);
    localStorage.setItem('projectEdits', JSON.stringify(localEdits));
    localStorage.setItem('projectArchived', JSON.stringify(archived));
    localStorage.setItem('projectNotes', JSON.stringify(notes));
    localStorage.setItem('projectChecked', JSON.stringify(checked));
    localStorage.setItem('deletedIds', JSON.stringify(deletedIds));
    updateStats();
    renderTable();
    initResourceSearch();
    showSaved();
  }
}

// 一键清理测试项目
function saveData() {
  localStorage.setItem('projectEdits', JSON.stringify(localEdits));
  localStorage.setItem('projectNotes', JSON.stringify(notes));
  localStorage.setItem('projectChecked', JSON.stringify(checked));
  showSaved();
}

async function resetAll() {
  if (confirm('确定要清空所有本地修改（备注、点检、计划编辑）吗？\\n（归档和已删除的项目不会受影响）')) {
    // 1. 清空本地修改状态（【关键修复】保留 archived 和 deletedIds，
    //    因为这些是已同步到Excel的持久化状态，不是"本地修改"）
    localEdits = {};
    notes = {};
    checked = {};
    // archived = {};  // ← 不再清空！归档是持久化状态
    // deletedIds = [];  // ← 不再清空！删除是持久化状态
    newProjects = [];
    customEmails = {};

    // 2. 从 localStorage 移除相关项（保留归档和删除状态）
    localStorage.removeItem('projectEdits');
    localStorage.removeItem('projectNotes');
    localStorage.removeItem('projectChecked');
    // localStorage.removeItem('projectArchived');  // ← 保留
    // localStorage.removeItem('deletedIds');  // ← 保留
    localStorage.removeItem('newProjects');
    localStorage.removeItem('customEmails');

    // 3. 【关键修复】从服务器获取最新数据并同步 Excel 中的归档标志
    //    重置后，Excel中标记为"已归档"的项目仍然应该在归档看板中
    await syncFromExcel();

    // 4. 协作模式：从服务器重新拉取数据（获取服务器端的协作状态）
    if (collabIsEnabled()) {
      await collabLoadData();
      // 服务器数据加载完成后，再次同步Excel归档状态（防止服务器覆盖）
      await syncFromExcel();
      updateStats();
      renderTable();
      initHoursPanel();
      initResourceSearch();
    } else {
      // 非协作模式：直接刷新
      updateStats();
      renderTable();
      initHoursPanel();
      initResourceSearch();
    }
  }
}

function showSaved() {
  const el = document.getElementById('saveIndicator');
  el.classList.remove('show');
  void el.offsetWidth;
  el.classList.add('show');
  // 协作模式：标记需要同步到服务器
  collabMarkDirty();
}

// 整理报表邮件 - 包含工时统计、资源看板、项目看板
function generateReportEmail() {
  const allData = RAW_DATA.allProjects
    .map(getProject)
    .map(p => ({ ...p, _days: recalcDays(p) }))
    .filter(p => !isProjectArchived(p));
  
  const todayStr = RAW_DATA.today;
  
  // ============ 1. 工时统计 ============
  // 使用默认范围：本周
  const today = new Date(RAW_DATA.today);
  const dayOfWeek = today.getDay() || 7;
  const weekStart = new Date(today);
  weekStart.setDate(today.getDate() - dayOfWeek + 1);
  const weekEnd = new Date(weekStart);
  weekEnd.setDate(weekStart.getDate() + 6);
  
  const startDate = weekStart;
  const endDate = weekEnd;
  const totalDays = Math.ceil((endDate - startDate) / (1000 * 60 * 60 * 24)) + 1;
  
  let totalHours = 0;
  const uniqueResources = new Set();
  const personHours = {};
  
  allData.forEach(p => {
    const hoursPerDay = parseFloat(p.日平均工时) || 0;
    if (hoursPerDay <= 0) return;
    const pStart = p.资源开始时间 ? new Date(p.资源开始时间) : null;
    const pEnd = p.资源结束时间 ? new Date(p.资源结束时间) : null;
    if (!pStart && !pEnd) return;
    const overlapStart = pStart ? (pStart > startDate ? pStart : startDate) : startDate;
    const overlapEnd = pEnd ? (pEnd < endDate ? pEnd : endDate) : endDate;
    if (overlapStart > overlapEnd) return;
    const overlapDays = Math.ceil((overlapEnd - overlapStart) / (1000 * 60 * 60 * 24)) + 1;
    const ph = overlapDays * hoursPerDay;
    totalHours += ph;
    const resName = p.资源名称 || '未分配';
    uniqueResources.add(resName);
    if (!personHours[resName]) personHours[resName] = 0;
    personHours[resName] += ph;
  });
  
  const resourceCount = uniqueResources.size;
  const avgPerDay = totalDays > 0 ? (totalHours / totalDays) : 0;
  const avgPerPerson = resourceCount > 0 && totalDays > 0 ? (totalHours / resourceCount / totalDays) : 0;
  
  const formatDate = (d) => d.toISOString().split('T')[0];
  
  // ============ 2. 资源看板（按人员） ============
  const byPerson = {};
  allData.forEach(p => {
    const name = p.资源名称 || '未分配';
    if (!byPerson[name]) byPerson[name] = [];
    byPerson[name].push(p);
  });
  
  // 按人员统计延期/正常
  const personStats = Object.keys(byPerson).sort().map(name => {
    const tasks = byPerson[name];
    const delayed = tasks.filter(t => t._days !== null && t._days < 0).length;
    const urgent = tasks.filter(t => t._days !== null && t._days >= 0 && t._days <= 3).length;
    const normal = tasks.filter(t => t._days === null || t._days > 3).length;
    return { name, total: tasks.length, delayed, urgent, normal, tasks };
  });
  
  // ============ 3. 项目看板（按部门-项目） ============
  const byDeptProj = {};
  allData.forEach(p => {
    const key = (p.部门 || '未分配') + '||' + (p.项目 || '未命名');
    if (!byDeptProj[key]) byDeptProj[key] = { dept: p.部门 || '未分配', project: p.项目 || '未命名', tasks: [] };
    byDeptProj[key].tasks.push(p);
  });
  
  const projectStats = Object.keys(byDeptProj).sort().map(key => {
    const info = byDeptProj[key];
    const tasks = info.tasks;
    const delayed = tasks.filter(t => t._days !== null && t._days < 0).length;
    const urgent = tasks.filter(t => t._days !== null && t._days >= 0 && t._days <= 3).length;
    const normal = tasks.filter(t => t._days === null || t._days > 3).length;
    return { ...info, delayed, urgent, normal };
  });
  
  // ============ 生成邮件内容 ============
  const subject = `【项目点检周报】${todayStr} 工时统计与项目进度`;
  
  let body = `项家祺 您好：\n\n`;
  body += `以下是${todayStr}的项目点检周报，包含工时统计、资源看板和项目看板，请查阅：\n\n`;
  
  // 工时统计
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n`;
  body += `📊 一、工时统计（${formatDate(startDate)} ~ ${formatDate(endDate)}）\n`;
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  body += `  统计范围：${totalDays} 天\n`;
  body += `  参与资源：${resourceCount} 人\n`;
  body += `  总工时投入：${totalHours.toFixed(1)} h\n`;
  body += `  日均投入：${avgPerDay.toFixed(1)} h/天\n`;
  body += `  单资源日均：${avgPerPerson.toFixed(1)} h/人/天\n\n`;
  
  if (Object.keys(personHours).length > 0) {
    body += `  个人工时排行：\n`;
    const sortedByHours = Object.keys(personHours).sort((a, b) => personHours[b] - personHours[a]);
    sortedByHours.slice(0, 10).forEach((name, i) => {
      body += `    ${i + 1}. ${name}：${personHours[name].toFixed(1)} h\n`;
    });
    body += `\n`;
  }
  
  // 资源看板
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n`;
  body += `👥 二、资源看板（共 ${personStats.length} 人）\n`;
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  
  personStats.forEach(ps => {
    const statusParts = [];
    if (ps.delayed > 0) statusParts.push(`🔴延期${ps.delayed}`);
    if (ps.urgent > 0) statusParts.push(`🟠即将到期${ps.urgent}`);
    if (ps.normal > 0) statusParts.push(`🟢正常${ps.normal}`);
    body += `  ${ps.name}（${ps.total}项）：${statusParts.join('、') || '无数据'}\n`;
    
    // 列出延期任务
    const delayedTasks = ps.tasks.filter(t => t._days !== null && t._days <= 3);
    if (delayedTasks.length > 0) {
      delayedTasks.slice(0, 3).forEach(t => {
        const daysLabel = t._days < 0 ? `已延期${Math.abs(t._days)}天` : `${t._days}天后到期`;
        body += `    · [${daysLabel}] ${t.部门}/${t.项目} - ${t.资源类型}\n`;
      });
    }
    body += `\n`;
  });
  
  // 项目看板
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n`;
  body += `📋 三、项目看板（共 ${projectStats.length} 个项目）\n`;
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  
  // 按部门分组显示
  const byDept = {};
  projectStats.forEach(ps => {
    if (!byDept[ps.dept]) byDept[ps.dept] = [];
    byDept[ps.dept].push(ps);
  });
  
  Object.keys(byDept).sort().forEach(dept => {
    const projects = byDept[dept];
    const totalDelayed = projects.reduce((s, p) => s + p.delayed, 0);
    const totalUrgent = projects.reduce((s, p) => s + p.urgent, 0);
    body += `【${dept}】${projects.length}个项目（🔴${totalDelayed} 🟠${totalUrgent}）\n`;
    projects.forEach(p => {
      const statusParts = [];
      if (p.delayed > 0) statusParts.push(`🔴${p.delayed}`);
      if (p.urgent > 0) statusParts.push(`🟠${p.urgent}`);
      if (p.normal > 0) statusParts.push(`🟢${p.normal}`);
      body += `  · ${p.project}（${p.tasks.length}项）：${statusParts.join(' ') || '无数据'}\n`;
    });
    body += `\n`;
  });
  
  body += `━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n`;
  body += `---\n`;
  body += `此邮件由项目点检系统自动生成\n`;
  body += `点检日期：${todayStr}\n`;
  body += `请参考超声波户用水表产品线研发群内钉钉文档「超声波户表脚本」\n`;
  
  // 获取项家祺的邮箱
  const allEmails = getAllEmails();
  const targetEmail = allEmails['项家祺'] || 'xiangjiaqi@hxgroup.com';
  
  // 生成预览弹窗
  const modal = document.createElement('div');
  modal.className = 'modal-overlay';
  modal.style.display = 'flex';
  modal.style.alignItems = 'center';
  modal.style.justifyContent = 'center';
  
  modal.innerHTML = `
    <div style="background:white;padding:30px;border-radius:12px;max-width:600px;max-height:85vh;overflow:auto;box-shadow:0 20px 60px rgba(0,0,0,0.3)">
      <h2 style="margin-bottom:20px;color:#1f2937">📧 整理报表邮件</h2>
      
      <div style="margin-bottom:16px">
        <div style="color:#6b7280;font-size:13px;margin-bottom:4px">收件人</div>
        <div style="padding:8px 12px;background:#f3f4f6;border-radius:6px;font-weight:500">项家祺 (${targetEmail})</div>
      </div>
      
      <div style="margin-bottom:16px">
        <div style="color:#6b7280;font-size:13px;margin-bottom:4px">抄送（点击选择成员）</div>
        <div id="ccSelectorContainer"></div>
        <div id="ccEmailDisplay" style="margin-top:8px;padding:8px 12px;background:#eff6ff;border-radius:6px;font-size:12px;color:#1e40af;display:none"></div>
      </div>
      
      <div style="margin-bottom:16px">
        <div style="color:#6b7280;font-size:13px;margin-bottom:4px">邮件主题</div>
        <div style="padding:8px 12px;background:#f3f4f6;border-radius:6px;font-weight:500">${subject}</div>
      </div>
      
      <div style="margin-bottom:20px">
        <div style="color:#6b7280;font-size:13px;margin-bottom:4px">邮件内容</div>
        <div style="padding:12px;background:#f9fafb;border-radius:6px;white-space:pre-wrap;font-family:monospace;font-size:12px;line-height:1.6;max-height:300px;overflow:auto;border:1px solid #e5e7eb">${body.replace(/</g, '&lt;').replace(/>/g, '&gt;')}</div>
      </div>
      
      <div style="display:flex;gap:10px;justify-content:flex-end">
        <button class="btn" style="background:#e5e7eb;color:#374151" onclick="this.closest('.modal-overlay').remove()">关闭</button>
        <button class="btn" style="background:#10b981;color:white" onclick="openReportEmailClient()">📨 打开邮件客户端发送</button>
      </div>
    </div>
  `;
  
  document.body.appendChild(modal);
  modal.addEventListener('click', (e) => {
    if (e.target === modal) modal.remove();
  });
  
  // 渲染抄送选择器
  renderCCSelector();
  
  // 保存邮件内容供打开客户端使用
  window._reportEmailData = { subject, body, targetEmail };
}

function openReportEmailClient() {
  if (!window._reportEmailData) return;
  const { subject, body, targetEmail } = window._reportEmailData;
  
  const ccList = (CC_EMAILS && CC_EMAILS.length > 0) ? CC_EMAILS.join(',') : '';
  
  let mailtoUrl = 'mailto:' + encodeURIComponent(targetEmail);
  const params = [];
  if (ccList) params.push('cc=' + encodeURIComponent(ccList));
  params.push('subject=' + encodeURIComponent(subject));
  params.push('body=' + encodeURIComponent(body));
  mailtoUrl += '?' + params.join('&');
  
  console.log('报表邮件 mailto URL:', mailtoUrl);
  console.log('抄送列表:', ccList);
  console.log('抄送数量:', CC_EMAILS ? CC_EMAILS.length : 0);
  
  window.location.href = mailtoUrl;
}

function exportCSV() {
  // 导出所有未归档的项目，按原始任务计划表格式（23列）
  const allData = RAW_DATA.allProjects
    .map(getProject)
    .filter(p => !isProjectArchived(p))
    .map(p => ({ ...p, _days: recalcDays(p) }));
  
  // 按部门-项目-资源开始时间排序（与原始Excel一致）
  allData.sort((a, b) => {
    if (a.部门 !== b.部门) return (a.部门 || '').localeCompare(b.部门 || '');
    if (a.项目 !== b.项目) return a.项目.localeCompare(b.项目);
    return (a.资源开始时间 || '').localeCompare(b.资源开始时间 || '');
  });
  
  // 生成CSV内容 - 按原始任务计划表23列格式
  let rows = [];
  
  // 行0：元数据行（原始格式，列5有值，其余为空）
  let row0 = new Array(23).fill('');
  row0[5] = '';  // 原始Excel这里有个数值，导出时空着
  rows.push(row0);
  
  // 行1：第一层表头
  let row1 = new Array(23).fill('');
  row1[5] = '项目';
  row1[6] = '开始时间';
  row1[7] = '结束时间';
  row1[8] = '项目描述';
  row1[9] = '资源描述';
  rows.push(row1);
  
  // 行2：第二层表头
  let row2 = new Array(23).fill('');
  row2[9] = '资源类型';
  row2[10] = '资源名称';
  row2[11] = '资源开始时间';
  row2[12] = '资源结束时间';
  row2[13] = '日平均工时投入\\n（h）';
  rows.push(row2);
  
  // 数据行：按部门和项目分组，同一项目的后续行部门和项目信息留空
  let lastDept = '';
  let lastProject = '';
  
  allData.forEach(p => {
    let row = new Array(23).fill('');
    
    // 列4：部门（新部门才填）
    if (p.部门 && p.部门 !== lastDept) {
      row[4] = p.部门;
      lastDept = p.部门;
      // 部门变化时项目也重置
      lastProject = '';
    }
    
    // 列5-8：项目信息（新项目才填）
    if (p.项目 && p.项目 !== lastProject) {
      row[5] = p.项目;
      row[6] = p.项目开始时间 || '';
      row[7] = p.项目结束时间 || '';
      row[8] = p.项目描述 || '';
      lastProject = p.项目;
    }
    
    // 列9-13：资源信息（每行都填）
    row[9] = p.资源类型 || '';
    row[10] = p.资源名称 || '';
    row[11] = p.资源开始时间 || '';
    row[12] = p.资源结束时间 || '';
    row[13] = p.日平均工时 || 0;
    
    // 列14-16：重复列（资源名称、开始、结束）
    row[14] = p.资源名称 || '';
    row[15] = p.资源开始时间 || '';
    row[16] = p.资源结束时间 || '';
    
    // 列17-22：数值列（保持为空）
    rows.push(row);
  });
  
  // 转换为CSV字符串
  let csv = rows.map(row => row.map(cell => {
    const s = String(cell ?? '');
    if (s.includes(',') || s.includes('"') || s.includes('\\n')) {
      return '"' + s.replace(/"/g, '""') + '"';
    }
    return s;
  }).join(',')).join('\\n');
  
  const blob = new Blob(['\ufeff' + csv], { type: 'text/csv;charset=utf-8;' });
  const link = document.createElement('a');
  link.href = URL.createObjectURL(blob);
  link.download = `任务计划表_${RAW_DATA.today}.csv`;
  link.click();
  alert(`✅ 已导出 ${allData.length} 条资源记录到 CSV！\n格式与原始任务计划表一致（23列）。`);
}

function exportShareData() {
  const shareData = {
    version: '1.1',
    exportDate: new Date().toISOString(),
    localEdits: localEdits,
    notes: notes,
    checked: checked,
    archived: archived,
    customEmails: customEmails,
    description: '项目点检表共享数据 - 包含所有备注、点检、编辑、归档记录和成员邮箱'
  };
  const blob = new Blob([JSON.stringify(shareData, null, 2)], { type: 'application/json' });
  const link = document.createElement('a');
  link.href = URL.createObjectURL(blob);
  link.download = `项目点检共享数据_${RAW_DATA.today}.json`;
  link.click();
  alert('✅ 共享数据已导出！\\n\\n将此 JSON 文件和 HTML 文件一起发给您的朋友，\\n对方打开 HTML 后点击「📥 导入共享数据」选择此文件即可同步所有修改（包括成员邮箱）。');
}

// 将所有数据同步到Excel（按原始任务计划表格式生成.xlsx文件）
function syncProjectsToExcel() {
  // 获取所有未归档的项目数据（应用本地编辑）
  const allData = RAW_DATA.allProjects
    .map(p => getProject(p))
    .filter(p => !isProjectArchived(p))
    .map(p => ({ ...p, _days: recalcDays(p) }));
  
  // 按部门-项目-资源开始时间排序（与原始Excel一致）
  allData.sort((a, b) => {
    if (a.部门 !== b.部门) return (a.部门 || '').localeCompare(b.部门 || '');
    if (a.项目 !== b.项目) return a.项目.localeCompare(b.项目);
    return (a.资源开始时间 || '').localeCompare(b.资源开始时间 || '');
  });
  
  // 生成任务计划表的数据行（23列）
  let taskRows = [];
  // 记录需要合并的单元格：{startRow, endRow, col} (0-based)
  let mergeCells = [];
  
  // 行0：元数据行
  let row0 = new Array(23).fill('');
  taskRows.push(row0);
  
  // 行1：第一层表头
  let row1 = new Array(23).fill('');
  row1[5] = '项目';
  row1[6] = '开始时间';
  row1[7] = '结束时间';
  row1[8] = '项目描述';
  row1[9] = '资源描述';
  taskRows.push(row1);
  
  // 行2：第二层表头
  let row2 = new Array(23).fill('');
  row2[9] = '资源类型';
  row2[10] = '资源名称';
  row2[11] = '资源开始时间';
  row2[12] = '资源结束时间';
  row2[13] = '日平均工时投入（h）';
  taskRows.push(row2);
  
  // 数据行
  let lastDept = '';
  let lastProject = '';
  let lastProjectDesc = '';
  let deptStartRow = -1;  // 部门合并起始行（0-based，相对于taskRows）
  let projStartRow = -1;  // 项目合并起始行
  let descStartRow = -1;  // 项目描述合并起始行
  
  const HEADER_ROWS = 3;  // 表头行数
  
  allData.forEach((p, idx) => {
    let row = new Array(23).fill('');
    const currentRowIdx = taskRows.length;  // 当前行在taskRows中的索引
    
    // 列4：部门（新部门才填）
    if (p.部门 && p.部门 !== lastDept) {
      // 如果之前有未结束的部门合并，先记录
      if (deptStartRow !== -1 && currentRowIdx - 1 > deptStartRow) {
        mergeCells.push({ startRow: deptStartRow, endRow: currentRowIdx - 1, col: 4 });
      }
      row[4] = p.部门;
      lastDept = p.部门;
      lastProject = '';
      deptStartRow = currentRowIdx;
      // 部门变化时，之前的项目合并也要结束
      if (projStartRow !== -1 && currentRowIdx - 1 > projStartRow) {
        mergeCells.push({ startRow: projStartRow, endRow: currentRowIdx - 1, col: 5 });
      }
      projStartRow = -1;
    }
    
    // 列5-8：项目信息（新项目才填）
    if (p.项目 && p.项目 !== lastProject) {
      // 如果之前有未结束的项目合并，先记录
      if (projStartRow !== -1 && currentRowIdx - 1 > projStartRow) {
        mergeCells.push({ startRow: projStartRow, endRow: currentRowIdx - 1, col: 5 });
      }
      // 结束之前的项目描述合并
      if (descStartRow !== -1 && currentRowIdx - 1 > descStartRow) {
        mergeCells.push({ startRow: descStartRow, endRow: currentRowIdx - 1, col: 8 });
      }
      row[5] = p.项目;
      row[6] = p.项目开始时间 || '';
      row[7] = p.项目结束时间 || '';
      row[8] = p.项目描述 || '';
      lastProject = p.项目;
      lastProjectDesc = p.项目描述 || '';
      projStartRow = currentRowIdx;
      descStartRow = p.项目描述 ? currentRowIdx : -1;
    } else if (p.项目 && p.项目 === lastProject) {
      // 同一项目，检查项目描述是否变化
      const curDesc = p.项目描述 || '';
      if (curDesc !== lastProjectDesc) {
        // 描述变化，结束之前的合并
        if (descStartRow !== -1 && currentRowIdx - 1 > descStartRow) {
          mergeCells.push({ startRow: descStartRow, endRow: currentRowIdx - 1, col: 8 });
        }
        row[8] = curDesc;
        lastProjectDesc = curDesc;
        descStartRow = curDesc ? currentRowIdx : -1;
      }
    }
    
    // 列9-13：资源信息
    row[9] = p.资源类型 || '';
    row[10] = p.资源名称 || '';
    row[11] = p.资源开始时间 || '';
    row[12] = p.资源结束时间 || '';
    row[13] = p.日平均工时 || 0;
    
    // 列14-16：重复列
    row[14] = p.资源名称 || '';
    row[15] = p.资源开始时间 || '';
    row[16] = p.资源结束时间 || '';
    
    taskRows.push(row);
  });
  
  // 结束时处理最后一个部门、项目和项目描述的合并
  const lastRowIdx = taskRows.length - 1;
  if (deptStartRow !== -1 && lastRowIdx > deptStartRow) {
    mergeCells.push({ startRow: deptStartRow, endRow: lastRowIdx, col: 4 });
  }
  if (projStartRow !== -1 && lastRowIdx > projStartRow) {
    mergeCells.push({ startRow: projStartRow, endRow: lastRowIdx, col: 5 });
  }
  if (descStartRow !== -1 && lastRowIdx > descStartRow) {
    mergeCells.push({ startRow: descStartRow, endRow: lastRowIdx, col: 8 });
  }
  
  // 使用SpreadsheetML格式生成Excel（支持多sheet和单元格合并）
  const sheet2Xml = generateSheetXml('任务计划表', taskRows, mergeCells);
  
  // 人员资源情况sheet（简单格式）
  const personRows = [
    ['开始时间', '结束时间', '人员', '工时']
  ];
  
  // 按人员汇总
  const personMap = {};
  allData.forEach(p => {
    const name = p.资源名称 || '未分配';
    if (!personMap[name]) personMap[name] = { start: '', end: '', hours: 0 };
    if (p.资源开始时间 && (!personMap[name].start || p.资源开始时间 < personMap[name].start)) {
      personMap[name].start = p.资源开始时间;
    }
    if (p.资源结束时间 && (!personMap[name].end || p.资源结束时间 > personMap[name].end)) {
      personMap[name].end = p.资源结束时间;
    }
    personMap[name].hours += parseFloat(p.日平均工时) || 0;
  });
  
  Object.keys(personMap).sort().forEach(name => {
    const info = personMap[name];
    personRows.push([info.start, info.end, name, info.hours.toFixed(1)]);
  });
  
  const sheet1Xml = generateSheetXml('人员资源情况', personRows);
  
  // 完整的SpreadsheetML XML
  const xmlContent = `<?xml version="1.0" encoding="UTF-8"?>
<?mso-application progid="Excel.Sheet"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:html="http://www.w3.org/TR/REC-html40">
 <Styles>
  <Style ss:ID="Default" ss:Name="Normal">
   <Alignment ss:Vertical="Center"/>
   <Borders/>
   <Font/>
   <Interior/>
   <NumberFormat/>
   <Protection/>
  </Style>
  <Style ss:ID="Header">
   <Font ss:Bold="1"/>
   <Interior ss:Color="#D9E1F2" ss:Pattern="Solid"/>
  </Style>
 </Styles>
 ${sheet1Xml}
 ${sheet2Xml}
 <Worksheet ss:Name="Sheet3">
  <Table ss:ExpandedColumnCount="1" ss:ExpandedRowCount="1" x:FullColumns="1" x:FullRows="1">
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
  </Table>
 </Worksheet>
</Workbook>`;
  
  const blob = new Blob(['\ufeff' + xmlContent], { type: 'application/vnd.ms-excel;charset=utf-8;' });
  const link = document.createElement('a');
  link.href = URL.createObjectURL(blob);
  link.download = `超声波户表脚本_${RAW_DATA.today}.xls`;
  link.click();
  
  alert(`✅ 已生成 Excel 文件！\n\n共 ${allData.length} 条资源记录\n格式与原始「超声波户表脚本」一致\n包含：人员资源情况、任务计划表、Sheet3 三个Sheet`);
}

// 辅助函数：生成单个Sheet的XML（支持单元格合并）
// mergeCells: [{startRow, endRow, col}] (0-based)
function generateSheetXml(sheetName, rows, mergeCells) {
  mergeCells = mergeCells || [];
  const maxCols = Math.max(...rows.map(r => r.length));
  
  // 构建合并映射：key = "row,col" -> mergeDown 数量
  const mergeMap = {};
  mergeCells.forEach(mc => {
    // 只在起始行设置 MergeDown
    const key = mc.startRow + ',' + mc.col;
    mergeMap[key] = mc.endRow - mc.startRow;
  });
  
  let rowsXml = rows.map((row, rowIdx) => {
    let cellsXml = row.map((cell, colIdx) => {
      // 检查是否是合并单元格的一部分（非起始行，需要跳过）
      const isMergedChild = mergeCells.some(mc => 
        mc.col === colIdx && rowIdx > mc.startRow && rowIdx <= mc.endRow
      );
      if (isMergedChild) return '';  // 被合并的单元格不输出
      
      const val = String(cell ?? '');
      const isNum = !isNaN(val) && val !== '' && !isNaN(parseFloat(val));
      
      let cellAttrs = `ss:Index="${colIdx + 1}"`;
      
      // 检查是否需要设置 MergeDown
      const mergeKey = rowIdx + ',' + colIdx;
      if (mergeMap[mergeKey] !== undefined) {
        cellAttrs += ` ss:MergeDown="${mergeMap[mergeKey]}"`;
      }
      
      if (isNum) {
        return `<Cell ${cellAttrs}><Data ss:Type="Number">${val}</Data></Cell>`;
      }
      return `<Cell ${cellAttrs}><Data ss:Type="String">${val.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}</Data></Cell>`;
    }).join('');
    return `<Row>${cellsXml}</Row>`;
  }).join('\\n  ');
  
  return `<Worksheet ss:Name="${sheetName}">
  <Table ss:ExpandedColumnCount="${maxCols}" ss:ExpandedRowCount="${rows.length}" x:FullColumns="1" x:FullRows="1">
   ${rowsXml}
  </Table>
 </Worksheet>`;
}

function importShareData(event) {
  const file = event.target.files[0];
  if (!file) return;
  
  const reader = new FileReader();
  reader.onload = function(e) {
    try {
      const data = JSON.parse(e.target.result);
      if (!data.version) {
        alert('❌ 文件格式不正确，请确认是通过「导出共享数据」生成的文件。');
        return;
      }
      if (data.localEdits) localEdits = data.localEdits;
      if (data.notes) notes = data.notes;
      if (data.checked) checked = data.checked;
      if (data.archived) archived = data.archived;
      if (data.customEmails) {
        customEmails = { ...customEmails, ...data.customEmails };
        localStorage.setItem('customEmails', JSON.stringify(customEmails));
      }
      localStorage.setItem('projectEdits', JSON.stringify(localEdits));
      localStorage.setItem('projectNotes', JSON.stringify(notes));
      localStorage.setItem('projectChecked', JSON.stringify(checked));
      localStorage.setItem('projectArchived', JSON.stringify(archived));
      mergeLocalNewProjects();  // 合并导入的新增项目
      updateStats();
      renderTable();
      let msg = '✅ 共享数据导入成功！\\n\\n所有备注、点检记录、编辑内容和归档状态已同步。';
      if (data.customEmails && Object.keys(data.customEmails).length > 0) {
        msg += '\\n\\n同时已同步 ' + Object.keys(data.customEmails).length + ' 位成员的邮箱信息。';
      }
      alert(msg);
    } catch (err) {
      alert('❌ 导入失败：' + err.message);
    }
  };
  reader.readAsText(file);
  event.target.value = '';
}

// ==================== 新增：成员邮箱管理 ====================
// 自定义成员邮箱（本地存储，同事导入的新成员）

// 获取完整的工程师邮箱列表（内置 + 自定义）
function getAllEmails() {
  return { ...ENGINEER_EMAILS, ...customEmails };
}

// 中文姓氏拼音对照表（常见姓氏）
const SURNAME_PINYIN = {
  '王':'wang','李':'li','张':'zhang','刘':'liu','陈':'chen','杨':'yang','赵':'zhao',
  '黄':'huang','周':'zhou','吴':'wu','徐':'xu','孙':'sun','胡':'hu','朱':'zhu',
  '高':'gao','林':'lin','何':'he','郭':'guo','马':'ma','罗':'luo','梁':'liang',
  '宋':'song','郑':'zheng','谢':'xie','韩':'han','唐':'tang','冯':'feng','于':'yu',
  '董':'dong','萧':'xiao','程':'cheng','曹':'cao','袁':'yuan','邓':'deng','许':'xu',
  '傅':'fu','沈':'shen','曾':'zeng','彭':'peng','吕':'lv','苏':'su','卢':'lu',
  '蒋':'jiang','蔡':'cai','贾':'jia','丁':'ding','魏':'wei','薛':'xue','叶':'ye',
  '阎':'yan','余':'yu','潘':'pan','杜':'du','戴':'dai','夏':'xia','钟':'zhong',
  '汪':'wang','田':'tian','任':'ren','姜':'jiang','范':'fan','方':'fang','石':'shi',
  '姚':'yao','谭':'tan','廖':'liao','邹':'zou','熊':'xiong','金':'jin','陆':'lu',
  '郝':'hao','孔':'kong','白':'bai','崔':'cui','康':'kang','毛':'mao','邱':'qiu',
  '秦':'qin','江':'jiang','史':'shi','顾':'gu','侯':'hou','邵':'shao','孟':'meng',
  '龙':'long','万':'wan','段':'duan','钱':'qian','汤':'tang','尹':'yin','黎':'li',
  '易':'yi','常':'chang','武':'wu','乔':'qiao','贺':'he','赖':'lai','龚':'gong',
  '文':'wen','庞':'pang','樊':'fan','兰':'lan','殷':'yin','施':'shi','陶':'tao',
  '洪':'hong','翟':'zhai','安':'an','颜':'yan','倪':'ni','严':'yan','牛':'niu',
  '温':'wen','芦':'lu','季':'ji','俞':'yu','章':'zhang','鲁':'lu','葛':'ge',
  '韦':'wei','申':'shen','尤':'you','毕':'bi','聂':'nie','丛':'cong','焦':'jiao',
  '向':'xiang','柳':'liu','邢':'xing','路':'lu','岳':'yue','齐':'qi','沿':'yan',
  '梅':'mei','莫':'mo','庄':'zhuang','辛':'xin','管':'guan','祝':'zhu','左':'zuo',
  '涂':'tu','谷':'gu','祁':'qi','时':'shi','舒':'shu','耿':'geng','牟':'mou',
  '卜':'bu','上官':'shangguan','欧阳':'ouyang','夏侯':'xiahou','诸葛':'zhuge',
  '闻人':'wenren','东方':'dongfang','赫连':'helian','皇甫':'huangfu','尉迟':'yuchi',
  '公羊':'gongyang','澹台':'tantai','公冶':'gongye','宗政':'zongzheng','濮阳':'puyang',
  '淳于':'chunyu','单于':'chanyu','太叔':'taishu','申屠':'shentu','公孙':'gongsun',
  '仲孙':'zhongsun','轩辕':'xuanyuan','令狐':'linghu','钟离':'zhongli','宇文':'yuwen',
  '长孙':'zhangsun','慕容':'murong','鲜于':'xianyu','闾丘':'lvqiu','司徒':'situ',
  '司空':'sikong','亓官':'qiguan','司寇':'sikou','仉':'zhang','督':'du','子车':'juche',
  '颛孙':'zhuansun','端木':'duanmu','巫马':'wuma','公西':'gongxi','漆雕':'qidiao',
  '乐正':'yuezheng','壤驷':'rangsi','公良':'gongliang','拓跋':'tuoba','夹谷':'jiagu',
  '宰父':'zaifu','谷梁':'guliang','晋':'jin','楚':'chu','闫':'yan','法':'fa',
  '汝':'ru','鄢':'yan','涂':'tu','钦':'qin','段干':'duangan','百里':'baili',
  '东郭':'dongguo','南门':'nanmen','呼延':'huyan','归':'gui','海':'hai','羊舌':'yangshe',
  '微生':'weisheng','帅':'shuai','缑':'gou','亢':'kang','况':'kuang','后':'hou',
  '有':'you','琴':'qin','梁丘':'liangqiu','左丘':'zuoqiu','东门':'dongmen',
  '西门':'ximen','商':'shang','牟':'mou','佘':'she','佴':'nai','伯':'bo','赏':'shang',
  '南宫':'nangong','墨':'mo','哈':'ha','谯':'qiao','笪':'da','年':'nian','爱':'ai',
  '阳':'yang','佟':'tong','答':'da','项':'xiang','危':'wei','童':'tong','毛':'mao',
  '苑':'yuan','鲍':'bao','华':'hua','岑':'cen','滕':'teng','殷':'yin','罗':'luo','毕':'bi',
  '郝':'hao','邬':'wu','乐':'le','于':'yu','时':'shi','傅':'fu','皮':'pi','卞':'bian',
  '齐':'qi','康':'kang','伍':'wu','余':'yu','元':'yuan','卜':'bu','顾':'gu','孟':'meng',
  '黄':'huang','邵':'shao','戚':'qi','谢':'xie','邹':'zou','喻':'yu','柏':'bai','水':'shui',
  '窦':'dou','章':'zhang','云':'yun','苏':'su','潘':'pan','葛':'ge','奚':'xi','范':'fan',
  '彭':'peng','郎':'lang','鲁':'lu','韦':'wei','昌':'chang','马':'ma','苗':'miao','凤':'feng',
  '花':'hua','方':'fang','俞':'yu','任':'ren','袁':'yuan','柳':'liu','酆':'feng','鲍':'bao',
  '史':'shi','唐':'tang','费':'fei','廉':'lian','岑':'cen','薛':'xue','雷':'lei','贺':'he',
  '倪':'ni','汤':'tang','滕':'teng','殷':'yin','罗':'luo','毕':'bi','郝':'hao','邬':'wu'
};

// 中文常用字拼音库（用于名字拼音生成）
const CHAR_PINYIN = {
  // A
  '阿':'a','啊':'a','埃':'ai','哀':'ai','哎':'ai','唉':'ai','矮':'ai','爱':'ai','碍':'ai','安':'an','按':'an','暗':'an','岸':'an','案':'an',
  // B
  '八':'ba','巴':'ba','拔':'ba','把':'ba','爸':'ba','霸':'ba','白':'bai','百':'bai','拜':'bai','班':'ban','搬':'ban','板':'ban','半':'ban','办':'ban','版':'ban',
  '半':'ban','伴':'ban','瓣':'ban','邦':'bang','帮':'bang','榜':'bang','傍':'bang','棒':'bang','包':'bao','宝':'bao','保':'bao','报':'bao','抱':'bao','豹':'bao','暴':'bao','爆':'bao','杯':'bei','悲':'bei','碑':'bei','北':'bei','贝':'bei','备':'bei','背':'bei','辈':'bei','贝':'bei','奔':'ben','本':'ben','笨':'ben','崩':'beng','绷':'beng','甭':'beng','蹦':'beng','逼':'bi','鼻':'bi','鼻':'bi','比':'bi','笔':'bi','彼':'bi','碧':'bi','蔽':'bi','毕':'bi','闭':'bi','必':'bi','辟':'bi','壁':'bi','臂':'bi','避':'bi','陛':'bi','鞭':'bian','边':'bian','编':'bian','贬':'bian','扁':'bian','便':'bian','变':'bian','卞':'bian','遍':'bian','辨':'bian','辩':'bian','辫':'bian','标':'biao','彪':'biao','表':'biao','别':'bie','宾':'bin','滨':'bin','彬':'bin','冰':'bing','兵':'bing','丙':'bing','柄':'bing','炳':'bing','病':'bing','并':'bing','拨':'bo','波':'bo','博':'bo','伯':'bo','剥':'bo','播':'bo','伯':'bo','驳':'bo','泊':'bo','勃':'bo','博':'bo','搏':'bo','薄':'bo','卜':'bu','补':'bu','捕':'bu','不':'bu','布':'bu','步':'bu','部':'bu','埠':'bu',
  // C
  '擦':'ca','猜':'cai','才':'cai','材':'cai','财':'cai','裁':'cai','采':'cai','彩':'cai','菜':'cai','蔡':'cai','参':'can','餐':'can','残':'can','蚕':'can','惨':'can','灿':'can','仓':'cang','苍':'cang','舱':'cang','藏':'cang','操':'cao','曹':'cao','草':'cao','册':'ce','测':'ce','侧':'ce','策':'ce','层':'ceng','曾':'ceng','插':'cha','茶':'cha','查':'cha','搽':'cha','察':'cha','岔':'cha','差':'cha','拆':'chai','柴':'chai','搀':'chan','掺':'chan','蝉':'chan','馋':'chan','缠':'chan','产':'chan','铲':'chan','颤':'chan','昌':'chang','猖':'chang','长':'chang','尝':'chang','常':'chang','偿':'chang','唱':'chang','倡':'chang','抄':'chao','超':'chao','巢':'chao','朝':'chao','潮':'chao','吵':'chao','炒':'chao','车':'che','扯':'che','彻':'che','撤':'che','尘':'chen','辰':'chen','晨':'chen','沉':'chen','陈':'chen','趁':'chen','衬':'chen','撑':'cheng','称':'cheng','城':'cheng','诚':'cheng','承':'cheng','乘':'cheng','程':'cheng','惩':'cheng','秤':'cheng','吃':'chi','痴':'chi','持':'chi','池':'chi','迟':'chi','驰':'chi','弛':'chi','耻':'chi','齿':'chi','侈':'chi','尺':'chi','赤':'chi','翅':'chi','斥':'chi','炽':'chi','充':'chong','冲':'chong','虫':'chong','崇':'chong','宠':'chong','抽':'chou','仇':'chou','筹':'chou','畴':'chou','踌':'chou','愁':'chou','筹':'chou','酬':'chou','瞅':'chou','丑':'chou','臭':'chou','出':'chu','初':'chu','除':'chu','础':'chu','储':'chu','楚':'chu','处':'chu','畜':'chu','触':'chu','川':'chuan','穿':'chuan','传':'chuan','船':'chuan','喘':'chuan','串':'chuan','疮':'chuang','窗':'chuang','床':'chuang','创':'chuang','吹':'chui','炊':'chui','垂':'chui','锤':'chui','春':'chun','春':'chun','纯':'chun','唇':'chun','醇':'chun','蠢':'chun','戳':'chuo','绰':'chuo','疵':'ci','差':'ci','词':'ci','瓷':'ci','慈':'ci','雌':'ci','辞':'ci','磁':'ci','雌':'ci','此':'ci','次':'ci','刺':'ci','赐':'ci','葱':'cong','聪':'cong','从':'cong','丛':'cong','凑':'cou','粗':'cu','促':'cu','醋':'cu','蹿':'cuan','窜':'cuan','篡':'cuan','崔':'cui','催':'cui','摧':'cui','脆':'cui','翠':'cui','粹':'cui','村':'cun','存':'cun','寸':'cun','措':'cuo','错':'cuo','挫':'cuo',
  // D
  '搭':'da','答':'da','达':'da','打':'da','大':'da','呆':'dai','待':'dai','逮':'dai','戴':'dai','带':'dai','代':'dai','袋':'dai','待':'dai','贷':'dai','黛':'dai','丹':'dan','担':'dan','单':'dan','耽':'dan','郸':'dan','胆':'dan','旦':'dan','但':'dan','担':'dan','弹':'dan','淡':'dan','蛋':'dan','诞':'dan','弹':'dan','淡':'dan','当':'dang','挡':'dang','党':'dang','档':'dang','刀':'dao','导':'dao','岛':'dao','倒':'dao','盗':'dao','道':'dao','稻':'dao','到':'dao','悼':'dao','得':'de','德':'de','的':'de','灯':'deng','登':'deng','等':'deng','凳':'deng','邓':'deng','低':'di','堤':'di','滴':'di','迪':'di','敌':'di','笛':'di','狄':'di','涤':'di','底':'di','抵':'di','地':'di','弟':'di','帝':'di','第':'di','递':'di','蒂':'di','缔':'di','颠':'dian','掂':'dian','滇':'dian','颠':'dian','典':'dian','点':'dian','碘':'dian','电':'dian','店':'dian','殿':'dian','垫':'dian','淀':'dian','奠':'dian','殿':'dian','雕':'diao','刁':'diao','叼':'diao','掉':'diao','吊':'diao','钓':'diao','调':'diao','爹':'die','跌':'die','叠':'die','碟':'die','蝶':'die','丁':'ding','叮':'ding','盯':'ding','钉':'ding','顶':'ding','鼎':'ding','定':'ding','订':'ding','丢':'diu','东':'dong','冬':'dong','董':'dong','懂':'dong','动':'dong','洞':'dong','栋':'dong','冻':'dong','都':'du','督':'du','毒':'du','独':'du','读':'du','犊':'du','独':'du','堵':'du','睹':'du','赌':'du','杜':'du','肚':'du','度':'du','渡':'du','镀':'du','端':'duan','短':'duan','断':'duan','段':'duan','缎':'duan','堆':'dui','队':'dui','对':'dui','兑':'dui','怼':'dui','墩':'dun','吨':'dun','蹲':'dun','敦':'dun','盾':'dun','钝':'dun','顿':'dun','多':'duo','夺':'duo','朵':'duo','躲':'duo','堕':'duo','舵':'duo',
  // E
  '蛾':'e','鹅':'e','俄':'e','额':'e','讹':'e','恶':'e','饿':'e','恩':'en','而':'er','儿':'er','耳':'er','尔':'er','饵':'er','洱':'er','二':'er','贰':'er',
  // F
  '发':'fa','罚':'fa','伐':'fa','乏':'fa','法':'fa','阀':'fa','法':'fa','帆':'fan','番':'fan','翻':'fan','凡':'fan','烦':'fan','繁':'fan','反':'fan','返':'fan','犯':'fan','泛':'fan','范':'fan','贩':'fan','方':'fang','芳':'fang','方':'fang','防':'fang','妨':'fang','房':'fang','肪':'fang','防':'fang','仿':'fang','访':'fang','放':'fang','飞':'fei','非':'fei','妃':'fei','肥':'fei','腓':'fei','匪':'fei','诽':'fei','肺':'fei','沸':'fei','费':'fei','废':'fei','沸':'fei','分':'fen','芬':'fen','吩':'fen','氛':'fen','纷':'fen','坟':'fen','焚':'fen','粉':'fen','奋':'fen','份':'fen','忿':'fen','愤':'fen','粪':'fen','丰':'feng','风':'feng','封':'feng','枫':'feng','疯':'feng','蜂':'feng','峰':'feng','锋':'feng','冯':'feng','逢':'feng','凤':'feng','奉':'feng','俸':'feng','佛':'fo','否':'fou','夫':'fu','肤':'fu','肤':'fu','服':'fu','浮':'fu','符':'fu','幅':'fu','福':'fu','蝠':'fu','抚':'fu','辅':'fu','俯':'fu','腐':'fu','府':'fu','父':'fu','赴':'fu','付':'fu','妇':'fu','负':'fu','附':'fu','富':'fu','副':'fu','赋':'fu','复':'fu','傅':'fu','富':'fu','腹':'fu','覆':'fu',
  // G
  '该':'gai','改':'gai','盖':'gai','概':'gai','钙':'gai','溉':'gai','干':'gan','甘':'gan','杆':'gan','肝':'gan','赶':'gan','感':'gan','敢':'gan','感':'gan','刚':'gang','钢':'gang','岗':'gang','港':'gang','纲':'gang','杠':'gang','高':'gao','膏':'gao','糕':'gao','搞':'gao','稿':'gao','告':'gao','诰':'gao','戈':'ge','哥':'ge','歌':'ge','阁':'ge','革':'ge','葛':'ge','格':'ge','蛤':'ge','隔':'ge','个':'ge','各':'ge','铬':'ge','给':'gei','根':'gen','跟':'gen','亘':'gen','更':'geng','耕':'geng','庚':'geng','羹':'geng','埂':'geng','耿':'geng','梗':'geng','工':'gong','弓':'gong','公':'gong','功':'gong','攻':'gong','供':'gong','宫':'gong','恭':'gong','龚':'gong','躬':'gong','公':'gong','巩':'gong','拱':'gong','共':'gong','贡':'gong','勾':'gou','钩':'gou','沟':'gou','苟':'gou','狗':'gou','构':'gou','购':'gou','够':'gou','购':'gou','构':'gou','购':'gou','辜':'gu','姑':'gu','辜':'gu','孤':'gu','沽':'gu','菇':'gu','咕':'gu','箍':'gu','估':'gu','股':'gu','骨':'gu','谷':'gu','蛊':'gu','鼓':'gu','古':'gu','蛊':'gu','鼓':'gu','谷':'gu','故':'gu','顾':'gu','固':'gu','雇':'gu','瓜':'gua','刮':'gua','寡':'gua','挂':'gua','卦':'gua','乖':'guai','拐':'guai','怪':'guai','关':'guan','官':'guan','冠':'guan','观':'guan','管':'guan','馆':'guan','罐':'guan','惯':'guan','贯':'guan','灌':'guan','光':'guang','广':'guang','逛':'guang','归':'gui','龟':'gui','规':'gui','圭':'gui','硅':'gui','瑰':'gui','闺':'gui','轨':'gui','鬼':'gui','桂':'gui','柜':'gui','贵':'gui','跪':'gui','桂':'gui','滚':'gun','棍':'gun','锅':'guo','郭':'guo','国':'guo','果':'guo','裹':'guo','过':'guo',
  // H
  '哈':'ha','骸':'hai','孩':'hai','海':'hai','害':'hai','亥':'hai','骇':'hai','氦':'hai','还':'hai','含':'han','寒':'han','韩':'han','涵':'han','旱':'han','罕':'han','喊':'han','汉':'han','汗':'han','旱':'han','瀚':'han','夯':'hang','行':'hang','杭':'hang','航':'hang','壕':'hao','豪':'hao','毫':'hao','郝':'hao','好':'hao','号':'hao','浩':'hao','昊':'hao','皓':'hao','耗':'hao','呵':'he','喝':'he','禾':'he','和':'he','何':'he','合':'he','河':'he','核':'he','荷':'he','涸':'he','赫':'he','褐':'he','鹤':'he','贺':'he','黑':'hei','痕':'hen','很':'hen','狠':'hen','恨':'hen','恨':'hen','横':'heng','衡':'heng','恒':'heng','衡':'heng','轰':'hong','哄':'hong','红':'hong','虹':'hong','洪':'hong','宏':'hong','鸿':'hong','泓':'hong','弘':'hong','红':'hong','侯':'hou','喉':'hou','猴':'hou','吼':'hou','后':'hou','厚':'hou','候':'hou','后':'hou','乎':'hu','呼':'hu','忽':'hu','狐':'hu','胡':'hu','壶':'hu','湖':'hu','糊':'hu','蝴':'hu','虎':'hu','唬':'hu','护':'hu','互':'hu','户':'hu','花':'hua','华':'hua','哗':'hua','骅':'hua','化':'hua','划':'hua','画':'hua','话':'hua','桦':'hua','怀':'huai','淮':'huai','槐':'huai','坏':'huai','欢':'huan','还':'huan','环':'huan','桓':'huan','缓':'huan','换':'huan','唤':'huan','患':'huan','焕':'huan','幻':'huan','荒':'huang','慌':'huang','皇':'huang','黄':'huang','凰':'huang','惶':'huang','煌':'huang','晃':'huang','恍':'huang','谎':'huang','灰':'hui','挥':'hui','恢':'hui','挥':'hui','辉':'hui','徽':'hui','回':'hui','茴':'hui','悔':'hui','毁':'hui','汇':'hui','会':'hui','讳':'hui','诲':'hui','绘':'hui','贿':'hui','慧':'hui','秽':'hui','惠':'hui','昏':'hun','婚':'hun','魂':'hun','浑':'hun','混':'hun','活':'huo','火':'huo','伙':'huo','或':'huo','货':'huo','获':'huo','祸':'huo','惑':'huo',
  // J
  '几':'ji','机':'ji','饥':'ji','机':'ji','肌':'ji','鸡':'ji','迹':'ji','积':'ji','基':'ji','绩':'ji','激':'ji','缉':'ji','吉':'ji','即':'ji','急':'ji','疾':'ji','棘':'ji','集':'ji','籍':'ji','及':'ji','极':'ji','急':'ji','疾':'ji','棘':'ji','集':'ji','籍':'ji','几':'ji','挤':'ji','脊':'ji','己':'ji','纪':'ji','记':'ji','技':'ji','际':'ji','济':'ji','忌':'ji','继':'ji','寄':'ji','祭':'ji','寂':'ji','寄':'ji','加':'jia','佳':'jia','嘉':'jia','家':'jia','枷':'jia','夹':'jia','荚':'jia','颊':'jia','甲':'jia','假':'jia','价':'jia','架':'jia','驾':'jia','嫁':'jia','尖':'jian','坚':'jian','间':'jian','肩':'jian','艰':'jian','奸':'jian','监':'jian','兼':'jian','笺':'jian','渐':'jian','剪':'jian','捡':'jian','简':'jian','减':'jian','检':'jian','碱':'jian','简':'jian','见':'jian','件':'jian','建':'jian','剑':'jian','健':'jian','舰':'jian','渐':'jian','荐':'jian','践':'jian','鉴':'jian','键':'jian','江':'jiang','姜':'jiang','将':'jiang','浆':'jiang','僵':'jiang','缰':'jiang','讲':'jiang','奖':'jiang','蒋':'jiang','匠':'jiang','降':'jiang','酱':'jiang','交':'jiao','浇':'jiao','骄':'jiao','娇':'jiao','胶':'jiao','椒':'jiao','焦':'jiao','蕉':'jiao','礁':'jiao','角':'jiao','脚':'jiao','狡':'jiao','绞':'jiao','铰':'jiao','搅':'jiao','剿':'jiao','叫':'jiao','轿':'jiao','较':'jiao','教':'jiao','酵':'jiao','窖':'jiao','接':'jie','阶':'jie','皆':'jie','结':'jie','街':'jie','揭':'jie','截':'jie','节':'jie','杰':'jie','捷':'jie','睫':'jie','竭':'jie','洁':'jie','结':'jie','姐':'jie','解':'jie','介':'jie','戒':'jie','界':'jie','借':'jie','届':'jie','金':'jin','巾':'jin','斤':'jin','今':'jin','津':'jin','筋':'jin','襟':'jin','紧':'jin','锦':'jin','谨':'jin','进':'jin','近':'jin','晋':'jin','浸':'jin','尽':'jin','劲':'jin','晋':'jin','禁':'jin','靳':'jin','京':'jing','经':'jing','茎':'jing','惊':'jing','鲸':'jing','睛':'jing','精':'jing','菁':'jing','兢':'jing','景':'jing','井':'jing','阱':'jing','颈':'jing','景':'jing','警':'jing','竟':'jing','竞':'jing','敬':'jing','静':'jing','境':'jing','镜':'jing','径':'jing','靖':'jing','净':'jing','炯':'jiong','迥':'jiong','窘':'jiong','纠':'jiu','究':'jiu','鸠':'jiu','九':'jiu','久':'jiu','玖':'jiu','灸':'jiu','酒':'jiu','旧':'jiu','臼':'jiu','救':'jiu','就':'jiu','舅':'jiu','拘':'ju','居':'ju','驹':'ju','狙':'ju','疽':'ju','居':'ju','掬':'ju','局':'ju','菊':'ju','橘':'ju','局':'ju','咀':'ju','举':'ju','矩':'ju','沮':'ju','聚':'ju','拒':'ju','据':'ju','巨':'ju','具':'ju','剧':'ju','距':'ju','惧':'ju','据':'ju','卷':'juan','捐':'juan','涓':'juan','鹃':'juan','镌':'juan','卷':'juan','倦':'juan','眷':'juan','圈':'juan','卷':'juan','撅':'jue','决':'jue','诀':'jue','觉':'jue','角':'jue','脚':'jue','嚼':'jue','爵':'jue','倔':'jue','掘':'jue','崛':'jue','厥':'jue','蕨':'jue','獗':'jue','军':'jun','均':'jun','君':'jun','钧':'jun','菌':'jun','筠':'jun','俊':'jun','峻':'jun','浚':'jun','骏':'jun','竣':'jun',
  // K
  '咖':'ka','卡':'ka','喀':'ka','卡':'ka','开':'kai','凯':'kai','慨':'kai','楷':'kai','刊':'kan','堪':'kan','看':'kan','勘':'kan','坎':'kan','砍':'kan','侃':'kan','瞰':'kan','康':'kang','慷':'kang','糠':'kang','扛':'kang','抗':'kang','亢':'kang','考':'kao','烤':'kao','拷':'kao','烤':'kao','靠':'kao','苛':'ke','柯':'ke','棵':'ke','科':'ke','壳':'ke','颗':'ke','磕':'ke','颗':'ke','壳':'ke','咳':'ke','可':'ke','渴':'ke','克':'ke','刻':'ke','客':'ke','课':'ke','肯':'ken','啃':'ken','垦':'ken','恳':'ken','坑':'keng','吭':'keng','铿':'keng','空':'kong','孔':'kong','恐':'kong','控':'kong','抠':'kou','口':'kou','寇':'kou','扣':'kou','寇':'kou','蔻':'kou','枯':'ku','哭':'ku','窟':'ku','苦':'ku','酷':'ku','库':'ku','裤':'ku','夸':'kua','跨':'kua','垮':'kua','挎':'kua','跨':'kua','快':'kuai','块':'kuai','筷':'kuai','宽':'kuan','款':'kuan','匡':'kuang','筐':'kuang','狂':'kuang','框':'kuang','矿':'kuang','旷':'kuang','况':'kuang','亏':'kui','盔':'kui','窥':'kui','葵':'kui','奎':'kui','魁':'kui','馈':'kui','愧':'kui','溃':'kui','坤':'kun','昆':'kun','捆':'kun','困':'kun','廓':'kuo','扩':'kuo','括':'kuo',
  // L
  '垃':'la','拉':'la','啦':'la','拉':'la','喇':'la','腊':'la','辣':'la','辣':'la','来':'lai','莱':'la','赖':'lai','癞':'lai','兰':'lan','蓝':'lan','岚':'lan','栏':'lan','澜':'lan','婪':'lan','览':'lan','懒':'lan','烂':'lan','滥':'lan','琅':'lang','狼':'lang','郎':'lang','廊':'lang','朗':'lang','浪':'lang','捞':'lao','劳':'lao','牢':'lao','老':'lao','姥':'lao','烙':'lao','涝':'lao','乐':'le','雷':'lei','雷':'lei','垒':'lei','泪':'lei','类':'lei','累':'lei','肋':'lei','擂':'lei','磊':'lei','蕾':'lei','磊':'lei','泪':'lei','类':'lei','棱':'leng','冷':'leng','愣':'leng','厘':'li','狸':'li','离':'li','梨':'li','犁':'li','黎':'li','篱':'li','漓':'li','理':'li','里':'li','鲤':'li','礼':'li','李':'li','里':'li','理':'li','鲤':'li','力':'li','历':'li','立':'li','粒':'li','丽':'li','利':'li','励':'li','例':'li','俐':'li','莉':'li','荔':'li','吏':'li','栗':'li','砾':'li','痢':'li','连':'lian','帘':'lian','莲':'lian','连':'lian','联':'lian','廉':'lian','怜':'lian','涟':'lian','莲':'lian','脸':'lian','敛':'lian','链':'lian','练':'lian','恋':'lian','良':'liang','凉':'liang','梁':'liang','粱':'liang','粮':'liang','梁':'liang','两':'liang','亮':'liang','谅':'liang','辆':'liang','量':'liang','聊':'liao','辽':'liao','疗':'liao','聊':'liao','僚':'liao','寥':'liao','僚':'liao','廖':'liao','料':'liao','撂':'liao','列':'lie','烈':'lie','裂':'lie','劣':'lie','猎':'lie','裂':'lie','林':'lin','邻':'lin','林':'lin','临':'lin','淋':'lin','琳':'lin','霖':'lin','鳞':'lin','凛':'lin','凛':'lin','吝':'lin','赁':'lin','淋':'lin','灵':'ling','玲':'ling','凌':'ling','铃':'ling','龄':'ling','凌':'ling','陵':'ling','零':'ling','羚':'ling','零':'ling','领':'ling','岭':'ling','令':'ling','另':'ling','溜':'liu','刘':'liu','留':'liu','流':'liu','琉':'liu','硫':'liu','留':'liu','六':'liu','龙':'long','隆':'long','笼':'long','聋':'long','珑':'long','隆':'long','笼':'long','楼':'lou','楼':'lou','卢':'lu','芦':'lu','庐':'lu','炉':'lu','掳':'lu','卤':'lu','虏':'lu','鲁':'lu','橹':'lu','录':'lu','禄':'lu','陆':'lu','路':'lu','录':'lu','鹿':'lu','禄':'lu','碌':'lu','路':'lu','露':'lu','赂':'lu','鹭':'lu','驴':'lv','闾':'lv','旅':'lv','屡':'lv','缕':'lv','履':'lv','律':'lv','虑':'lv','率':'lv','绿':'lv','峦':'luan','孪':'luan','滦':'luan','卵':'luan','乱':'luan','掠':'lve','略':'lve','抡':'lun','轮':'lun','伦':'lun','沦':'lun','论':'lun','轮':'lun','萝':'luo','罗':'luo','逻':'luo','螺':'luo','罗':'luo','络':'luo','落':'luo','骆':'luo','洛':'luo','络':'luo','骆':'luo',
  // M
  '妈':'ma','麻':'ma','马':'ma','码':'ma','玛':'ma','骂':'ma','嘛':'ma','吗':'ma','埋':'mai','买':'mai','麦':'mai','卖':'mai','迈':'mai','麦':'mai','埋':'mai','脉':'mai','蛮':'man','满':'man','蔓':'man','曼':'man','慢':'man','漫':'man','忙':'mang','芒':'mang','茫':'mang','盲':'mang','茫':'mang','猫':'mao','毛':'mao','矛':'mao','茅':'mao','锚':'mao','毛':'mao','卯':'mao','茂':'mao','冒':'mao','帽':'mao','贸':'mao','貌':'mao','么':'me','没':'mei','眉':'mei','梅':'mei','玫':'mei','枚':'mei','玫':'mei','梅':'mei','媒':'mei','煤':'mei','霉':'mei','每':'mei','美':'mei','镁':'mei','妹':'mei','媚':'mei','魅':'mei','门':'men','闷':'men','们':'men','萌':'meng','蒙':'meng','盟':'meng','蒙':'meng','猛':'meng','孟':'meng','梦':'meng','迷':'mi','眯':'mi','谜':'mi','弥':'mi','糜':'mi','迷':'mi','靡':'mi','米':'mi','觅':'mi','泌':'mi','密':'mi','蜜':'mi','密':'mi','棉':'mian','眠':'mian','绵':'mian','棉':'mian','免':'mian','勉':'mian','冕':'mian','缅':'mian','面':'mian','苗':'miao','描':'miao','瞄':'miao','苗':'miao','秒':'miao','渺':'miao','庙':'miao','妙':'miao','庙':'miao','灭':'mie','蔑':'mie','民':'min','闽':'min','岷':'min','敏':'min','悯':'min','敏':'min','明':'ming','名':'ming','鸣':'ming','茗':'ming','冥':'ming','铭':'ming','明':'ming','冥':'ming','命':'ming','摸':'mo','摸':'mo','摩':'mo','磨':'mo','魔':'mo','磨':'mo','魔':'mo','抹':'mo','末':'mo','莫':'mo','墨':'mo','默':'mo','莫':'mo','漠':'mo','寞':'mo','陌':'mo','谋':'mou','某':'mou','眸':'mou','谋':'mou','母':'mu','牡':'mu','亩':'mu','拇':'mu','木':'mu','目':'mu','牧':'mu','墓':'mu','幕':'mu','慕':'mu','木':'mu','穆':'mu',
  // N
  '拿':'na','哪':'na','那':'na','纳':'na','娜':'na','乃':'nai','奶':'nai','耐':'nai','奈':'nai','男':'nan','南':'nan','男':'nan','难':'nan','囊':'nang','囊':'nang','挠':'nao','恼':'nao','脑':'nao','闹':'nao','淖':'nao','呢':'ne','馁':'nei','内':'nei','嫩':'nen','能':'neng','妮':'ni','妮':'ni','霓':'ni','泥':'ni','你':'ni','拟':'ni','逆':'ni','年':'nian','念':'nian','念':'nian','捻':'nian','鸟':'niao','尿':'niao','捏':'nie','涅':'nie','镍':'nie','涅':'nie','您':'nin','您':'nin','宁':'ning','凝':'ning','柠':'ning','宁':'ning','牛':'niu','扭':'niu','纽':'niu','浓':'nong','农':'nong','浓':'nong','弄':'nong','奴':'nu','努':'nu','努':'nu','怒':'nu','女':'nv','暖':'nuan','虐':'nue','疟':'nue','诺':'nuo','糯':'nuo',
  // O
  '哦':'o','欧':'ou','欧':'ou','藕':'ou','偶':'ou','呕':'ou','鸥':'ou',
  // P
  '趴':'pa','怕':'pa','拍':'pai','排':'pai','派':'pai','攀':'pan','潘':'pan','盘':'pan','判':'pan','盼':'pan','畔':'pan','盘':'pan','胖':'pang','旁':'pang','庞':'pang','旁':'pang','胖':'pang','抛':'pao','泡':'pao','炮':'pao','袍':'pao','跑':'pao','泡':'pao','陪':'pei','培':'pei','赔':'pei','佩':'pei','配':'pei','佩':'pei','喷':'pen','盆':'pen','朋':'peng','彭':'peng','蓬':'peng','棚':'peng','蓬':'peng','鹏':'peng','捧':'peng','碰':'peng','批':'pi','坯':'pi','披':'pi','劈':'pi','皮':'pi','疲':'pi','皮':'pi','匹':'pi','痞':'pi','僻':'pi','屁':'pi','偏':'pian','篇':'pian','片':'pian','骗':'pian','票':'piao','漂':'piao','飘':'piao','漂':'piao','瓢':'piao','票':'piao','拼':'pin','贫':'pin','频':'pin','品':'pin','聘':'pin','乒':'ping','平':'ping','评':'ping','凭':'ping','瓶':'ping','苹':'ping','屏':'ping','萍':'ping','平':'ping','评':'ping','坡':'po','泼':'po','婆':'po','破':'po','魄':'po','剖':'pou','扑':'pu','铺':'pu','仆':'pu','葡':'pu','蒲':'pu','朴':'pu','浦':'pu','普':'pu','谱':'pu','浦':'pu',
  // Q
  '七':'qi','妻':'qi','栖':'qi','凄':'qi','妻':'qi','戚':'qi','欺':'qi','漆':'qi','七':'qi','其':'qi','奇':'qi','歧':'qi','棋':'qi','旗':'qi','歧':'qi','旗':'qi','祈':'qi','骑':'qi','崎':'qi','琦':'qi','琪':'qi','麒':'qi','企':'qi','启':'qi','起':'qi','气':'qi','汽':'qi','弃':'qi','器':'qi','契':'qi','泣':'qi','恰':'qia','洽':'qia','千':'qian','迁':'qian','签':'qian','牵':'qian','铅':'qian','谦':'qian','前':'qian','钱':'qian','浅':'qian','遣':'qian','欠':'qian','歉':'qian','枪':'qiang','枪':'qiang','强':'qiang','墙':'qiang','抢':'qiang','强':'qiang','悄':'qiao','敲':'qiao','桥':'qiao','乔':'qiao','侨':'qiao','荞':'qiao','桥':'qiao','巧':'qiao','切':'qie','且':'qie','窃':'qie','茄':'qie','且':'qie','切':'qie','怯':'qie','窃':'qie','亲':'qin','侵':'qin','钦':'qin','琴':'qin','秦':'qin','寝':'qin','青':'qing','轻':'qing','氢':'qing','清':'qing','蜻':'qing','情':'qing','晴':'qing','顷':'qing','请':'qing','情':'qing','擎':'qing','顷':'qing','请':'qing','庆':'qing','亲':'qing','穷':'qiong','穹':'qiong','琼':'qiong','秋':'qiu','丘':'qiu','邱':'qiu','求':'qiu','球':'qiu','求':'qiu','囚':'qiu','酋':'qiu','区':'qu','曲':'qu','驱':'qu','躯':'qu','趋':'qu','曲':'qu','取':'qu','娶':'qu','去':'qu','趣':'qu','全':'quan','权':'quan','泉':'quan','拳':'quan','全':'quan','犬':'quan','劝':'quan','缺':'que','却':'que','确':'que','鹊':'que','雀':'que','确':'que','裙':'qun','群':'qun',
  // R
  '然':'ran','燃':'ran','染':'ran','嚷':'rang','壤':'rang','让':'rang','饶':'rao','扰':'rao','绕':'rao','惹':'re','热':'re','人':'ren','仁':'ren','人':'ren','忍':'ren','认':'ren','任':'ren','韧':'ren','刃':'ren','认':'ren','扔':'reng','仍':'reng','日':'ri','荣':'rong','容':'rong','绒':'rong','融':'rong','熔':'rong','溶':'rong','容':'rong','柔':'rou','肉':'rou','如':'ru','茹':'ru','儒':'ru','乳':'ru','辱':'ru','入':'ru','软':'ruan','瑞':'rui','蕊':'rui','锐':'rui','瑞':'rui','润':'run','闰':'run','若':'ruo','弱':'ruo',
  // S
  '撒':'sa','洒':'sa','萨':'sa','塞':'sai','腮':'sai','赛':'sai','三':'san','叁':'san','伞':'san','散':'san','桑':'sang','丧':'sang','桑':'sang','嗓':'sang','丧':'sang','骚':'sao','扫':'sao','嫂':'sao','色':'se','涩':'se','瑟':'se','森':'sen','僧':'seng','杀':'sha','沙':'sha','纱':'sha','傻':'sha','啥':'sha','煞':'sha','晒':'shai','山':'shan','删':'shan','杉':'shan','衫':'shan','珊':'shan','闪':'shan','善':'shan','膳':'shan','擅':'shan','伤':'shang','商':'shang','伤':'shang','裳':'shang','晌':'shang','上':'shang','尚':'shang','裳':'shang','梢':'shao','捎':'shao','烧':'shao','稍':'shao','勺':'shao','少':'shao','邵':'shao','绍':'shao','哨':'shao','舌':'she','蛇':'she','舍':'she','射':'she','涉':'she','社':'she','设':'she','射':'she','摄':'she','申':'shen','伸':'shen','身':'shen','深':'shen','申':'shen','绅':'shen','神':'shen','沈':'shen','审':'shen','婶':'shen','甚':'shen','慎':'shen','渗':'shen','肾':'shen','升':'sheng','生':'sheng','声':'sheng','牲':'sheng','甥':'sheng','绳':'sheng','省':'sheng','圣':'sheng','胜':'sheng','圣':'sheng','盛':'sheng','剩':'sheng','尸':'shi','失':'shi','师':'shi','诗':'shi','施':'shi','湿':'shi','诗':'shi','狮':'shi','施':'shi','湿':'shi','十':'shi','石':'shi','时':'shi','识':'shi','实':'shi','拾':'shi','食':'shi','蚀':'shi','史':'shi','使':'shi','始':'shi','屎':'shi','史':'shi','使':'shi','始':'shi','士':'shi','氏':'shi','世':'shi','市':'shi','示':'shi','式':'shi','事':'shi','侍':'shi','势':'shi','是':'shi','适':'shi','室':'shi','视':'shi','试':'shi','饰':'shi','室':'shi','逝':'shi','释':'shi','誓':'shi','收':'shou','手':'shou','守':'shou','首':'shou','寿':'shou','受':'shou','售':'shou','授':'shou','兽':'shou','瘦':'shou','书':'shu','叔':'shu','殊':'shu','梳':'shu','淑':'shu','舒':'shu','疏':'shu','输':'shu','蔬':'shu','熟':'shu','熟':'shu','暑':'shu','黍':'shu','署':'shu','蜀':'shu','鼠':'shu','术':'shu','树':'shu','束':'shu','竖':'shu','墅':'shu','数':'shu','术':'shu','刷':'shua','耍':'shua','衰':'shuai','摔':'shuai','甩':'shuai','帅':'shuai','栓':'shuan','拴':'shuan','双':'shuang','爽':'shuang','霜':'shuang','谁':'shui','水':'shui','睡':'shui','顺':'shun','舜':'shun','瞬':'shun','说':'shuo','朔':'shuo','硕':'shuo','朔':'shuo','斯':'si','司':'si','私':'si','思':'si','丝':'si','斯':'si','撕':'si','死':'si','四':'si','寺':'si','似':'si','饲':'si','驷':'si','肆':'si','松':'song','宋':'song','颂':'song','讼':'song','送':'song','诵':'song','颂':'song','搜':'sou','艘':'sou','苏':'su','酥':'su','俗':'su','速':'su','素':'su','诉':'su','宿':'su','肃':'su','粟':'su','诉':'su','素':'su','速':'su','酸':'suan','算':'suan','蒜':'suan','虽':'sui','隋':'sui','随':'sui','绥':'sui','髓':'sui','岁':'sui','碎':'sui','穗':'sui','遂':'sui','隧':'sui','孙':'sun','损':'sun','笋':'sun','损':'sun','缩':'suo','所':'suo','锁':'suo','索':'suo',
  // T
  '他':'ta','她':'ta','它':'ta','塌':'ta','塔':'ta','踏':'ta','台':'tai','抬':'tai','太':'tai','态':'tai','泰':'tai','汰':'tai','谈':'tan','潭':'tan','谭':'tan','坛':'tan','檀':'tan','痰':'tan','毯':'tan','叹':'tan','探':'tan','汤':'tang','堂':'tang','棠':'tang','塘':'tang','糖':'tang','膛':'tang','唐':'tang','堂':'tang','淌':'tang','躺':'tang','烫':'tang','趟':'tang','涛':'tao','掏':'tao','逃':'tao','桃':'tao','淘':'tao','陶':'tao','淘':'tao','套':'tao','特':'te','疼':'teng','腾':'teng','藤':'teng','腾':'teng','剔':'ti','踢':'ti','锑':'ti','梯':'ti','提':'ti','题':'ti','啼':'ti','蹄':'ti','啼':'ti','体':'ti','替':'ti','天':'tian','添':'tian','田':'tian','甜':'tian','填':'tian','恬':'tian','田':'tian','舔':'tian','挑':'tiao','条':'tiao','调':'tiao','跳':'tiao','眺':'tiao','跳':'tiao','贴':'tie','铁':'tie','帖':'tie','铁':'tie','厅':'ting','听':'ting','烃':'ting','停':'ting','亭':'ting','庭':'ting','婷':'ting','廷':'ting','挺':'ting','艇':'ting','通':'tong','同':'tong','彤':'tong','童':'tong','铜':'tong','桐':'tong','童':'tong','统':'tong','桶':'tong','捅':'tong','痛':'tong','偷':'tou','头':'tou','投':'tou','透':'tou','凸':'tu','突':'tu','秃':'tu','图':'tu','途':'tu','涂':'tu','土':'tu','吐':'tu','兔':'tu','团':'tuan','湍':'tuan','推':'tui','颓':'tui','腿':'tui','退':'tui','吞':'tun','屯':'tun','拖':'tuo','托':'tuo','脱':'tuo','拖':'tuo','托':'tuo','妥':'tuo','拓':'tuo',
  // W
  '挖':'wa','哇':'wa','蛙':'wa','瓦':'wa','袜':'wa','歪':'wai','外':'wai','湾':'wan','弯':'wan','丸':'wan','完':'wan','玩':'wan','顽':'wan','碗':'wan','宛':'wan','婉':'wan','惋':'wan','晚':'wan','万':'wan','腕':'wan','汪':'wang','王':'wang','亡':'wang','枉':'wang','网':'wang','往':'wang','旺':'wang','望':'wang','妄':'wang','威':'wei','危':'wei','偎':'wei','微':'wei','薇':'wei','巍':'wei','为':'wei','围':'wei','唯':'wei','惟':'wei','维':'wei','伟':'wei','伪':'wei','尾':'wei','纬':'wei','伟':'wei','委':'wei','炜':'wei','卫':'wei','味':'wei','畏':'wei','胃':'wei','尉':'wei','谓':'wei','魏':'wei','位':'wei','渭':'wei','谓':'wei','温':'wen','文':'wen','纹':'wen','闻':'wen','蚊':'wen','雯':'wen','稳':'wen','问':'wen','翁':'weng','嗡':'weng','瓮':'weng','我':'wo','握':'wo','蜗':'wo','窝':'wo','卧':'wo','沃':'wo','握':'wo','乌':'wu','巫':'wu','呜':'wu','污':'wu','屋':'wu','无':'wu','吴':'wu','五':'wu','吾':'wu','伍':'wu','武':'wu','午':'wu','侮':'wu','舞':'wu','务':'wu','物':'wu','误':'wu','悟':'wu','雾':'wu','晤':'wu',
  // X
  '西':'xi','吸':'xi','希':'xi','息':'xi','悉':'xi','惜':'xi','晰':'xi','溪':'xi','锡':'xi','熙':'xi','稀':'xi','曦':'xi','习':'xi','袭':'xi','席':'xi','喜':'xi','洗':'xi','玺':'xi','铣':'xi','喜':'xi','戏':'xi','系':'xi','细':'xi','隙':'xi','瞎':'xia','虾':'xia','侠':'xia','狭':'xia','霞':'xia','下':'xia','夏':'xia','吓':'xia','仙':'xian','先':'xian','仙':'xian','掀':'xian','鲜':'xian','贤':'xian','闲':'xian','弦':'xian','衔':'xian','嫌':'xian','显':'xian','险':'xian','藓':'xian','现':'xian','限':'xian','线':'xian','宪':'xian','陷':'xian','献':'xian','腺':'xian','馅':'xian','羡':'xian','乡':'xiang','相':'xiang','香':'xiang','箱':'xiang','厢':'xiang','湘':'xiang','祥':'xiang','翔':'xiang','详':'xiang','想':'xiang','响':'xiang','享':'xiang','项':'xiang','巷':'xiang','橡':'xiang','向':'xiang','像':'xiang','象':'xiang','消':'xiao','萧':'xiao','硝':'xiao','销':'xiao','潇':'xiao','霄':'xiao','小':'xiao','晓':'xiao','孝':'xiao','肖':'xiao','效':'xiao','校':'xiao','啸':'xiao','笑':'xiao','些':'xie','歇':'xie','蝎':'xie','斜':'xie','携':'xie','协':'xie','挟':'xie','邪':'xie','鞋':'xie','写':'xie','血':'xie','泻':'xie','谢':'xie','懈':'xie','械':'xie','蟹':'xie','新':'xin','心':'xin','欣':'xin','辛':'xin','锌':'xin','薪':'xin','馨':'xin','鑫':'xin','信':'xin','芯':'xin','衅':'xin','星':'xing','兴':'xing','猩':'xing','刑':'xing','型':'xing','形':'xing','邢':'xing','醒':'xing','性':'xing','姓':'xing','幸':'xing','杏':'xing','性':'xing','凶':'xiong','兄':'xiong','胸':'xiong','雄':'xiong','熊':'xiong','休':'xiu','修':'xiu','秀':'xiu','绣':'xiu','袖':'xiu','秀':'xiu','锈':'xiu','袖':'xiu','墟':'xu','虚':'xu','嘘':'xu','需':'xu','徐':'xu','许':'xu','须':'xu','需':'xu','徐':'xu','许':'xu','序':'xu','叙':'xu','绪':'xu','续':'xu','蓄':'xu','宣':'xuan','轩':'xuan','悬':'xuan','玄':'xuan','旋':'xuan','选':'xuan','癣':'xuan','选':'xuan','绚':'xuan','穴':'xue','学':'xue','薛':'xue','血':'xue','穴':'xue','学':'xue','雪':'xue','靴':'xue','谑':'xue','勋':'xun','寻':'xun','巡':'xun','循':'xun','询':'xun','荀':'xun','训':'xun','讯':'xun','迅':'xun','逊':'xun',
  // Y
  '压':'ya','呀':'ya','鸦':'ya','牙':'ya','芽':'ya','雅':'ya','亚':'ya','讶':'ya','烟':'yan','咽':'yan','严':'yan','言':'yan','岩':'yan','炎':'yan','沿':'yan','研':'yan','盐':'yan','阎':'yan','颜':'yan','衍':'yan','眼':'yan','演':'yan','厌':'yan','宴':'yan','艳':'yan','验':'yan','彦':'yan','焰':'yan','燕':'yan','央':'yang','秧':'yang','鸯':'yang','杨':'yang','扬':'yang','阳':'yang','洋':'yang','仰':'yang','养':'yang','氧':'yang','仰':'yang','样':'yang','漾':'yang','妖':'yao','邀':'yao','腰':'yao','邀':'yao','瑶':'yao','摇':'yao','遥':'yao','瑶':'yao','咬':'yao','舀':'yao','药':'yao','要':'yao','耀':'yao','钥':'yao','爷':'ye','耶':'ye','椰':'ye','也':'ye','野':'ye','业':'ye','叶':'ye','页':'ye','夜':'ye','液':'ye','叶':'ye','一':'yi','衣':'yi','依':'yi','医':'yi','伊':'yi','宜':'yi','壹':'yi','颐':'yi','以':'yi','已':'yi','以':'yi','蚁':'yi','倚':'yi','椅':'yi','义':'yi','亿':'yi','艺':'yi','忆':'yi','议':'yi','亦':'yi','异':'yi','易':'yi','翼':'yi','译':'yi','驿':'yi','疫':'yi','益':'yi','溢':'yi','谊':'yi','肄':'yi','毅':'yi','裔':'yi','意':'yi','熠':'yi','因':'yin','阴':'yin','姻':'yin','音':'yin','殷':'yin','茵':'yin','银':'yin','淫':'yin','寅':'yin','引':'yin','饮':'yin','隐':'yin','印':'yin','荫':'yin','胤':'yin','英':'ying','应':'ying','莺':'ying','婴':'ying','赢':'ying','鹰':'ying','迎':'ying','盈':'ying','营':'ying','莹':'ying','荧':'ying','蝇':'ying','影':'ying','颖':'ying','影':'ying','硬':'ying','映':'ying','拥':'yong','庸':'yong','永':'yong','泳':'yong','勇':'yong','涌':'yong','咏':'yong','踊':'yong','用':'yong','优':'you','悠':'you','忧':'you','悠':'you','尤':'you','由':'you','犹':'you','油':'you','游':'you','友':'you','有':'you','酉':'you','有':'you','佑':'you','右':'you','又':'you','幼':'you','于':'yu','迂':'yu','淤':'yu','于':'yu','余':'yu','予':'yu','娱':'yu','鱼':'yu','渔':'yu','愉':'yu','逾':'yu','瑜':'yu','愚':'yu','舆':'yu','鱼':'yu','与':'yu','宇':'yu','羽':'yu','雨':'yu','语':'yu','禹':'yu','屿':'yu','语':'yu','御':'yu','玉':'yu','域':'yu','育':'yu','浴':'yu','浴':'yu','预':'yu','寓':'yu','遇':'yu','誉':'yu','御':'yu','欲':'yu','裕':'yu','裕':'yu','豫':'yu','鸳':'yuan','渊':'yuan','冤':'yuan','元':'yuan','园':'yuan','原':'yuan','员':'yuan','圆':'yuan','援':'yuan','缘':'yuan','袁':'yuan','源':'yuan','远':'yuan','苑':'yuan','怨':'yuan','院':'yuan','愿':'yuan','月':'yue','曰':'yue','约':'yue','月':'yue','岳':'yue','阅':'yue','悦':'yue','跃':'yue','越':'yue','粤':'yue','云':'yun','匀':'yun','耘':'yun','云':'yun','芸':'yun','允':'yun','孕':'yun','运':'yun','韵':'yun','蕴':'yun',
  // Z
  '杂':'za','砸':'za','灾':'zai','栽':'zai','载':'zai','宰':'zai','载':'zai','再':'zai','在':'zai','咱':'zan','暂':'zan','赞':'zan','脏':'zang','葬':'zang','遭':'zao','糟':'zao','早':'zao','澡':'zao','枣':'zao','澡':'zao','皂':'zao','造':'zao','灶':'zao','燥':'zao','则':'ze','择':'ze','责':'ze','择':'ze','贼':'zei','怎':'zen','增':'zeng','曾':'zeng','赠':'zeng','扎':'zha','眨':'zha','炸':'zha','渣':'zha','札':'zha','轧':'zha','铡':'zha','闸':'zha','眨':'zha','榨':'zha','乍':'zha','炸':'zha','摘':'zhai','宅':'zhai','窄':'zhai','债':'zhai','寨':'zhai','沾':'zhan','毡':'zhan','粘':'zhan','詹':'zhan','展':'zhan','崭':'zhan','展':'zhan','占':'zhan','站':'zhan','战':'zhan','栈':'zhan','绽':'zhan','张':'zhang','章':'zhang','彰':'zhang','漳':'zhang','蟑':'zhang','杖':'zhang','丈':'zhang','帐':'zhang','账':'zhang','仗':'zhang','胀':'zhang','障':'zhang','招':'zhao','昭':'zhao','找':'zhao','沼':'zhao','照':'zhao','罩':'zhao','肇':'zhao','赵':'zhao','哲':'zhe','遮':'zhe','折':'zhe','哲':'zhe','辙':'zhe','者':'zhe','锗':'zhe','蔗':'zhe','这':'zhe','浙':'zhe','珍':'zhen','真':'zhen','斟':'zhen','甄':'zhen','臻':'zhen','诊':'zhen','枕':'zhen','阵':'zhen','振':'zhen','镇':'zhen','阵':'zhen','震':'zhen','正':'zheng','争':'zheng','征':'zheng','征':'zheng','挣':'zheng','睁':'zheng','蒸':'zheng','整':'zheng','拯':'zheng','正':'zheng','政':'zheng','症':'zheng','之':'zhi','之':'zhi','支':'zhi','枝':'zhi','知':'zhi','芝':'zhi','织':'zhi','肢':'zhi','脂':'zhi','职':'zhi','植':'zhi','殖':'zhi','执':'zhi','值':'zhi','侄':'zhi','旨':'zhi','址':'zhi','指':'zhi','纸':'zhi','制':'zhi','志':'zhi','挚':'zhi','致':'zhi','置':'zhi','智':'zhi','秩':'zhi','稚':'zhi','质':'zhi','炙':'zhi','中':'zhong','忠':'zhong','钟':'zhong','衷':'zhong','终':'zhong','种':'zhong','肿':'zhong','重':'zhong','仲':'zhong','众':'zhong','周':'zhou','州':'zhou','舟':'zhou','周':'zhou','粥':'zhou','轴':'zhou','肘':'zhou','帚':'zhou','咒':'zhou','宙':'zhou','皱':'zhou','昼':'zhou','骤':'zhou','珠':'zhu','朱':'zhu','株':'zhu','蛛':'zhu','珠':'zhu','诸':'zhu','猪':'zhu','竹':'zhu','烛':'zhu','逐':'zhu','竹':'zhu','主':'zhu','拄':'zhu','煮':'zhu','瞩':'zhu','嘱':'zhu','主':'zhu','注':'zhu','住':'zhu','驻':'zhu','祝':'zhu','著':'zhu','筑':'zhu','铸':'zhu','注':'zhu','贮':'zhu','抓':'zhua','爪':'zhua','拽':'zhuai','转':'zhuan','专':'zhuan','砖':'zhuan','转':'zhuan','赚':'zhuan','篆':'zhuan','桩':'zhuang','装':'zhuang','庄':'zhuang','妆':'zhuang','装':'zhuang','壮':'zhuang','状':'zhuang','撞':'zhuang','准':'zhun','桌':'zhuo','拙':'zhuo','捉':'zhuo','灼':'zhuo','卓':'zhuo','桌':'zhuo','酌':'zhuo','啄':'zhuo','浊':'zhuo','孜':'zi','咨':'zi','资':'zi','姿':'zi','滋':'zi','孜':'zi','紫':'zi','姊':'zi','梓':'zi','滓':'zi','子':'zi','梓':'zi','自':'zi','字':'zi','恣':'zi','宗':'zong','棕':'zong','踪':'zong','综':'zong','总':'zong','粽':'zong','总':'zong','纵':'zong','走':'zou','奏':'zou','邹':'zou','租':'zu','族':'zu','阻':'zu','组':'zu','祖':'zu','阻':'zu','钻':'zuan','嘴':'zui','醉':'zui','最':'zui','罪':'zui','尊':'zun','遵':'zun','昨':'zuo','左':'zuo','作':'zuo','坐':'zuo','座':'zuo','做':'zuo','作':'zuo',
};

// 根据中文姓名生成邮箱（规则：名.姓@hxgroup.com）
function generateEmailFromName(name) {
  if (!name || name.trim() === '') return '';
  
  // 检查是否已在列表中
  const allEmails = getAllEmails();
  if (allEmails[name]) return allEmails[name];
  
  // 处理复姓（2个字的姓氏）
  let surname = '';
  let givenName = '';
  
  if (name.length >= 2) {
    // 先尝试匹配复姓
    const doubleSurname = name.substring(0, 2);
    if (SURNAME_PINYIN[doubleSurname]) {
      surname = SURNAME_PINYIN[doubleSurname];
      givenName = name.substring(2);
    } else {
      // 单姓
      const firstChar = name.substring(0, 1);
      surname = SURNAME_PINYIN[firstChar] || CHAR_PINYIN[firstChar] || firstChar;
      givenName = name.substring(1);
    }
  }
  
  // 如果姓氏无法识别，就用全名的某种简化形式
  if (!surname) {
    return name.toLowerCase().replace(/\s+/g, '') + '@hxgroup.com';
  }
  
  // 从已有邮箱列表中提取字→拼音映射（最准确）
  const charMap = {};
  for (const existingName in allEmails) {
    const email = allEmails[existingName];
    const emailParts = email.split('@')[0].split('.');
    if (emailParts.length >= 2) {
      const givenPinyinPart = emailParts[0];
      const surnamePinyinPart = emailParts[1];
      // 姓氏映射
      const surnameChar = existingName.substring(0, SURNAME_PINYIN[existingName.substring(0, 2)] ? 2 : 1);
      charMap[surnameChar] = surnamePinyinPart;
      // 名字映射（按字符位置尝试匹配）
      const givenChars = existingName.substring(surnameChar.length);
      if (givenChars.length > 0 && givenPinyinPart.length >= givenChars.length) {
        // 简单的按字符分割（基于常见2字词拼音长度）
        let remaining = givenPinyinPart;
        for (let i = 0; i < givenChars.length; i++) {
          const ch = givenChars[i];
          // 如果只剩最后一个字，剩下的全归它
          if (i === givenChars.length - 1) {
            charMap[ch] = remaining;
            break;
          }
          // 尝试从常见拼音结尾字母分割
          const endings = ['a', 'e', 'i', 'o', 'u', 'n', 'g'];
          let found = false;
          for (let j = 2; j <= remaining.length; j++) {
            if (endings.includes(remaining.charAt(j - 1)) && !endings.includes(remaining.charAt(j))) {
              charMap[ch] = remaining.substring(0, j);
              remaining = remaining.substring(j);
              found = true;
              break;
            }
          }
          if (!found) {
            // 保守估计：每个字2-3个字母
            const est = Math.max(2, Math.floor(remaining.length / (givenChars.length - i)));
            charMap[ch] = remaining.substring(0, est);
            remaining = remaining.substring(est);
          }
        }
      }
    }
  }
  
  // 生成名字拼音：优先从已有映射，其次从CHAR_PINYIN库
  let givenPinyin = '';
  let hasUnknown = false;
  for (let i = 0; i < givenName.length; i++) {
    const ch = givenName[i];
    if (charMap[ch]) {
      givenPinyin += charMap[ch];
    } else if (CHAR_PINYIN[ch]) {
      givenPinyin += CHAR_PINYIN[ch];
    } else {
      hasUnknown = true;
      givenPinyin += '[' + ch + ']';
    }
  }
  
  if (hasUnknown) {
    return givenPinyin + '.' + surname + '@hxgroup.com';
  }
  
  return givenPinyin + '.' + surname + '@hxgroup.com';
}

// ==================== 新增：从Project导入 ====================
let importParsedData = [];  // 存储解析后的预览数据
let importFeatures = { text: true, mpp: false, pdf: false, ocr: false };  // 可用功能

async function openImportModal() {
  const depts = Object.keys(RAW_DATA.depts || {}).sort();
  
  // 先检测可用功能
  try {
    const resp = await fetch('/api/import/features');
    if (resp.ok) {
      const data = await resp.json();
      importFeatures = { ...importFeatures, ...data };
    }
  } catch (e) {
    // 功能检测失败，使用默认值
  }
  
  const modal = document.createElement('div');
  modal.className = 'modal-overlay';
  modal.innerHTML = `
    <div class="modal-box" style="max-width:900px">
      <button class="modal-close" onclick="this.closest('.modal-overlay').remove()">×</button>
      <h3>📥 从Project导入</h3>
      
      <!-- Tab 按钮 -->
      <div style="display:flex;gap:4px;margin-bottom:16px;border-bottom:2px solid #e5e7eb">
        <button class="import-tab-btn" data-tab="text" data-active="true" style="padding:8px 16px;border:none;background:none;cursor:pointer;border-bottom:3px solid #8b5cf6;color:#8b5cf6;font-weight:600">📋 文本粘贴</button>
        <button class="import-tab-btn" data-tab="mpp" style="padding:8px 16px;border:none;background:none;cursor:pointer;border-bottom:3px solid transparent;color:#6b7280;font-weight:600">📁 .mpp文件${importFeatures.mpp ? '' : ' 🔒'}</button>
        <button class="import-tab-btn" data-tab="pdf" style="padding:8px 16px;border:none;background:none;cursor:pointer;border-bottom:3px solid transparent;color:#6b7280;font-weight:600">📄 PDF文件${importFeatures.pdf ? '' : ' 🔒'}</button>
        <button class="import-tab-btn" data-tab="screenshot" style="padding:8px 16px;border:none;background:none;cursor:pointer;border-bottom:3px solid transparent;color:#6b7280;font-weight:600">🖼️ 截图${importFeatures.ocr ? '' : ' 🔒'}</button>
      </div>
      
      <!-- Tab: 文本粘贴 -->
      <div id="import-tab-text" class="import-tab-content">
        <div class="form-group">
          <label>粘贴Project表格数据（制表符或逗号分隔）</label>
          <textarea id="importTextArea" rows="8" placeholder="从Project中复制资源工作表，粘贴到这里...&#10;支持列：项目名称、资源名称、资源类型、开始时间、结束时间、工时等"></textarea>
        </div>
      </div>
      
      <!-- Tab: .mpp文件 -->
      <div id="import-tab-mpp" class="import-tab-content" style="display:none">
        <div class="form-group">
          <label>选择.mpp文件</label>
          <input type="file" id="importMppFile" accept=".mpp">
          ${importFeatures.mpp ? '' : '<div class="form-hint" style="color:#dc2626">⚠️ .mpp解析功能暂不可用，请先在服务器端安装相关依赖</div>'}
        </div>
      </div>
      
      <!-- Tab: PDF文件 -->
      <div id="import-tab-pdf" class="import-tab-content" style="display:none">
        <div class="form-group">
          <label>选择PDF文件（Microsoft Project导出的PDF）</label>
          <input type="file" id="importPdfFile" accept=".pdf">
          <div class="form-hint" style="color:#6b7280;font-size:12px;margin-top:4px">💡 从Project中选择"文件→导出→创建PDF/XPS"，确保左侧任务表格包含在PDF中</div>
          ${importFeatures.pdf ? '' : '<div class="form-hint" style="color:#dc2626">⚠️ PDF解析功能暂不可用，请先在服务器端安装pdfplumber依赖</div>'}
        </div>
      </div>
      
      <!-- Tab: 截图 -->
      <div id="import-tab-screenshot" class="import-tab-content" style="display:none">
        <div class="form-group">
          <label>选择截图文件</label>
          <input type="file" id="importScreenshotFile" accept="image/*">
          <div class="form-hint" style="color:#dc2626">⚠️ OCR功能暂不可用</div>
        </div>
      </div>
      
      <div class="form-actions" style="border:none;padding:0;margin:0 0 16px 0">
        <button class="btn" style="background:#8b5cf6;color:white" onclick="parseImportData()">🔍 解析预览</button>
      </div>
      
      <!-- 警告区域 -->
      <div id="importWarnings" style="display:none;background:#fef3c7;border:1px solid #f59e0b;color:#92400e;padding:10px 14px;border-radius:8px;margin-bottom:16px;font-size:13px"></div>
      
      <!-- 部门统一选择 -->
      <div id="importDeptRow" style="display:none;margin-bottom:12px">
        <div class="form-group">
          <label>部门（E列）- 统一应用到所有导入项</label>
          <select id="importDeptSelect">
            <option value="">请选择部门</option>
            ${depts.map(d => `<option value="${d}">${d}</option>`).join('')}
          </select>
        </div>
      </div>
      
      <!-- 预览表格 -->
      <div id="importPreviewArea" style="display:none">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
          <label style="font-size:13px;font-weight:600;color:#374151">解析结果预览（可编辑，共 <span id="importPreviewCount">0</span> 条）</label>
        </div>
        <div id="importPreviewTable" style="max-height:300px;overflow-y:auto;border:1px solid #e5e7eb;border-radius:8px"></div>
      </div>
      
      <div class="form-actions">
        <button class="btn btn-secondary" onclick="this.closest('.modal-overlay').remove()">取消</button>
        <button class="btn btn-success" onclick="submitImport()">✅ 确认导入</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  
  // Tab切换
  modal.querySelectorAll('.import-tab-btn').forEach(btn => {
    btn.addEventListener('click', function() {
      const tab = this.dataset.tab;
      // 检查功能是否可用
      if (tab === 'mpp' && !importFeatures.mpp) {
        alert('.mpp解析功能暂不可用');
        return;
      }
      if (tab === 'pdf' && !importFeatures.pdf) {
        alert('PDF解析功能暂不可用');
        return;
      }
      if (tab === 'screenshot' && !importFeatures.ocr) {
        alert('OCR功能暂不可用');
        return;
      }
      // 切换Tab样式
      modal.querySelectorAll('.import-tab-btn').forEach(b => {
        b.style.borderBottom = '3px solid transparent';
        b.style.color = '#6b7280';
        b.removeAttribute('data-active');
      });
      this.style.borderBottom = '3px solid #8b5cf6';
      this.style.color = '#8b5cf6';
      this.setAttribute('data-active', 'true');
      // 切换内容
      modal.querySelectorAll('.import-tab-content').forEach(c => {
        c.style.display = 'none';
      });
      document.getElementById('import-tab-' + tab).style.display = 'block';
    });
  });
}

function renderImportPreview() {
  const tableEl = document.getElementById('importPreviewTable');
  const countEl = document.getElementById('importPreviewCount');
  if (!tableEl) return;
  
  countEl.textContent = importParsedData.length;
  
  if (importParsedData.length === 0) {
    tableEl.innerHTML = '<div style="padding:20px;text-align:center;color:#9ca3af">暂无数据</div>';
    return;
  }
  
  let html = `
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <thead style="position:sticky;top:0;background:#f3f4f6;z-index:1">
        <tr>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">项目名称</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">资源名称</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">资源类型</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">开始时间</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">结束时间</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:left">日工时</th>
          <th style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:center">操作</th>
        </tr>
      </thead>
      <tbody>
  `;
  
  importParsedData.forEach((row, idx) => {
    html += `
      <tr style="border-bottom:1px solid #f3f4f6">
        <td style="padding:6px"><input type="text" value="${escapeHtml(row['项目'] || '')}" onchange="updateImportRow(${idx}, '项目', this.value)" style="width:100%;padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px"><input type="text" value="${escapeHtml(row['资源名称'] || '')}" onchange="updateImportRow(${idx}, '资源名称', this.value)" style="width:100%;padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px"><input type="text" value="${escapeHtml(row['资源类型'] || '')}" onchange="updateImportRow(${idx}, '资源类型', this.value)" style="width:100%;padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px"><input type="date" value="${row['资源开始时间'] || ''}" onchange="updateImportRow(${idx}, '资源开始时间', this.value)" style="padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px"><input type="date" value="${row['资源结束时间'] || ''}" onchange="updateImportRow(${idx}, '资源结束时间', this.value)" style="padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px"><input type="number" value="${(row['日平均工时'] !== undefined && row['日平均工时'] !== null && row['日平均工时'] !== '') ? row['日平均工时'] : 8}" min="0" max="24" step="0.5" onchange="updateImportRow(${idx}, '日平均工时', parseFloat(this.value) || 0)" style="width:70px;padding:4px 6px;border:1px solid #d1d5db;border-radius:4px;font-size:12px"></td>
        <td style="padding:6px;text-align:center"><button class="btn btn-secondary" style="padding:2px 8px;font-size:11px" onclick="deleteImportRow(${idx})">🗑️</button></td>
      </tr>
    `;
  });
  
  html += '</tbody></table>';
  tableEl.innerHTML = html;
}

function escapeHtml(str) {
  if (str === null || str === undefined) return '';
  return String(str).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;').replace(/"/g, '&quot;').replace(/'/g, '&#39;');
}

function updateImportRow(idx, field, value) {
  if (importParsedData[idx]) {
    importParsedData[idx][field] = value;
  }
}

function deleteImportRow(idx) {
  importParsedData.splice(idx, 1);
  renderImportPreview();
}

async function parseImportData() {
  try {
    // 判断当前Tab
    const activeTab = document.querySelector('.import-tab-btn[data-active="true"]');
    if (!activeTab) { alert('请选择导入方式'); return; }
    
    const tabType = activeTab.dataset.tab;
    let payload = {};
    
    if (tabType === 'text') {
      const text = document.getElementById('importTextArea').value.trim();
      if (!text) { alert('请粘贴Project数据'); return; }
      payload = { type: 'text', text: text };
    } else if (tabType === 'mpp') {
      const fileInput = document.getElementById('importMppFile');
      if (!fileInput.files || fileInput.files.length === 0) { alert('请选择.mpp文件'); return; }
      const formData = new FormData();
      formData.append('type', 'mpp');
      formData.append('file', fileInput.files[0]);
      const resp = await fetch('/api/import/parse', { method: 'POST', body: formData });
      if (!resp.ok) throw new Error('解析失败: ' + resp.status);
      const result = await resp.json();
      handleParseResult(result);
      return;
    } else if (tabType === 'pdf') {
      const fileInput = document.getElementById('importPdfFile');
      if (!fileInput.files || fileInput.files.length === 0) { alert('请选择PDF文件'); return; }
      const file = fileInput.files[0];
      console.log('[PDF导入] 选择文件:', file.name, '大小:', file.size, '字节');
      // 读取文件为base64
      const reader = new FileReader();
      reader.onload = async function(e) {
        try {
          const base64 = e.target.result.split(',')[1];  // 去掉data:application/pdf;base64,前缀
          console.log('[PDF导入] base64长度:', base64.length);
          const payload = { type: 'pdf', text: base64, filename: file.name };
          const resp = await fetch('/api/import/parse', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify(payload)
          });
          if (!resp.ok) throw new Error('解析失败: ' + resp.status);
          const result = await resp.json();
          console.log('[PDF导入] 解析结果:', result);
          handleParseResult(result);
        } catch (err) {
          console.error('[PDF导入] 错误:', err);
          alert('❌ 解析失败：' + (err.message || err));
        }
      };
      reader.onerror = function(e) {
        console.error('[PDF导入] 文件读取错误:', e);
        alert('❌ 文件读取失败');
      };
      reader.readAsDataURL(file);
      return;
    } else if (tabType === 'screenshot') {
      alert('OCR功能暂不可用');
      return;
    }
    
    const resp = await fetch('/api/import/parse', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(payload)
    });
    
    if (!resp.ok) throw new Error('解析失败: ' + resp.status);
    const result = await resp.json();
    handleParseResult(result);
  } catch (e) {
    console.error('[导入解析] 错误:', e);
    alert('❌ 解析失败：' + (e.message || e));
  }
}

function handleParseResult(result) {
  console.log('[导入] handleParseResult:', result);
  if (!result || result.success === false) {
    alert('解析失败：' + (result?.error || '未知错误'));
    return;
  }
  if (!result.rows || result.rows.length === 0) {
    alert('未解析到任何数据，请检查输入内容');
    return;
  }
  
  importParsedData = result.rows.map(r => ({
    '项目': r['项目'] || r['项目名称'] || r['project'] || '',
    '项目描述': r['项目描述'] || r['描述'] || '',
    '资源类型': r['资源类型'] || r['类型'] || r['resource_type'] || '',
    '资源名称': r['资源名称'] || r['负责人'] || r['resource'] || r['name'] || '',
    '资源开始时间': r['资源开始时间'] || r['开始时间'] || r['start'] || '',
    '资源结束时间': r['资源结束时间'] || r['结束时间'] || r['end'] || '',
    '日平均工时': (function() {
      const v = r['日平均工时'];
      if (v !== undefined && v !== null && v !== '' && !isNaN(parseFloat(v))) {
        return parseFloat(v);
      }
      // 后端未提供时，用工时自动计算（总工时÷工作日数）
      const work = parseFloat(r['工时'] || r['hours'] || 0);
      if (work > 0) {
        const s = r['开始时间'] || r['资源开始时间'] || '';
        const e = r['结束时间'] || r['资源结束时间'] || '';
        if (s && e) {
          // 粗略计算工作日（排除周末）
          let days = 0;
          const sd = new Date(s);
          const ed = new Date(e);
          if (!isNaN(sd.getTime()) && !isNaN(ed.getTime())) {
            for (let d = new Date(sd); d <= ed; d.setDate(d.getDate() + 1)) {
              if (d.getDay() !== 0 && d.getDay() !== 6) days++;
            }
          }
          if (days > 0) return Math.round(work / days * 10) / 10;
        }
      }
      return work > 0 ? work : 8;  // 兜底
    })()
  }));
  
  // 显示警告
  const warningsEl = document.getElementById('importWarnings');
  const warnings = [];
  if (result.warnings) {
    warnings.push(...result.warnings);
  }
  
  // === STEP 4: 检查TR标识（按SKILL.MD） ===
  const unresolvedTR = result.unresolved_tr || [];
  if (unresolvedTR.length > 0) {
    const trList = unresolvedTR.map((r, i) => `${i+1}. ${r}`).join('\\n');
    const msg = '⚠️ 有 ' + unresolvedTR.length + ' 条资源上级找不到 TR 标识：\\n\\n' + trList + '\\n\\n点击"确定"继续，点击"取消"停止。';
    if (!confirm(msg)) {
      warningsEl.innerHTML = '<div style="color:#ef4444">❌ 已取消：资源缺少TR标识，请确认数据后重试</div>';
      warningsEl.style.display = 'block';
      return;
    }
  }
  
  // === STEP 5: 检查project标识（按SKILL.MD） ===
  const unresolvedProject = result.unresolved_project || [];
  if (unresolvedProject.length > 0) {
    const projList = unresolvedProject.map((r, i) => `${i+1}. ${r}`).join('\\n');
    const msg = '⚠️ 有 ' + unresolvedProject.length + ' 条资源上级找不到 project 标识：\\n\\n' + projList + '\\n\\n点击"确定"继续，点击"取消"停止。';
    if (!confirm(msg)) {
      warningsEl.innerHTML = '<div style="color:#ef4444">❌ 已取消：资源缺少project标识，请确认数据后重试</div>';
      warningsEl.style.display = 'block';
      return;
    }
  }
  
  if (warnings.length > 0) {
    warningsEl.innerHTML = warnings.map(w => `<div>${w}</div>`).join('');
    warningsEl.style.display = 'block';
  } else {
    warningsEl.style.display = 'none';
  }
  
  // 显示部门选择和预览区域
  document.getElementById('importDeptRow').style.display = 'block';
  document.getElementById('importPreviewArea').style.display = 'block';
  renderImportPreview();
}

async function submitImport() {
  try {
    if (importParsedData.length === 0) {
      alert('没有可导入的数据');
      return;
    }
    
    const deptSelect = document.getElementById('importDeptSelect');
    const dept = deptSelect ? deptSelect.value : '';
    if (!dept) {
      if (!confirm('未选择部门，是否继续？（可后续在表格中补充）')) return;
    }
    
    // 按项目分组，计算G列（项目开始）和H列（项目结束）
    const projectMap = {};
    importParsedData.forEach(row => {
      const projName = row['项目'] || '未命名项目';
      if (!projectMap[projName]) {
        projectMap[projName] = { rows: [], startDates: [], endDates: [] };
      }
      projectMap[projName].rows.push(row);
      if (row['资源开始时间']) projectMap[projName].startDates.push(row['资源开始时间']);
      if (row['资源结束时间']) projectMap[projName].endDates.push(row['资源结束时间']);
    });
    
    // 为每个项目计算最早开始和最晚结束
    const finalRows = [];
    for (const projName in projectMap) {
      const info = projectMap[projName];
      const projStart = info.startDates.length > 0 ? info.startDates.sort()[0] : '1900-01-01';
      const projEnd = info.endDates.length > 0 ? info.endDates.sort().reverse()[0] : '2100-01-01';
      
      info.rows.forEach(row => {
        finalRows.push({
          '部门': dept,
          '项目': row['项目'] || projName,
          '项目开始时间': projStart,
          '项目结束时间': projEnd,
          '项目描述': row['项目描述'] || '',
          '资源类型': row['资源类型'] || '',
          '资源名称': row['资源名称'] || '',
          '资源开始时间': row['资源开始时间'] || '',
          '资源结束时间': row['资源结束时间'] || '',
          '日平均工时': parseFloat(row['日平均工时']) || 0,
          '已归档': false
        });
      });
    }
    
    // 提交到服务器
    const resp = await fetch('/api/import/commit', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ projects: finalRows })
    });
    
    if (!resp.ok) throw new Error('导入失败: ' + resp.status);
    const result = await resp.json();
    
    if (result.success) {
      // 刷新数据
      await refreshRawData();
      updateStats();
      renderTable();
      initResourceSearch();
      showSaved();
      const overlay = document.querySelector('.modal-overlay');
      if (overlay) overlay.remove();
      alert(`✅ 成功导入 ${result.count || finalRows.length} 条数据`);
    } else {
      alert('❌ 导入失败：' + (result.message || '未知错误'));
    }
  } catch (e) {
    console.error('[确认导入] 错误:', e);
    alert('❌ 导入出错：' + (e.message || e) + '\\n\\n请打开浏览器控制台(F12)查看详细错误信息');
  }
}

// ==================== 新增：添加项目弹窗 ====================
function openAddProjectModal() {
  const depts = Object.keys(RAW_DATA.depts || {}).sort();
  const resourceTypes = ['项目经理', 'SE设计', '固件工程师', '固件工程师-开发', '固件工程师-测试跟踪', 
                         '固件工程师-认证跟踪', '固件工程师-协议确认', '固件工程师-开发设计',
                         '固件工程师-集成测试配合', '硬件工程师', '硬件工程师-压力传感器测试',
                         '项目CHAT', 'DLMS-BETA版本功能调整', '测试工程师'];
  const persons = Object.keys(getAllEmails()).sort();
  
  const modal = document.createElement('div');
  modal.className = 'modal-overlay';
  modal.innerHTML = `
    <div class="modal-box">
      <button class="modal-close" onclick="this.closest('.modal-overlay').remove()">×</button>
      <h3>➕ 添加新项目/资源</h3>
      <div class="form-row">
        <div class="form-group">
          <label>市场/部门</label>
          <select id="newProjDept">
            <option value="">请选择</option>
            ${depts.map(d => `<option value="${d}">${d}</option>`).join('')}
            <option value="__custom__">+ 新增市场...</option>
          </select>
        </div>
        <div class="form-group">
          <label>项目名称 *</label>
          <input type="text" id="newProjName" placeholder="如：巴西NB-V3">
        </div>
      </div>
      <div class="form-group">
        <label>项目描述</label>
        <input type="text" id="newProjDesc" placeholder="如：TR4A-测试阶段">
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>资源类型 *</label>
          <select id="newResType">
            <option value="">请选择</option>
            ${resourceTypes.map(t => `<option value="${t}">${t}</option>`).join('')}
            <option value="__custom__">+ 新增类型...</option>
          </select>
        </div>
        <div class="form-group">
          <label>负责人 *</label>
          <select id="newResPerson">
            <option value="">请选择</option>
            ${persons.map(p => `<option value="${p}">${p}</option>`).join('')}
            <option value="__custom__">+ 新增成员...</option>
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>开始时间 *</label>
          <input type="date" id="newStartDate">
        </div>
        <div class="form-group">
          <label>结束时间 *</label>
          <input type="date" id="newEndDate">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>日平均工时 (h/天)</label>
          <input type="number" id="newHours" value="8" min="0" max="24" step="0.5">
        </div>
        <div class="form-group">
          <label>项目开始时间</label>
          <input type="date" id="newProjStart">
        </div>
      </div>
      <div class="form-group">
        <label>项目结束时间</label>
        <input type="date" id="newProjEnd">
      </div>
      <div class="form-actions">
        <button class="btn btn-secondary" onclick="this.closest('.modal-overlay').remove()">取消</button>
        <button class="btn btn-success" onclick="submitNewProject()">✅ 添加项目</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  
  // 处理新增选项
  document.getElementById('newProjDept').addEventListener('change', function() {
    if (this.value === '__custom__') {
      const custom = prompt('请输入新市场/部门名称：');
      if (custom) {
        const opt = document.createElement('option');
        opt.value = custom;
        opt.textContent = custom;
        opt.selected = true;
        this.insertBefore(opt, this.querySelector('option[value="__custom__"]'));
      } else {
        this.value = '';
      }
    }
  });
  
  document.getElementById('newResType').addEventListener('change', function() {
    if (this.value === '__custom__') {
      const custom = prompt('请输入新资源类型：');
      if (custom) {
        const opt = document.createElement('option');
        opt.value = custom;
        opt.textContent = custom;
        opt.selected = true;
        this.insertBefore(opt, this.querySelector('option[value="__custom__"]'));
      } else {
        this.value = '';
      }
    }
  });
  
  document.getElementById('newResPerson').addEventListener('change', function() {
    if (this.value === '__custom__') {
      const name = prompt('请输入新成员姓名：');
      if (name) {
        const suggestedEmail = generateEmailFromName(name);
        const email = prompt('请确认/修改邮箱地址：', suggestedEmail);
        if (email) {
          customEmails[name] = email;
          localStorage.setItem('customEmails', JSON.stringify(customEmails));
          collabMarkDirty();  // 协作模式：标记需要同步
          const opt = document.createElement('option');
          opt.value = name;
          opt.textContent = name;
          opt.selected = true;
          this.insertBefore(opt, this.querySelector('option[value="__custom__"]'));
          initResourceSearch();  // 刷新人员检索下拉列表
          alert(`✅ 已添加新成员：${name} <${email}>`);
        } else {
          this.value = '';
        }
      } else {
        this.value = '';
      }
    }
  });
}

async function submitNewProject() {
  try {
    const deptEl = document.getElementById('newProjDept');
    const projNameEl = document.getElementById('newProjName');
    const projDescEl = document.getElementById('newProjDesc');
    const resTypeEl = document.getElementById('newResType');
    const resPersonEl = document.getElementById('newResPerson');
    const startDateEl = document.getElementById('newStartDate');
    const endDateEl = document.getElementById('newEndDate');
    const hoursEl = document.getElementById('newHours');
    const projStartEl = document.getElementById('newProjStart');
    const projEndEl = document.getElementById('newProjEnd');
    
    if (!deptEl || !projNameEl) {
      alert('错误：表单元素未找到，请刷新页面重试');
      return;
    }
    
    const dept = deptEl.value;
    const projName = projNameEl.value.trim();
    const projDesc = projDescEl ? projDescEl.value.trim() : '';
    const resType = resTypeEl.value;
    const resPerson = resPersonEl.value;
    const startDate = startDateEl.value;
    const endDate = endDateEl.value;
    const hours = parseFloat(hoursEl.value) || 0;
    const projStart = projStartEl ? (projStartEl.value || '') : '';
    const projEnd = projEndEl ? (projEndEl.value || '') : '';
    
    if (!projName) { alert('请填写项目名称'); return; }
    if (!resType) { alert('请选择资源类型'); return; }
    if (!resPerson) { alert('请选择负责人'); return; }
    if (!startDate) { alert('请选择开始时间'); return; }
    if (!endDate) { alert('请选择结束时间'); return; }
    
    const newProject = {
      部门: dept,
      项目: projName,
      项目开始时间: projStart || '1900-01-01',
      项目结束时间: projEnd || '2100-01-01',
      项目描述: projDesc,
      资源类型: resType,
      资源名称: resPerson,
      资源开始时间: startDate,
      资源结束时间: endDate,
      日平均工时: hours,
      已归档: false
    };
    
    if (collabIsEnabled()) {
      // 【简化方案】协作模式：直接调用API
      const result = await callActionApi('add', newProject);
      if (result && result.success) {
        updateStats();
        renderTable();
        initResourceSearch();
        showSaved();
        document.querySelector('.modal-overlay').remove();
        alert(`✅ 已添加新项目：${projName} - ${resType}（${resPerson}）`);
      } else if (result && !result.success) {
        alert('❌ 添加失败：' + (result.message || '未知错误'));
      }
    } else {
    // 非协作模式：本地逻辑（兼容）
    const today = new Date(RAW_DATA.today);
    const end = new Date(endDate);
    const resDays = Math.ceil((end - today) / (1000 * 60 * 60 * 24));
    let projDays = null;
    if (projEnd) {
      const pEnd = new Date(projEnd);
      projDays = Math.ceil((pEnd - today) / (1000 * 60 * 60 * 24));
    }
    const newId = Math.max(...RAW_DATA.allProjects.map(p => p.id)) + 1;
    const proj = { ...newProject, id: newId, 项目剩余天数: projDays, 资源剩余天数: resDays };
    RAW_DATA.allProjects.push(proj);
    if (!RAW_DATA.depts[dept]) RAW_DATA.depts[dept] = [];
    RAW_DATA.depts[dept].push(proj);
    if (resDays !== null && resDays <= 3) {
      RAW_DATA.delayedProjects.push(proj);
    }
    newProjects.push(proj);
    localStorage.setItem('newProjects', JSON.stringify(newProjects));
    updateStats();
    renderTable();
    initResourceSearch();
    showSaved();
    collabMarkDirty();
    document.querySelector('.modal-overlay').remove();
    alert(`✅ 已添加新项目：${projName} - ${resType}（${resPerson}）`);
    }
  } catch (e) {
    console.error('[添加项目] 错误:', e);
    alert('❌ 添加项目出错：' + (e.message || e) + '\\n\\n请打开浏览器控制台(F12)查看详细错误信息');
  }
}

// ==================== 新增：成员管理弹窗 ====================
function openMemberManager() {
  const allEmails = getAllEmails();
  const names = Object.keys(allEmails).sort();
  
  const modal = document.createElement('div');
  modal.className = 'modal-overlay';
  
  let listHtml = '';
  names.forEach(name => {
    const email = allEmails[name];
    const isBuiltin = ENGINEER_EMAILS.hasOwnProperty(name);
    const sourceTag = isBuiltin 
      ? '<span class="tag-builtin">内置</span>' 
      : '<span class="tag-custom">自定义</span>';
    const deleteBtn = isBuiltin 
      ? '' 
      : `<button class="btn btn-secondary" style="padding:4px 10px;font-size:12px" onclick="deleteMember('${name}')">删除</button>`;
    
    listHtml += `
      <div class="member-item">
        <div class="member-info">
          <span class="member-name">${name}</span>
          <span class="member-email">${email}</span>
          ${sourceTag}
        </div>
        ${deleteBtn}
      </div>
    `;
  });
  
  modal.innerHTML = `
    <div class="modal-box">
      <button class="modal-close" onclick="this.closest('.modal-overlay').remove()">×</button>
      <h3>👥 成员管理（共 ${names.length} 人）</h3>
      <div class="form-group">
        <label>添加新成员</label>
        <div class="form-row">
          <input type="text" id="newMemberName" placeholder="姓名（如：张三）" style="flex:1">
          <input type="text" id="newMemberEmail" placeholder="邮箱（自动生成）" style="flex:2">
        </div>
        <div id="emailSuggestBox" style="display:none"></div>
        <div class="form-actions" style="border:none;padding:0;margin:10px 0 16px 0">
          <button class="btn btn-secondary" onclick="document.getElementById('newMemberName').value='';document.getElementById('newMemberEmail').value=''">清空</button>
          <button class="btn" style="background:#06b6d4;color:white" onclick="document.getElementById('importMembersFile').click()">📥 导入成员</button>
          <input type="file" id="importMembersFile" accept=".xlsx,.xls,.csv,.json" style="display:none" onchange="importMembersFromFile(event)">
          <button class="btn btn-success" onclick="addNewMember()">➕ 添加成员</button>
        </div>
        <div style="font-size:12px;color:#6b7280;background:#f0f9ff;padding:8px 12px;border-radius:6px">
          💡 支持导入 Excel/CSV/JSON：每行格式为「姓名,邮箱」或「姓名, 邮箱」，也支持从原始户表脚本Excel自动提取资源人员
        </div>
      </div>
      <div class="form-group">
        <label>成员列表</label>
        <div class="member-list">
          ${listHtml || '<div style="padding:20px;text-align:center;color:#9ca3af">暂无成员</div>'}
        </div>
      </div>
      <div class="form-actions">
        <button class="btn btn-secondary" onclick="this.closest('.modal-overlay').remove()">关闭</button>
      </div>
    </div>
  `;
  document.body.appendChild(modal);
  
  // 姓名输入时自动生成邮箱建议
  const nameInput = document.getElementById('newMemberName');
  const emailInput = document.getElementById('newMemberEmail');
  nameInput.addEventListener('input', function() {
    const suggested = generateEmailFromName(this.value);
    emailInput.value = suggested;
    if (suggested && !suggested.includes('[') && this.value.length >= 2) {
      document.getElementById('emailSuggestBox').innerHTML = 
        `<div class="email-suggest" onclick="document.getElementById('newMemberEmail').value='${suggested}'">
           💡 建议邮箱：${suggested}（点击使用）
         </div>`;
      document.getElementById('emailSuggestBox').style.display = 'block';
    } else {
      document.getElementById('emailSuggestBox').style.display = 'none';
    }
  });
}

function addNewMember() {
  const name = document.getElementById('newMemberName').value.trim();
  const email = document.getElementById('newMemberEmail').value.trim();
  
  if (!name) { alert('请输入成员姓名'); return; }
  if (!email || !email.includes('@')) { alert('请输入有效的邮箱地址'); return; }
  
  const allEmails = getAllEmails();
  if (allEmails[name]) {
    if (!confirm(`成员「${name}」已存在（${allEmails[name]}），是否覆盖为新邮箱？`)) {
      return;
    }
  }
  
  customEmails[name] = email;
  localStorage.setItem('customEmails', JSON.stringify(customEmails));
  collabMarkDirty();  // 协作模式：标记需要同步
  
  alert(`✅ 已添加成员：${name} <${email}>`);
  
  // 刷新弹窗
  document.querySelector('.modal-overlay').remove();
  initResourceSearch();  // 刷新人员检索下拉列表
  openMemberManager();
}

function deleteMember(name) {
  if (!confirm(`确定要删除成员「${name}」吗？`)) return;
  delete customEmails[name];
  localStorage.setItem('customEmails', JSON.stringify(customEmails));
  collabMarkDirty();  // 协作模式：标记需要同步
  document.querySelector('.modal-overlay').remove();
  initResourceSearch();  // 刷新人员检索下拉列表
  openMemberManager();
}

// 从文件导入成员（支持 Excel/CSV/JSON）
function importMembersFromFile(event) {
  const file = event.target.files[0];
  if (!file) return;
  
  const fileName = file.name.toLowerCase();
  const reader = new FileReader();
  
  reader.onload = function(e) {
    try {
      let members = [];
      
      if (fileName.endsWith('.json')) {
        // JSON 格式
        const data = JSON.parse(e.target.result);
        if (Array.isArray(data)) {
          members = data.map(item => {
            const name = item.name || item.姓名 || item['成员'] || '';
            const email = item.email || item.邮箱 || item['邮件'] || '';
            return { name: name.trim(), email: email.trim() };
          }).filter(m => m.name);
        } else if (typeof data === 'object') {
          members = Object.keys(data).map(name => ({
            name: name.trim(),
            email: (data[name] || '').trim()
          })).filter(m => m.name);
        }
      } else if (fileName.endsWith('.csv')) {
        // CSV 格式
        const text = e.target.result;
        const lines = text.split(/\\r?\\n/).filter(l => l.trim());
        for (const line of lines) {
          // 跳过表头
          if (line.includes('姓名') && line.includes('邮箱')) continue;
          const parts = line.split(/[,，\\t]/).map(p => p.trim());
          if (parts.length >= 2) {
            const name = parts[0];
            // 找邮箱（含 @ 的字段）
            let email = parts[1];
            for (const p of parts) {
              if (p.includes('@')) { email = p; break; }
            }
            if (name) members.push({ name, email });
          } else if (parts.length === 1 && parts[0]) {
            // 只有姓名，自动生成邮箱
            members.push({ name: parts[0], email: generateEmailFromName(parts[0]) });
          }
        }
      } else if (fileName.endsWith('.xlsx') || fileName.endsWith('.xls')) {
        // Excel 格式
        const data = new Uint8Array(e.target.result);
        const workbook = XLSX.read(data, { type: 'array' });
        const sheetName = workbook.SheetNames[0];
        const sheet = workbook.Sheets[sheetName];
        const rows = XLSX.utils.sheet_to_json(sheet, { header: 1 });
        
        // 查找姓名和邮箱列的索引
        let nameColIdx = -1;
        let emailColIdx = -1;
        let resourceColIdx = -1;
        let startRow = 0;
        
        if (rows.length > 0) {
          const headerRow = rows[0];
          for (let i = 0; i < headerRow.length; i++) {
            const h = String(headerRow[i] || '').trim();
            if (h.includes('姓名') || h.includes('名字') || h === 'name' || h === 'Name') {
              nameColIdx = i;
            } else if (h.includes('邮箱') || h.includes('邮件') || h === 'email' || h === 'Email') {
              emailColIdx = i;
            } else if (h.includes('资源') || h.includes('负责人') || h.includes('人员')) {
              resourceColIdx = i;
            }
          }
          startRow = 1;
        }
        
        // 如果没找到明确的姓名/邮箱列，尝试从资源人员列提取
        if (nameColIdx === -1 && resourceColIdx !== -1) {
          nameColIdx = resourceColIdx;
        }
        
        // 如果还找不到，默认第0列为姓名，第1列为邮箱
        if (nameColIdx === -1) nameColIdx = 0;
        if (emailColIdx === -1) emailColIdx = 1;
        
        const seenNames = new Set();
        for (let r = startRow; r < rows.length; r++) {
          const row = rows[r];
          if (!row) continue;
          
          let name = String(row[nameColIdx] || '').trim();
          let email = String(row[emailColIdx] || '').trim();
          
          // 处理多个资源人员（可能是"张三、李四"或"张三;李四"格式）
          if (name && (name.includes('、') || name.includes(';') || name.includes('；') || name.includes(','))) {
            const names = name.split(/[、;；,]/).map(n => n.trim()).filter(n => n);
            for (const n of names) {
              if (!seenNames.has(n)) {
                seenNames.add(n);
                members.push({ name: n, email: generateEmailFromName(n) });
              }
            }
            continue;
          }
          
          if (name && !seenNames.has(name)) {
            seenNames.add(name);
            // 如果没有邮箱则自动生成
            if (!email || !email.includes('@')) {
              email = generateEmailFromName(name);
            }
            members.push({ name, email });
          }
        }
      }
      
      // 过滤有效成员
      const validMembers = members.filter(m => m.name && m.name.length >= 2);
      
      if (validMembers.length === 0) {
        alert('未找到有效的成员数据，请检查文件格式。\\n\\n支持格式：\\n- CSV: 姓名,邮箱\\n- Excel: 含「姓名」「邮箱」列或「资源人员」列\\n- JSON: [{name,email} 或 {"姓名":"邮箱"}');
        return;
      }
      
      // 统计新增和已存在的
      const allEmails = getAllEmails();
      let newCount = 0;
      let updateCount = 0;
      let skipped = [];
      
      for (const m of validMembers) {
        if (ENGINEER_EMAILS.hasOwnProperty(m.name)) {
          skipped.push(m.name + '（内置）');
        } else if (allEmails[m.name]) {
            if (allEmails[m.name] !== m.email) {
              customEmails[m.name] = m.email;
              updateCount++;
            }
          } else {
            customEmails[m.name] = m.email;
            newCount++;
          }
        }
      
      localStorage.setItem('customEmails', JSON.stringify(customEmails));
      collabMarkDirty();
      
      let msg = `✅ 导入完成！\\n\\n共解析 ${validMembers.length} 人\\n新增：${newCount} 人`;
      if (updateCount > 0) msg += `\\n更新：${updateCount} 人`;
      if (skipped.length > 0) msg += `\\n跳过（内置成员）：${skipped.length} 人`;
      alert(msg);
      
      // 刷新弹窗
      document.querySelector('.modal-overlay').remove();
      initResourceSearch();
      openMemberManager();
      
    } catch (err) {
      console.error(err);
      alert('导入失败：' + err.message);
    }
    
    // 清空文件输入，允许重复选择同一文件
    event.target.value = '';
  };
  
  if (fileName.endsWith('.xlsx') || fileName.endsWith('.xls')) {
    reader.readAsArrayBuffer(file);
  } else {
    reader.readAsText(file, 'UTF-8');
  }
}

init();
</script>
</body>
</html>
'''

html = html_template.replace('__DATA_PLACEHOLDER__', json.dumps(data_json, ensure_ascii=False))
html = html.replace('__SOURCE_FILE__', display_source_name)

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"✅ HTML已生成: {OUTPUT_HTML}")

# ==================== 5. 生成Excel ====================
print(f"📊 生成Excel: {OUTPUT_EXCEL}")

wb = Workbook()

HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
DELAYED_FILL = PatternFill('solid', fgColor='FEE2E2')
URGENT_FILL = PatternFill('solid', fgColor='FFEDD5')
ZEBRA1 = PatternFill('solid', fgColor='FFFFFF')
ZEBRA2 = PatternFill('solid', fgColor='F7F9FC')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9DEE7'),
    right=Side(style='thin', color='D9DEE7'),
    top=Side(style='thin', color='D9DEE7'),
    bottom=Side(style='thin', color='D9DEE7')
)
TITLE_FONT = Font(bold=True, size=14, color='1F2937')
SUBTITLE_FONT = Font(size=11, color='6B7280')

ws1 = wb.active
ws1.title = '延期预警'

ws1['A1'] = '📋 项目延期点检表'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:M1')
ws1['A2'] = f'点检日期: {today.strftime("%Y-%m-%d")}  |  预警截止: {(today + timedelta(days=3)).strftime("%Y-%m-%d")}  |  数据来源: {os.path.basename(SOURCE_FILE)}'
ws1['A2'].font = SUBTITLE_FONT
ws1.merge_cells('A2:M2')

ws1['A4'] = '统计概览'
ws1['A4'].font = Font(bold=True, size=12)

delayed_count = len([p for p in delayed_projects if get_effective_days(p) is not None and get_effective_days(p) < 0])
urgent_count = len([p for p in delayed_projects if get_effective_days(p) is not None and 0 <= get_effective_days(p) <= 3])

stats_data = [
    ['已延期', '3天内到期', '7天内到期', '总资源记录'],
    [delayed_count, urgent_count, stats['warning'], stats['total']]
]
for r, row in enumerate(stats_data):
    for c, val in enumerate(row):
        cell = ws1.cell(row=4+r, column=1+c, value=val)
        cell.border = THIN_BORDER
        if r == 0:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='EEF2FF')
        cell.alignment = Alignment(horizontal='center')

headers = ['预警级别', '状态', '部门', '项目名称', '项目描述', '项目结束时间', '资源类型', '负责人', '资源结束时间', '剩余天数', '日平均工时', '点检备注', '点检确认']
for c, h in enumerate(headers):
    cell = ws1.cell(row=7, column=1+c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

for r, p in enumerate(delayed_projects):
    row_num = 8 + r
    days = get_effective_days(p)
    status_label = '已延期' if days < 0 else f'{days}天后到期'
    level = '🔴 严重' if days < 0 else ('🟠 紧急' if days <= 1 else '🟡 注意')
    
    fill = DELAYED_FILL if days < 0 else (URGENT_FILL if days <= 1 else ZEBRA2)
    if r % 2 == 0 and days >= 2:
        fill = ZEBRA1
    
    def parse_date_for_excel(s):
        if not s: return None
        try:
            return datetime.strptime(s, '%Y-%m-%d')
        except:
            return s
    
    values = [
        level, status_label, p['部门'], p['项目'], p['项目描述'],
        parse_date_for_excel(p['项目结束时间']), p['资源类型'], p['资源名称'],
        parse_date_for_excel(p['资源结束时间']), days, p['日平均工时'], '', ''
    ]
    
    for c, val in enumerate(values):
        cell = ws1.cell(row=row_num, column=1+c, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if c in [9, 10]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

col_widths = [12, 12, 14, 22, 30, 14, 18, 10, 14, 10, 10, 20, 10]
for i, w in enumerate(col_widths):
    ws1.column_dimensions[get_column_letter(i+1)].width = w

ws1.row_dimensions[7].height = 30
ws1.freeze_panes = 'A8'

ws2 = wb.create_sheet('全部项目')
all_headers = ['部门', '项目名称', '项目开始时间', '项目结束时间', '项目剩余天数', '项目描述', '资源类型', '负责人', '资源开始时间', '资源结束时间', '资源剩余天数', '日平均工时']
for c, h in enumerate(all_headers):
    cell = ws2.cell(row=1, column=1+c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

for r, p in enumerate(projects):
    row_num = 2 + r
    fill = ZEBRA1 if r % 2 == 0 else ZEBRA2
    days = p['项目剩余天数'] if p['项目剩余天数'] is not None else p['资源剩余天数']
    if days is not None and days < 0:
        fill = DELAYED_FILL
    elif days is not None and days <= 3:
        fill = URGENT_FILL
    
    def parse_date_for_excel(s):
        if not s: return None
        try:
            return datetime.strptime(s, '%Y-%m-%d')
        except:
            return s
    
    values = [p['部门'], p['项目'], parse_date_for_excel(p['项目开始时间']), parse_date_for_excel(p['项目结束时间']), p['项目剩余天数'],
              p['项目描述'], p['资源类型'], p['资源名称'], parse_date_for_excel(p['资源开始时间']), parse_date_for_excel(p['资源结束时间']),
              p['资源剩余天数'], p['日平均工时']]
    for c, val in enumerate(values):
        cell = ws2.cell(row=row_num, column=1+c, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER

all_widths = [14, 22, 14, 14, 10, 30, 18, 10, 14, 14, 10, 10]
for i, w in enumerate(all_widths):
    ws2.column_dimensions[get_column_letter(i+1)].width = w
ws2.freeze_panes = 'A2'

ws3 = wb.create_sheet('更新说明')
ws3['A1'] = '📖 计划更新操作指南'
ws3['A1'].font = TITLE_FONT
instructions = [
    '',
    '【方式A：修改原始Excel（推荐用于大量变更）】',
    '  1. 打开原始文件：' + os.path.basename(SOURCE_FILE),
    '  2. 修改「任务计划表」中的开始时间、结束时间等',
    '  3. 保存Excel文件',
    '  4. 双击运行：更新点检表.py （或命令行执行 python 更新点检表.py）',
    '  5. 脚本会自动重新生成 HTML 和 Excel 版本的点检表',
    '',
    '【方式B：HTML页面直接编辑（推荐用于日常微调）】',
    '  1. 打开：项目延期点检表.html',
    '  2. 点击工具栏的「✏️ 编辑模式」按钮',
    '  3. 直接点击单元格修改：',
    '     • 负责人：点击名字即可修改',
    '     • 结束时间：点击日期选择器修改',
    '  4. 修改后自动保存到浏览器本地',
    '  5. 需要时可「导出CSV」备份',
    '',
    '【两种方式对比】',
    '  方式A：数据准确，源头统一，适合批量修改',
    '  方式B：方便快捷，无需打开Excel，适合日常调整',
    '',
    '【注意事项】',
    '  • 方式B的修改仅保存在当前浏览器，清除浏览器数据会丢失',
    '  • 重要变更建议使用方式A，修改原始Excel并重新生成',
    '  • 每次运行脚本会用最新的Excel数据覆盖旧版本',
]
for i, text in enumerate(instructions):
    ws3.cell(row=2+i, column=1, value=text)

ws3.column_dimensions['A'].width = 100

wb.save(OUTPUT_EXCEL)
print(f"✅ Excel已生成: {OUTPUT_EXCEL}")

# ==================== 邮件发送 ====================
import sys

# 解析命令行参数
send_email = '--send-email' in sys.argv or '-e' in sys.argv
test_email = '--test-email' in sys.argv or '-t' in sys.argv

if test_email:
    print(f"\n{'='*50}")
    print("📧 测试邮件连接...")
    print(f"{'='*50}")
    try:
        from email_sender import test_email_connection
        test_email_connection()
    except ImportError:
        print("❌ 找不到 email_sender.py 模块")
    sys.exit(0)

if send_email:
    print(f"\n{'='*50}")
    print("📧 发送延期提醒邮件...")
    print(f"{'='*50}")
    try:
        from email_sender import send_delay_emails
        send_delay_emails(delayed_projects, projects)
    except ImportError:
        print("❌ 找不到 email_sender.py 模块")
    except Exception as e:
        print(f"❌ 邮件发送异常: {e}")

# ==================== 完成 ====================
print(f"\n{'='*50}")
print(f"🎉 更新完成！生成日期：{today.strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*50}")
print(f"📄 HTML版本：{OUTPUT_HTML}")
print(f"📊 Excel版本：{OUTPUT_EXCEL}")
print(f"🔄 下次更新：修改原始Excel后运行 python 更新点检表.py")
print(f"📧 发送邮件：python 更新点检表.py --send-email")
print(f"🔌 测试邮件：python 更新点检表.py --test-email")
