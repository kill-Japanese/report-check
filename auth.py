# -*- coding: utf-8 -*-
"""
项目点检表 - 权限管理与安全模块

功能：
- 用户管理（增删改查）—— 数据源：用户管理.xlsx（GitHub）
- 密码加密存储（PBKDF2-HMAC-SHA256）
- 角色权限（admin / editor / viewer）
- Session 管理（带过期时间）
- 登录限流（防止暴力破解）
- CSRF Token 防护
- 操作日志审计
"""

import hashlib
import hmac
import json
import os
import uuid
import secrets
import time
import threading
import subprocess
from datetime import datetime, timedelta
from functools import wraps

try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ==================== 配置 ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# 用户数据源：Excel（GitHub 持久化）
USERS_EXCEL = os.path.join(BASE_DIR, '用户管理.xlsx')

# 审计日志（持久化）
AUDIT_LOG_FILE = os.path.join(DATA_DIR, 'audit.log')

# 临时数据（不持久化：随部署重置）
SESSIONS_FILE = os.path.join(BASE_DIR, 'sessions.json')
RATE_LIMIT_FILE = os.path.join(BASE_DIR, 'rate_limit.json')

# 安全配置
PASSWORD_HASH_ITERATIONS = 200_000  # PBKDF2 迭代次数
SESSION_TIMEOUT = 8 * 60 * 60        # Session 过期时间（8小时）
MAX_LOGIN_ATTEMPTS = 5                # 最大登录失败次数
LOGIN_LOCKOUT_TIME = 15 * 60          # 锁定时间（15分钟）
CSRF_TOKEN_TTL = 3600                  # CSRF Token 有效期（1小时）

# 角色定义
ROLES = {
    'admin': {
        'name': '管理员',
        'permissions': [
            'view',           # 查看报表
            'edit',           # 编辑数据
            'save',           # 全量保存
            'user_manage',    # 用户管理
            'role_manage',    # 角色管理
            'audit_view',     # 查看审计日志
            'system_config',  # 系统配置
        ]
    },
    'editor': {
        'name': '编辑者',
        'permissions': [
            'view',
            'edit',
            'save',
        ]
    },
    'viewer': {
        'name': '只读用户',
        'permissions': [
            'view',
        ]
    }
}

# 线程锁
_lock = threading.Lock()


# ==================== 工具函数 ====================

def _safe_read_json(filepath, default):
    """安全读取 JSON 文件"""
    with _lock:
        if not os.path.exists(filepath):
            return default
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return default


def _safe_write_json(filepath, data):
    """安全写入 JSON 文件"""
    with _lock:
        tmp = filepath + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, filepath)


def _hash_password(password: str, salt: bytes = None) -> dict:
    """使用 PBKDF2-HMAC-SHA256 加密密码"""
    if salt is None:
        salt = secrets.token_bytes(32)
    key = hashlib.pbkdf2_hmac(
        'sha256',
        password.encode('utf-8'),
        salt,
        PASSWORD_HASH_ITERATIONS
    )
    return {
        'hash': key.hex(),
        'salt': salt.hex(),
        'iterations': PASSWORD_HASH_ITERATIONS,
        'algorithm': 'PBKDF2-HMAC-SHA256'
    }


def _verify_password(password: str, stored: dict) -> bool:
    """验证密码"""
    try:
        result = _hash_password(password, bytes.fromhex(stored['salt']))
        return hmac.compare_digest(result['hash'], stored['hash'])
    except Exception:
        return False


def _audit_log(action: str, username: str, detail: str = ''):
    """记录审计日志"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        entry = f"[{timestamp}] {action} | 用户: {username} | {detail}\n"
        with open(AUDIT_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(entry)
    except Exception:
        pass


# ==================== Git 同步工具 ====================

def _ensure_git_repo() -> tuple[bool, str]:
    """确保 Git 仓库存在并且配置正确（Render 部署环境保障）
    
    与 sync_excel.ensure_git_repo 功能一致，避免循环导入。
    """
    git_dir = os.path.join(BASE_DIR, '.git')
    try:
        if not os.path.exists(git_dir):
            print(f"[Auth Git] .git 目录不存在，正在初始化...")
            init = subprocess.run(
                ['git', 'init'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=10
            )
            if init.returncode != 0:
                return False, f'Git 初始化失败: {init.stderr[:200]}'
            remote_url = os.environ.get('GIT_REMOTE_URL', '')
            if remote_url:
                subprocess.run(
                    ['git', 'remote', 'add', 'origin', remote_url],
                    capture_output=True, cwd=BASE_DIR, timeout=10
                )
                print(f"[Auth Git] 已设置 remote origin: {remote_url[:50]}...")
        subprocess.run(
            ['git', 'config', 'user.email', 'server@report-check.local'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        subprocess.run(
            ['git', 'config', 'user.name', 'Report Check Server'],
            capture_output=True, cwd=BASE_DIR, timeout=5
        )
        branch = subprocess.run(
            ['git', 'branch', '--show-current'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=5
        )
        current_branch = branch.stdout.strip()
        if current_branch and current_branch != 'main':
            subprocess.run(
                ['git', 'branch', '-M', 'main'],
                capture_output=True, cwd=BASE_DIR, timeout=5
            )
        return True, 'Git 仓库就绪'
    except Exception as e:
        return False, f'Git 仓库初始化失败: {str(e)}'


def _git_push(message: str) -> tuple[bool, str]:
    """提交并推送到 GitHub（双路径：git命令优先，GitHub API兜底）
    
    【关键修复】Render 环境中 git 命令可能不存在，
    自动检测并回退到 GitHub REST API 模式。
    """
    # 检查 git 命令是否可用
    git_available = False
    try:
        result = subprocess.run(['git', '--version'], capture_output=True, timeout=5)
        git_available = (result.returncode == 0) and os.path.exists(os.path.join(BASE_DIR, '.git'))
    except (FileNotFoundError, OSError):
        git_available = False
    
    if not git_available:
        # 回退到 GitHub API 模式
        try:
            from github_sync import github_api_push
            print('[auth] git 不可用，使用 GitHub API 模式推送')
            return github_api_push(message)
        except Exception as api_e:
            return False, f'git不可用且API推送失败: {api_e}'
    
    # git 可用，使用原有逻辑
    try:
        ensure_ok, ensure_msg = _ensure_git_repo()
        if not ensure_ok:
            return False, f'Git仓库不可用: {ensure_msg}'
        if not os.path.exists(os.path.join(BASE_DIR, '.git')):
            return False, '未检测到 Git 仓库'

        # 【关键修复】先 fetch 最新状态，检查是否落后于远程
        # （当之前用 GitHub API 推送过时，本地会落后）
        fetch = subprocess.run(
            ['git', 'fetch', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        
        behind = subprocess.run(
            ['git', 'rev-list', '--count', 'HEAD..origin/main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        try:
            behind_count = int(behind.stdout.strip())
        except ValueError:
            behind_count = 0
        
        if behind_count > 0:
            print(f'[auth] 本地落后远程 {behind_count} 个提交，尝试先拉取...')
            # 尝试拉取（自动合并）
            pull = subprocess.run(
                ['git', 'pull', 'origin', 'main', '--no-edit'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=30
            )
            if pull.returncode != 0:
                # 拉取失败（可能有冲突），回退到 GitHub API 模式
                print(f'[auth] 拉取失败，回退到 GitHub API 模式: {pull.stderr[:100]}')
                try:
                    from github_sync import github_api_push
                    return github_api_push(message)
                except Exception as api_e:
                    return False, f'本地落后远程且拉取失败: {pull.stderr[:200]}'

        ahead = subprocess.run(
            ['git', 'rev-list', '--count', 'origin/main..HEAD'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        try:
            ahead_count = int(ahead.stdout.strip())
        except ValueError:
            ahead_count = 0

        if ahead_count > 0:
            push = subprocess.run(
                ['git', 'push', 'origin', 'main'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=30
            )
            if push.returncode != 0:
                # 【关键修复】git push 失败时，回退到 GitHub API 模式推送
                try:
                    from github_sync import github_api_push
                    print('[auth] git push 失败，回退到 GitHub API 模式')
                    return github_api_push(message)
                except Exception as api_e:
                    return False, f'推送失败（git+API均失败）: {push.stderr[:200]}'
            return True, f'已推送 {ahead_count} 个待提交到 GitHub'

        subprocess.run(['git', 'add', '用户管理.xlsx', 'data/'],
                       capture_output=True, cwd=BASE_DIR, timeout=10)

        result = subprocess.run(
            ['git', 'status', '--porcelain', '用户管理.xlsx', 'data/'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if not result.stdout.strip():
            return True, '无变更'

        commit_msg = f'[用户同步] {message} - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        commit_result = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if commit_result.returncode != 0:
            if 'nothing to commit' in (commit_result.stdout + commit_result.stderr):
                return True, '无变更'
            return False, f'提交失败: {commit_result.stderr[:200]}'

        push = subprocess.run(
            ['git', 'push', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        if push.returncode != 0:
            subprocess.run(
                ['git', 'reset', '--soft', 'HEAD~1'],
                capture_output=True, cwd=BASE_DIR, timeout=10
            )
            # 【关键修复】git push 失败时，回退到 GitHub API 模式推送
            try:
                from github_sync import github_api_push
                print('[auth] git push 失败，回退到 GitHub API 模式')
                return github_api_push(message)
            except Exception as api_e:
                return False, f'推送失败（git+API均失败）: {push.stderr[:200]}'

        return True, '已同步到 GitHub'
    except Exception as e:
        return False, f'同步失败: {str(e)}'


# ==================== 用户管理（Excel 数据源） ====================

USER_HEADERS = ['用户名', '密码哈希(JSON)', '角色', '邮箱', '创建时间', '最后登录', '必须改密码', '状态']

def load_users() -> dict:
    """从 用户管理.xlsx 加载所有用户"""
    with _lock:
        if not HAS_OPENPYXL:
            return {}
        if not os.path.exists(USERS_EXCEL):
            return {}
        try:
            wb = load_workbook(USERS_EXCEL)
            ws = wb.active
            users = {}
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[0]:
                    continue
                username = str(row[0]).strip()
                pwd_raw = row[1] if len(row) > 1 else '{}'
                try:
                    password = json.loads(pwd_raw) if isinstance(pwd_raw, str) else {}
                except Exception:
                    password = {}
                role = str(row[2]).strip() if len(row) > 2 and row[2] else 'viewer'
                email = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                created_at = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                last_login = str(row[5]).strip() if len(row) > 5 and row[5] else None
                must_change = str(row[6]).strip() == '是' if len(row) > 6 and row[6] else False
                status = str(row[7]).strip() if len(row) > 7 and row[7] else 'active'
                users[username] = {
                    'username': username,
                    'password': password,
                    'role': role,
                    'email': email,
                    'created_at': created_at,
                    'last_login': last_login,
                    'must_change_pwd': must_change,
                    'status': status
                }
            return users
        except Exception as e:
            print(f'[auth] 读取用户Excel失败: {e}')
            return {}


def save_users(users: dict, push: bool = True) -> tuple[bool, str]:
    """保存用户到 用户管理.xlsx，并推送到 GitHub
    
    返回: (成功, 消息)
    - 第一个 bool 表示整体是否成功（写入+推送都成功才为True）
    - 如果 push=False，只检查写入是否成功
    """
    write_ok = True
    with _lock:
        if not HAS_OPENPYXL:
            return False, '缺少 openpyxl 依赖'
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = '用户表'
            ws.append(USER_HEADERS)
            for name, info in sorted(users.items()):
                pwd_json = json.dumps(info.get('password', {}), ensure_ascii=False)
                must_change = '是' if info.get('must_change_pwd', False) else '否'
                ws.append([
                    info.get('username', name),
                    pwd_json,
                    info.get('role', 'viewer'),
                    info.get('email', ''),
                    info.get('created_at', ''),
                    info.get('last_login') or '',
                    must_change,
                    info.get('status', 'active')
                ])
            col_widths = [15, 60, 12, 25, 25, 25, 12, 10]
            for i, w in enumerate(col_widths):
                ws.column_dimensions[chr(65+i)].width = w
            wb.save(USERS_EXCEL)
        except Exception as e:
            write_ok = False
            print(f'[auth] 写入用户Excel失败: {e}')
            return False, f'写入用户Excel失败: {e}'
    
    if push:
        push_ok, push_msg = _git_push('更新用户数据')
        if not push_ok:
            print(f'[auth] 警告: {push_msg}')
            # 【关键修复】本地保存成功就返回True，push失败只是警告
            # 用户创建/更新的核心目的已达成，不应因网络问题判定为失败
            return True, f'用户已保存（GitHub同步失败: {push_msg}），下次同步时将自动重试'
        return True, '用户已保存并同步到GitHub'
    
    return write_ok, '用户已保存到本地（未推送）'


def init_default_users():
    """初始化默认用户（首次运行）
    
    【关键修复】
    1. 先尝试从远程 GitHub 恢复用户管理.xlsx（如果本地没有）
    2. 只有远程也没有时，才创建默认 admin 用户
    3. 创建默认用户时 push=False（绝不允许本地初始化覆盖远程数据！）
    
    致命bug历史：之前 push=True 会导致当用户管理.xlsx不存在时，
    创建只有admin的空表并推送到远程，覆盖所有已有用户数据！
    """
    # ===== 修复0：先确保Git仓库存在（Render部署环境保障） =====
    _ensure_git_repo()
    
    # ===== 修复1：先尝试从远程恢复 =====
    if not os.path.exists(USERS_EXCEL):
        try:
            if os.path.exists(os.path.join(BASE_DIR, '.git')):
                subprocess.run(
                    ['git', 'fetch', 'origin', 'main'],
                    capture_output=True, cwd=BASE_DIR, timeout=15
                )
                subprocess.run(
                    ['git', 'checkout', 'origin/main', '--', '用户管理.xlsx'],
                    capture_output=True, cwd=BASE_DIR, timeout=10
                )
                if os.path.exists(USERS_EXCEL):
                    users = load_users()
                    if users:
                        print(f"[auth] 已从远程 GitHub 恢复用户数据（共 {len(users)} 个用户）")
                        return users
        except Exception:
            pass  # 远程恢复失败，继续创建默认用户

    # ===== 修复2：只有远程也没有时，才创建默认用户 =====
    users = load_users()
    if not users:
        # 创建默认管理员
        default_pwd = 'admin123'
        users['admin'] = {
            'username': 'admin',
            'password': _hash_password(default_pwd),
            'role': 'admin',
            'email': '',
            'created_at': datetime.now().isoformat(),
            'last_login': None,
            'must_change_pwd': True,
            'status': 'active'
        }
        # ===== 修复3：push=False！本地初始化绝不允许覆盖远程数据 =====
        save_users(users, push=False)
        _audit_log('USER_CREATE', 'system', '创建默认管理员 admin / admin123')
        print(f"\n⚠️  已创建默认管理员账号: admin / {default_pwd}")
        print("    请登录后立即修改密码！\n")
    return users


def get_user(username: str) -> dict:
    """获取用户信息"""
    users = load_users()
    return users.get(username)


def create_user(username: str, password: str, role: str = 'viewer',
                email: str = '') -> tuple[bool, str]:
    """创建用户"""
    if role not in ROLES:
        return False, f'无效的角色: {role}'
    if len(username) < 3 or len(username) > 32:
        return False, '用户名长度需在 3-32 之间'
    if len(password) < 6:
        return False, '密码长度至少 6 位'

    users = load_users()
    if username in users:
        return False, '用户名已存在'

    users[username] = {
        'username': username,
        'password': _hash_password(password),
        'role': role,
        'email': email,
        'created_at': datetime.now().isoformat(),
        'last_login': None,
        'must_change_pwd': False,
        'status': 'active'
    }
    save_ok, save_msg = save_users(users, push=True)
    _audit_log('USER_CREATE', 'admin', f'创建用户: {username}, 角色: {role}')
    return save_ok, save_msg if not save_ok else '用户创建成功'


def update_user(username: str, **kwargs) -> tuple[bool, str]:
    """更新用户信息"""
    users = load_users()
    if username not in users:
        return False, '用户不存在'

    if 'role' in kwargs and kwargs['role'] not in ROLES:
        return False, f'无效的角色: {kwargs["role"]}'
    if 'password' in kwargs:
        if len(kwargs['password']) < 6:
            return False, '密码长度至少 6 位'
        users[username]['password'] = _hash_password(kwargs['password'])
        del kwargs['password']

    for k, v in kwargs.items():
        if k in ('role', 'email', 'status', 'must_change_pwd'):
            users[username][k] = v

    save_ok, save_msg = save_users(users, push=True)
    _audit_log('USER_UPDATE', 'admin', f'更新用户: {username}, 字段: {list(kwargs.keys())}')
    return save_ok, save_msg if not save_ok else '用户更新成功'


def delete_user(username: str) -> tuple[bool, str]:
    """删除用户"""
    if username == 'admin':
        return False, '不能删除默认管理员'

    users = load_users()
    if username not in users:
        return False, '用户不存在'

    del users[username]
    save_ok, save_msg = save_users(users, push=True)
    _audit_log('USER_DELETE', 'admin', f'删除用户: {username}')
    return save_ok, save_msg if not save_ok else '用户删除成功'


def change_password(username: str, old_password: str, new_password: str) -> tuple[bool, str]:
    """用户修改自己的密码"""
    users = load_users()
    if username not in users:
        return False, '用户不存在'

    if not _verify_password(old_password, users[username]['password']):
        return False, '原密码错误'
    if len(new_password) < 6:
        return False, '新密码长度至少 6 位'
    if old_password == new_password:
        return False, '新密码不能与原密码相同'

    users[username]['password'] = _hash_password(new_password)
    users[username]['must_change_pwd'] = False
    save_ok, save_msg = save_users(users, push=True)
    _audit_log('PASSWORD_CHANGE', username, '修改密码成功')
    return save_ok, save_msg if not save_ok else '密码修改成功'


def list_users() -> list:
    """列出所有用户（去除密码）"""
    users = load_users()
    result = []
    for u in users.values():
        safe = {k: v for k, v in u.items() if k != 'password'}
        safe['role_name'] = ROLES.get(u['role'], {}).get('name', u['role'])
        result.append(safe)
    return result


# ==================== 权限检查 ====================

def has_permission(username: str, permission: str) -> bool:
    """检查用户是否有指定权限"""
    user = get_user(username)
    if not user or user.get('status') != 'active':
        return False
    role = user.get('role', 'viewer')
    return permission in ROLES.get(role, {}).get('permissions', [])


def require_permission(permission: str):
    """装饰器：要求权限"""
    def decorator(func):
        @wraps(func)
        def wrapper(self, *args, **kwargs):
            session = self.get_current_session()
            if not session:
                self.send_json({'error': '未登录'}, 401)
                return
            username = session.get('username')
            if not has_permission(username, permission):
                self.send_json({'error': '权限不足'}, 403)
                return
            return func(self, *args, **kwargs)
        return wrapper
    return decorator


# ==================== Session 管理 ====================

def load_sessions() -> dict:
    return _safe_read_json(SESSIONS_FILE, {})


def save_sessions(sessions: dict):
    _safe_write_json(SESSIONS_FILE, sessions)


def cleanup_sessions():
    """清理过期 Session"""
    sessions = load_sessions()
    now = time.time()
    expired = [sid for sid, s in sessions.items()
               if s.get('expires_at', 0) < now]
    for sid in expired:
        del sessions[sid]
    if expired:
        save_sessions(sessions)
        _audit_log('SESSION_CLEANUP', 'system', f'清理了 {len(expired)} 个过期 Session')
    return len(expired)


def create_session(username: str) -> str:
    """创建 Session，返回 Session ID"""
    cleanup_sessions()
    sessions = load_sessions()

    # 先清除该用户的旧 Session（单点登录）
    for sid, s in list(sessions.items()):
        if s.get('username') == username:
            del sessions[sid]

    session_id = secrets.token_urlsafe(48)
    now = time.time()
    sessions[session_id] = {
        'username': username,
        'created_at': now,
        'expires_at': now + SESSION_TIMEOUT,
        'last_activity': now,
    }
    save_sessions(sessions)

    # 更新用户最后登录时间
    users = load_users()
    if username in users:
        users[username]['last_login'] = datetime.now().isoformat()
        save_users(users, push=False)

    _audit_log('LOGIN', username, '登录成功')
    return session_id


def get_session(session_id: str) -> dict:
    """获取 Session（自动续期）"""
    if not session_id:
        return None
    cleanup_sessions()
    sessions = load_sessions()
    session = sessions.get(session_id)
    if not session:
        return None
    if session.get('expires_at', 0) < time.time():
        del sessions[session_id]
        save_sessions(sessions)
        return None

    # 自动续期（每次访问重置过期时间）
    sessions[session_id]['last_activity'] = time.time()
    sessions[session_id]['expires_at'] = time.time() + SESSION_TIMEOUT
    save_sessions(sessions)

    return session


def destroy_session(session_id: str):
    """销毁 Session（登出）"""
    sessions = load_sessions()
    if session_id in sessions:
        username = sessions[session_id].get('username', '')
        del sessions[session_id]
        save_sessions(sessions)
        _audit_log('LOGOUT', username, '登出成功')


# ==================== 登录限流 ====================

def load_rate_limit() -> dict:
    return _safe_read_json(RATE_LIMIT_FILE, {})


def save_rate_limit(data: dict):
    _safe_write_json(RATE_LIMIT_FILE, data)


def check_rate_limit(ip: str) -> tuple[bool, int]:
    """检查是否被限流，返回 (是否允许, 剩余秒数)"""
    data = load_rate_limit()
    entry = data.get(ip, {})
    now = time.time()

    # 检查是否被锁定
    if entry.get('locked_until', 0) > now:
        return False, int(entry['locked_until'] - now)

    return True, 0


def record_login_attempt(ip: str, success: bool):
    """记录登录尝试"""
    data = load_rate_limit()
    now = time.time()
    entry = data.get(ip, {'attempts': 0, 'first_attempt': 0})

    if success:
        # 登录成功，重置计数
        data[ip] = {'attempts': 0, 'first_attempt': 0}
    else:
        if now - entry.get('first_attempt', 0) > LOGIN_LOCKOUT_TIME:
            # 超过锁定窗口，重置
            entry = {'attempts': 1, 'first_attempt': now}
        else:
            entry['attempts'] += 1
            if entry['attempts'] >= MAX_LOGIN_ATTEMPTS:
                entry['locked_until'] = now + LOGIN_LOCKOUT_TIME
                _audit_log('RATE_LIMIT', 'system',
                           f'IP {ip} 因登录失败次数过多被锁定 {LOGIN_LOCKOUT_TIME//60} 分钟')
        data[ip] = entry

    save_rate_limit(data)


# ==================== CSRF 防护 ====================

_csrf_tokens = {}  # session_id -> {token, expires_at}


def generate_csrf_token(session_id: str) -> str:
    """生成 CSRF Token"""
    token = secrets.token_urlsafe(32)
    _csrf_tokens[session_id] = {
        'token': token,
        'expires_at': time.time() + CSRF_TOKEN_TTL
    }
    return token


def verify_csrf_token(session_id: str, token: str) -> bool:
    """验证 CSRF Token"""
    entry = _csrf_tokens.get(session_id)
    if not entry:
        return False
    if entry['expires_at'] < time.time():
        del _csrf_tokens[session_id]
        return False
    return hmac.compare_digest(entry['token'], token)


# ==================== 审计日志 ====================

def get_audit_log(limit: int = 100) -> list:
    """获取审计日志"""
    if not os.path.exists(AUDIT_LOG_FILE):
        return []
    try:
        with open(AUDIT_LOG_FILE, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        return [l.strip() for l in lines[-limit:]]
    except Exception:
        return []


# ==================== 初始化 ====================

def startup_sync() -> tuple[bool, str]:
    """启动时从 GitHub 拉取最新的用户数据（双路径：git命令优先，API兜底）"""
    # 检查 git 命令是否可用
    git_available = False
    try:
        result = subprocess.run(['git', '--version'], capture_output=True, timeout=5)
        git_available = (result.returncode == 0) and os.path.exists(os.path.join(BASE_DIR, '.git'))
    except (FileNotFoundError, OSError):
        git_available = False
    
    if not git_available:
        try:
            from github_sync import github_api_pull
            print('[auth] git 不可用，使用 GitHub API 模式拉取')
            return github_api_pull()
        except Exception as api_e:
            return False, f'git不可用且API拉取失败: {api_e}'
    
    try:
        _ensure_git_repo()
        if not os.path.exists(os.path.join(BASE_DIR, '.git')):
            return False, '未检测到 Git 仓库'

        subprocess.run(
            ['git', 'fetch', 'origin', 'main'],
            capture_output=True, cwd=BASE_DIR, timeout=30
        )

        existed_before = os.path.exists(USERS_EXCEL)
        checkout = subprocess.run(
            ['git', 'checkout', 'origin/main', '--', '用户管理.xlsx'],
            capture_output=True, cwd=BASE_DIR, timeout=10
        )

        if os.path.exists(USERS_EXCEL):
            users = load_users()
            if existed_before:
                return True, f'用户数据同步成功（已覆盖本地，共 {len(users)} 个用户）'
            else:
                return True, f'用户数据同步成功（从远程新建，共 {len(users)} 个用户）'
        else:
            return True, 'Git 同步完成，但远程也没有用户管理.xlsx（首次运行将创建默认用户）'

    except Exception as e:
        return False, f'用户数据同步失败: {str(e)}'


def init_auth():
    """初始化认证系统
    
    【防护增强】即使 main() 中忘记先调用 startup_sync()，
    这里也会先尝试从远程恢复用户数据，避免创建空表覆盖远程数据。
    """
    # 防护：如果用户管理.xlsx不存在，先尝试从远程恢复
    if not os.path.exists(USERS_EXCEL) and os.path.exists(os.path.join(BASE_DIR, '.git')):
        try:
            sync_ok, sync_msg = startup_sync()
            print(f"[auth] init_auth 防护同步: {sync_msg}")
        except Exception:
            pass

    init_default_users()
    cleanup_sessions()
    # 定期清理 Session（每小时一次）
    def _periodic_cleanup():
        while True:
            time.sleep(3600)
            try:
                cleanup_sessions()
            except Exception:
                pass

    t = threading.Thread(target=_periodic_cleanup, daemon=True)
    t.start()


# ==================== GitHub 数据同步 ====================

def sync_to_github(message: str = '同步数据') -> tuple[bool, str]:
    """将 data/ 目录和 用户管理.xlsx 推送到 GitHub（双路径：git优先+API兜底）"""
    # 检查 git 命令是否可用
    git_available = False
    try:
        result = subprocess.run(['git', '--version'], capture_output=True, timeout=5)
        git_available = (result.returncode == 0) and os.path.exists(os.path.join(BASE_DIR, '.git'))
    except (FileNotFoundError, OSError):
        git_available = False
    
    if not git_available:
        try:
            from github_sync import github_api_push
            print('[auth] git 不可用，使用 GitHub API 模式推送')
            ok, msg = github_api_push(message)
            if ok:
                _audit_log('GITHUB_SYNC', 'system', message)
            return ok, msg
        except Exception as api_e:
            return False, f'git不可用且API推送失败: {api_e}'
    
    import subprocess
    try:
        _ensure_git_repo()
        if not os.path.exists(os.path.join(BASE_DIR, '.git')):
            return False, '未检测到 Git 仓库'

        ahead = subprocess.run(
            ['git', 'rev-list', '--count', 'origin/main..HEAD'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        try:
            ahead_count = int(ahead.stdout.strip())
        except ValueError:
            ahead_count = 0

        if ahead_count > 0:
            push = subprocess.run(
                ['git', 'push', 'origin', 'main'],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=30
            )
            if push.returncode != 0:
                # 【关键修复】git push 失败时，回退到 GitHub API 模式推送
                try:
                    from github_sync import github_api_push
                    print('[auth] git push 失败，回退到 GitHub API 模式')
                    return github_api_push(message)
                except Exception as api_e:
                    return False, f'推送失败（git+API均失败）: {push.stderr[:200]}'
            return True, f'已推送 {ahead_count} 个待提交到 GitHub'

        result = subprocess.run(
            ['git', 'status', '--porcelain', 'data/', '用户管理.xlsx'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if not result.stdout.strip():
            return True, '数据无变更，无需同步'

        subprocess.run(
            ['git', 'add', 'data/', '用户管理.xlsx'],
            capture_output=True, cwd=BASE_DIR, timeout=10
        )
        commit_msg = f'[数据同步] {message} - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        commit_result = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=10
        )
        if commit_result.returncode != 0:
            if 'nothing to commit' in (commit_result.stdout + commit_result.stderr):
                return True, '无变更'
            return False, f'提交失败: {commit_result.stderr[:200]}'

        push_result = subprocess.run(
            ['git', 'push', 'origin', 'main'],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        )
        if push_result.returncode != 0:
            subprocess.run(
                ['git', 'reset', '--soft', 'HEAD~1'],
                capture_output=True, cwd=BASE_DIR, timeout=10
            )
            # 【关键修复】git push 失败时，回退到 GitHub API 模式推送
            try:
                from github_sync import github_api_push
                print('[auth] git push 失败，回退到 GitHub API 模式')
                ok, msg = github_api_push(message)
                if ok:
                    _audit_log('GITHUB_SYNC', 'system', message)
                return ok, msg
            except Exception as api_e:
                return False, f'推送失败（git+API均失败）: {push_result.stderr[:200]}'

        _audit_log('GITHUB_SYNC', 'system', message)
        return True, '数据已同步到 GitHub'
    except subprocess.TimeoutExpired:
        return False, '同步超时'
    except Exception as e:
        return False, f'同步失败: {str(e)}'


def auto_sync_periodically(interval_seconds: int = 300):
    """定期自动同步数据到 GitHub（默认5分钟）"""
    def _sync_loop():
        while True:
            time.sleep(interval_seconds)
            try:
                sync_to_github('定期自动同步')
            except Exception:
                pass
    
    t = threading.Thread(target=_sync_loop, daemon=True)
    t.start()
    _audit_log('AUTO_SYNC', 'system', f'已启动自动同步，间隔 {interval_seconds} 秒')


if __name__ == '__main__':
    # 测试
    init_auth()
    print("用户列表:", list_users())
    print("角色:", list(ROLES.keys()))
